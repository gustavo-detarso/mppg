#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.request
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path


NS = {
    "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}

DEFAULT_OUT = Path("app_bundle/projetos/prisma_fluxo_pmf/output_pesquisa/relatorio_prisma_prisma_fluxo_pmf")
PREFIX = "relatorio_prisma_prisma_fluxo_pmf"

DEFAULT_TEMA = (
    "Seminário ATESTMED/Perícia Médica Federal sobre redesenho do fluxo de análise "
    "dos benefícios por incapacidade, com foco em avaliação médico-pericial, análise "
    "documental, teleperícia, capacidade/fila, qualidade, integridade e equidade."
)


def log(kind: str, msg: str) -> None:
    print(f"[{kind}] {msg}")


def load_dotenv(path: Path = Path(".env")) -> None:
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


def norm_text(s: object) -> str:
    s = str(s or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def norm_key(s: object) -> str:
    return norm_text(s).lower()


def title_key(row: dict[str, str]) -> str:
    return norm_key(row.get("titulo") or row.get("title") or "")


def doi_key(row: dict[str, str]) -> str:
    doi = norm_key(row.get("doi") or row.get("DOI") or "")
    return doi.replace("https://doi.org/", "").replace("http://dx.doi.org/", "")


def url_key(row: dict[str, str]) -> str:
    return norm_key(row.get("url") or row.get("URL") or "")


def col_idx(ref: str) -> int:
    m = re.match(r"([A-Z]+)", ref or "")
    if not m:
        return 0
    n = 0
    for ch in m.group(1):
        n = n * 26 + ord(ch) - 64
    return n - 1


def read_xlsx(path: Path, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, str]]]:
    with zipfile.ZipFile(path) as z:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in z.namelist():
            root = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for si in root.findall("a:si", NS):
                shared.append("".join(t.text or "" for t in si.findall(".//a:t", NS)))

        workbook = ET.fromstring(z.read("xl/workbook.xml"))
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rel_map = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels.findall("rel:Relationship", NS)
        }

        def resolve_target(target: str) -> str:
            target = (target or "").replace("\\\\", "/").lstrip("/")
            if target.startswith("xl/"):
                return target
            if target.startswith("../"):
                target = target[3:]
            return "xl/" + target

        target_xml = None
        available = []
        for sheet in workbook.findall("a:sheets/a:sheet", NS):
            name = sheet.attrib.get("name", "")
            rid = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            target = resolve_target(rel_map.get(rid, ""))
            available.append(name)
            if sheet_name is None or name == sheet_name:
                target_xml = target
                if sheet_name is not None:
                    break

        if target_xml is None:
            raise SystemExit(f"Aba '{sheet_name}' não encontrada. Abas disponíveis: {available}")

        root = ET.fromstring(z.read(target_xml))
        rows: list[list[str]] = []
        for r in root.findall("a:sheetData/a:row", NS):
            vals: list[str] = []
            for c in r.findall("a:c", NS):
                idx = col_idx(c.attrib.get("r", "A1"))
                while len(vals) <= idx:
                    vals.append("")
                t = c.attrib.get("t")
                if t == "inlineStr":
                    val = "".join(x.text or "" for x in c.findall(".//a:t", NS))
                else:
                    v = c.find("a:v", NS)
                    val = "" if v is None else (v.text or "")
                    if t == "s" and val:
                        val = shared[int(val)]
                vals[idx] = val
            rows.append(vals)

    if not rows:
        return [], []

    headers = [str(h or "").strip() for h in rows[0]]
    data = []
    for row in rows[1:]:
        row = row + [""] * (len(headers) - len(row))
        if not any(str(x).strip() for x in row):
            continue
        data.append({headers[i]: str(row[i] if i < len(row) else "") for i in range(len(headers))})
    return headers, data


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    for enc in ("utf-8-sig", "utf-8", "latin1"):
        try:
            with path.open("r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                return list(reader.fieldnames or []), [dict(r) for r in reader]
        except UnicodeDecodeError:
            continue
    raise SystemExit(f"Não consegui ler CSV: {path}")


def find_input(out_dir: Path) -> Path:
    candidates = [
        out_dir / f"{PREFIX}.triagem_titulo_resumo.xlsx",
        out_dir / f"{PREFIX}.triagem_titulo_resumo.csv",
        Path(f"{PREFIX}.triagem_titulo_resumo.xlsx"),
        Path(f"{PREFIX}.triagem_titulo_resumo.csv"),
    ]
    for c in candidates:
        if c.exists():
            return c
    raise SystemExit("Não encontrei triagem_titulo_resumo.xlsx/csv. Informe com --input.")


def load_triage(path: Path) -> tuple[list[str], list[dict[str, str]], str]:
    if path.suffix.lower() == ".xlsx":
        best_h, best_r, best_sheet = [], [], "primeira aba"
        for sheet in ("Triagem e matriz", "Triagem completa", None):
            try:
                h, r = read_xlsx(path, sheet)
            except Exception:
                continue
            if len(r) > len(best_r):
                best_h, best_r, best_sheet = h, r, sheet or "primeira aba"
            if len(r) >= 50:
                return h, r, f"{path} :: aba {sheet or 'primeira aba'}"
        return best_h, best_r, f"{path} :: aba {best_sheet}"
    h, r = read_csv(path)
    return h, r, str(path)


def clean(s: object, max_len: int | None = None) -> str:
    txt = re.sub(r"\s+", " ", str(s or "")).strip()
    if max_len and len(txt) > max_len:
        return txt[: max_len - 1].rstrip() + "…"
    return txt


def score_row(row: dict[str, str]) -> int:
    text = norm_key(" ".join([
        row.get("titulo", ""),
        row.get("resumo", ""),
        row.get("periodico", ""),
        row.get("palavras_chave", ""),
        row.get("bloco_tematico_ia", ""),
    ]))
    score = 0
    positives = {
        35: [
            "incapacity benefit", "disability benefit", "disability benefits",
            "sickness benefit", "social security disability", "beneficio por incapacidade",
            "benefício por incapacidade", "bpc", "inss", "national social security institute",
        ],
        30: [
            "work disability", "work capacity", "medical certification", "medical certificate",
            "medical evidence", "medical assessment", "medical evaluation", "certification of disability",
            "assessment of disability", "capacidade laboral", "incapacidade laboral", "pericia",
            "perícia", "elegibility", "eligibility",
        ],
        22: [
            "telehealth", "telemedicine", "remote assessment", "digital certification",
            "digitization", "artificial intelligence", "ai", "atestmed", "analise documental",
            "análise documental",
        ],
        18: [
            "waiting time", "wait time", "application costs", "screened out", "equity",
            "quality assurance", "audit", "regional trends", "access", "capacity",
        ],
    }
    negatives = {
        -35: ["dental", "cardiac rehabilitation", "haemophilia", "incontinence"],
        -28: ["mental health apps", "exercise training", "home rehabilitation", "shoulder"],
        -18: ["telemedicine intervention" if "disability" not in text and "certificate" not in text else "__nope__"],
    }
    for weight, terms in positives.items():
        for term in terms:
            if term in text:
                score += weight
                break
    for weight, terms in negatives.items():
        for term in terms:
            if term in text:
                score += weight
                break

    rec = norm_key(row.get("recomendacao_ia", ""))
    if "prioridade_alta" in rec:
        score += 25
    elif "revisar_humano" in rec:
        score += 10
    elif "incerto" in rec:
        score -= 8

    try:
        score += int(float(row.get("escore_aderencia_ia") or 0) / 10)
    except Exception:
        pass

    return max(0, min(100, score))


def infer_axis(row: dict[str, str]) -> str:
    text = norm_key((row.get("titulo", "") + " " + row.get("resumo", "")))
    if any(t in text for t in ["telehealth", "telemedicine", "remote assessment", "digitization", "artificial intelligence"]):
        return "Teleperícia, digitalização e avaliação remota"
    if any(t in text for t in ["wait", "waiting", "screened out", "application costs", "equity", "regional trends", "access"]):
        return "Capacidade, filas, acesso e equidade"
    if any(t in text for t in ["certificate", "certification", "medical evidence", "document", "documental"]):
        return "Análise documental, certificação e elegibilidade"
    if any(t in text for t in ["incapacity benefit", "sickness benefit", "work disability", "bpc", "inss", "social security"]):
        return "Benefícios por incapacidade e avaliação médico-pericial"
    return row.get("bloco_tematico_ia") or "Benefícios por incapacidade e avaliação médico-pericial"


def infer_design(row: dict[str, str]) -> str:
    text = norm_key(row.get("titulo", "") + " " + row.get("resumo", ""))
    if "scoping review" in text:
        return "Revisão de escopo"
    if "systematic review" in text:
        return "Revisão sistemática"
    if any(t in text for t in ["qualitative", "interviews", "ethnographic", "text analysis"]):
        return "Estudo qualitativo/análise documental"
    if any(t in text for t in ["pilot", "evaluation of a pilot"]):
        return "Avaliação de estudo-piloto"
    if any(t in text for t in ["retrospective", "cross-sectional"]):
        return "Estudo observacional"
    if any(t in text for t in ["instrument", "scale", "measure"]):
        return "Desenvolvimento/validação de instrumento"
    if any(t in text for t in ["proposal", "propose", "protocol"]):
        return "Proposta de protocolo/modelo"
    if "review" in text:
        return "Revisão narrativa/documental"
    return "Estudo aplicado/documental"


def infer_country(row: dict[str, str]) -> str:
    text = norm_key(row.get("titulo", "") + " " + row.get("resumo", "") + " " + row.get("periodico", ""))
    checks = [
        ("Brasil/INSS", ["brazil", "brazilian", "brasil", "inss", "parana", "bpc"]),
        ("Itália/INPS", ["italy", "italian", "inps"]),
        ("Coreia do Sul", ["korea", "korean"]),
        ("Reino Unido", ["incapacity benefit", "britain", "uk", "great britain", "england"]),
        ("Países Baixos", ["netherlands", "dutch"]),
        ("Estados Unidos", ["social security disability insurance", "ssa", "supplemental security income", "united states"]),
        ("Portugal", ["portugal", "aveiro"]),
        ("Internacional/comparativo", ["international", "oecd", "countries", "worldwide"]),
    ]
    for label, terms in checks:
        if any(t in text for t in terms):
            return label
    return "Internacional ou não especificado"


def abnt_ref(row: dict[str, str]) -> str:
    autores = clean(row.get("autores", ""))
    titulo = clean(row.get("titulo", ""))
    periodico = clean(row.get("periodico", ""))
    ano = clean(row.get("ano", ""))
    doi = clean(row.get("doi", ""))
    url = clean(row.get("url", ""))
    parts = []
    if autores:
        parts.append(autores.upper())
    if titulo:
        parts.append(f"{titulo}.")
    if periodico:
        parts.append(f"{periodico},")
    if ano:
        parts.append(f"{ano}.")
    if doi:
        parts.append(f"DOI: {doi}.")
    elif url:
        parts.append(f"Disponível em: {url}.")
    return " ".join(parts)


def heuristic_decision(row: dict[str, str], max_score: int | None = None) -> dict[str, str]:
    score = score_row(row)
    axis = infer_axis(row)
    include = score >= 58
    return {
        "escore_curadoria": str(score),
        "prioridade_seminario": "APOIO" if include else "EXCLUÍDA",
        "eixo_seminario": axis,
        "decisao_titulo_resumo": "INCLUIR" if include else "EXCLUIR",
        "motivo_exclusao_titulo_resumo": "" if include else "Baixa aderência ao recorte ATESTMED/PMF ou tema periférico.",
        "texto_completo_local": row.get("url_pdf_aberto") or "PENDENTE_LOCALIZAR" if include else "",
        "decisao_texto_completo": "INCLUIR" if include else "NAO_APLICAVEL",
        "motivo_exclusao_texto_completo": "" if include else "Excluído na triagem por título/resumo.",
        "incluir_final": "SIM" if include else "NAO",
        "observacoes": "Decisão por heurística local; revisar manualmente se necessário.",
        "pais_contexto": infer_country(row),
        "objetivo_estudo": f"Analisar contribuição para o eixo: {axis}.",
        "desenho_metodo": infer_design(row),
        "amostra_base": "Não detalhado automaticamente; conferir texto completo.",
        "achados_principais": clean(row.get("resumo", ""), 420),
        "limitacoes": "Limitações não detalhadas automaticamente; confirmar no texto completo.",
        "contribuicao_pergunta": "Contribui para discutir fluxo de análise de benefícios por incapacidade, documentação, capacidade, qualidade ou equidade.",
        "como_usar_no_seminario": f"Usar no bloco: {axis}.",
        "referencia_abnt_rascunho": abnt_ref(row),
    }


def chunked(items: list[dict[str, str]], n: int):
    for i in range(0, len(items), n):
        yield items[i : i + n]


def extract_json(text: str):
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?", "", text).strip()
        text = re.sub(r"```$", "", text).strip()
    try:
        return json.loads(text)
    except Exception:
        pass
    m = re.search(r"(\[.*\]|\{.*\})", text, re.S)
    if not m:
        raise ValueError("Resposta da IA não contém JSON.")
    return json.loads(m.group(1))


def call_openai_batch(batch: list[dict[str, str]], tema: str, model: str, api_url: str, api_key: str) -> list[dict[str, str]]:
    payload_items = []
    for r in batch:
        payload_items.append({
            "row_id": r["_row_id"],
            "titulo": clean(r.get("titulo", ""), 260),
            "autores": clean(r.get("autores", ""), 160),
            "ano": clean(r.get("ano", ""), 20),
            "periodico": clean(r.get("periodico", ""), 120),
            "doi": clean(r.get("doi", ""), 100),
            "url": clean(r.get("url", ""), 160),
            "resumo": clean(r.get("resumo", ""), 1200),
            "recomendacao_previa": r.get("recomendacao_ia", ""),
            "escore_previo": r.get("escore_aderencia_ia", ""),
        })

    system = (
        "Você é um pesquisador sênior em revisão PRISMA e políticas de benefícios por incapacidade. "
        "Classifique referências para um seminário, com postura criteriosa: inclua apenas estudos diretamente úteis. "
        "Responda somente JSON válido."
    )
    user = {
        "tarefa": "Curadoria de referências para matriz PRISMA e seminário.",
        "tema": tema,
        "criterios_inclusao": [
            "benefício por incapacidade, sickness/incapacity/disability benefits",
            "avaliação médico-pericial, work disability assessment, work capacity",
            "análise documental, medical certification, medical evidence, atestados",
            "teleperícia/telemedicine/remote assessment vinculada a certificação, incapacidade ou deficiência",
            "capacidade/fila/acesso/equidade em avaliação de benefícios por incapacidade",
            "qualidade, auditoria, integridade ou padronização decisória",
        ],
        "criterios_exclusao": [
            "telemedicina assistencial genérica sem benefício/perícia/certificação",
            "reabilitação clínica sem interface com elegibilidade ou benefício",
            "deficiência em geral sem certificação, avaliação administrativa ou decisão de direito",
            "apps, tecnologia ou saúde digital sem vínculo com incapacidade/certificação",
            "tema periférico ou redundante",
        ],
        "saida_obrigatoria": {
            "tipo": "array",
            "campos_por_item": [
                "row_id", "decisao_titulo_resumo", "motivo_exclusao_titulo_resumo",
                "decisao_texto_completo", "motivo_exclusao_texto_completo",
                "incluir_final", "prioridade_seminario", "eixo_seminario",
                "pais_contexto", "objetivo_estudo", "desenho_metodo", "amostra_base",
                "achados_principais", "limitacoes", "contribuicao_pergunta",
                "como_usar_no_seminario"
            ],
            "valores_validos": {
                "decisao_titulo_resumo": ["INCLUIR", "EXCLUIR", "REVISAR"],
                "decisao_texto_completo": ["INCLUIR", "EXCLUIR", "PENDENTE", "NAO_APLICAVEL"],
                "incluir_final": ["SIM", "NAO"],
                "prioridade_seminario": ["NÚCLEO", "APOIO", "EXCLUÍDA"]
            }
        },
        "referencias": payload_items,
    }

    data = json.dumps({
        "model": model,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": json.dumps(user, ensure_ascii=False)},
        ],
        "temperature": 0.1,
    }, ensure_ascii=False).encode("utf-8")

    req = urllib.request.Request(
        api_url,
        data=data,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=180) as resp:
        raw = resp.read().decode("utf-8", errors="replace")
    obj = json.loads(raw)
    content = obj["choices"][0]["message"]["content"]
    parsed = extract_json(content)
    if isinstance(parsed, dict) and "items" in parsed:
        parsed = parsed["items"]
    if not isinstance(parsed, list):
        raise ValueError("JSON da IA não é lista.")
    return parsed


def apply_ai(rows: list[dict[str, str]], args: argparse.Namespace) -> tuple[dict[str, dict[str, str]], list[dict[str, str]]]:
    load_dotenv()
    api_key = os.environ.get("OPENAI_API_KEY", "")
    model = args.model or os.environ.get("OPENAI_MODEL") or "gpt-4.1-mini"
    api_url = args.api_url or os.environ.get("OPENAI_API_BASE") or "https://api.openai.com/v1/chat/completions"

    log_entries = []
    decisions: dict[str, dict[str, str]] = {}

    if not args.usar_ia:
        log("INFO", "IA desativada. Usando curadoria heurística local.")
        return decisions, log_entries

    if not api_key:
        log("WARN", "OPENAI_API_KEY ausente. Usando curadoria heurística local.")
        return decisions, log_entries

    ranked = sorted(rows, key=lambda r: int(r["_score"]), reverse=True)
    candidates = ranked[: args.top_n_candidatos]
    log("INFO", f"IA analisará {len(candidates)} candidato(s), em lotes de {args.batch_size}.")

    for batch_no, batch in enumerate(chunked(candidates, args.batch_size), start=1):
        for attempt in range(1, args.max_retries + 1):
            try:
                result = call_openai_batch(batch, args.tema, model, api_url, api_key)
                for item in result:
                    row_id = str(item.get("row_id", "")).strip()
                    if row_id:
                        decisions[row_id] = {k: clean(v) for k, v in item.items()}
                log_entries.append({"batch": batch_no, "attempt": attempt, "status": "ok", "items": len(result)})
                log("OK", f"Lote IA {batch_no} concluído: {len(result)} item(ns).")
                break
            except Exception as exc:
                log_entries.append({"batch": batch_no, "attempt": attempt, "status": "erro", "erro": str(exc)})
                wait = min(60, 3 * attempt)
                log("WARN", f"Erro no lote IA {batch_no}, tentativa {attempt}: {exc}. Aguardando {wait}s.")
                time.sleep(wait)
        else:
            log("WARN", f"Lote IA {batch_no} falhou. Esses itens usarão heurística local.")

    return decisions, log_entries


def ensure_header(headers: list[str], h: str) -> None:
    if h not in headers:
        headers.append(h)


def finalize_rows(headers: list[str], rows: list[dict[str, str]], ai_decisions: dict[str, dict[str, str]], max_incluir: int) -> tuple[list[str], list[dict[str, str]]]:
    curated_fields = [
        "escore_curadoria", "ordem_curadoria", "prioridade_seminario", "eixo_seminario",
        "decisao_titulo_resumo", "motivo_exclusao_titulo_resumo",
        "texto_completo_local", "decisao_texto_completo", "motivo_exclusao_texto_completo",
        "incluir_final", "observacoes", "pais_contexto", "objetivo_estudo",
        "desenho_metodo", "amostra_base", "achados_principais", "limitacoes",
        "contribuicao_pergunta", "como_usar_no_seminario", "referencia_abnt_rascunho",
    ]
    for h in curated_fields:
        ensure_header(headers, h)

    out = []
    for r in rows:
        row = dict(r)
        h = heuristic_decision(row)
        row.update(h)
        row_id = str(row["_row_id"])
        if row_id in ai_decisions:
            ai = ai_decisions[row_id]
            for field in curated_fields:
                if field in ai and ai[field] != "":
                    row[field] = ai[field]
            row["observacoes"] = "Decisão gerada por IA; revisar antes de submissão formal."
            row["referencia_abnt_rascunho"] = abnt_ref(row)
        out.append(row)

    def include_rank(r: dict[str, str]):
        prio = r.get("prioridade_seminario", "")
        prio_val = 0 if prio == "NÚCLEO" else 1 if prio == "APOIO" else 2
        return (prio_val, -int(float(r.get("escore_curadoria") or 0)), clean(r.get("titulo", "")))

    included = [r for r in out if str(r.get("incluir_final", "")).strip().upper() == "SIM" or str(r.get("decisao_titulo_resumo", "")).strip().upper() == "INCLUIR"]
    included = sorted(included, key=include_rank)
    keep_ids = {r["_row_id"] for r in included[:max_incluir]}

    order = 1
    for r in sorted(out, key=lambda x: int(x["_row_id"])):
        if r["_row_id"] in keep_ids:
            r["incluir_final"] = "SIM"
            r["decisao_titulo_resumo"] = "INCLUIR"
            if r.get("decisao_texto_completo") in ("", "NAO_APLICAVEL"):
                r["decisao_texto_completo"] = "INCLUIR"
            if r.get("prioridade_seminario") == "EXCLUÍDA":
                r["prioridade_seminario"] = "APOIO"
            r["ordem_curadoria"] = str(order)
            order += 1
        else:
            r["incluir_final"] = "NAO"
            if r.get("decisao_titulo_resumo") != "EXCLUIR":
                r["motivo_exclusao_titulo_resumo"] = "Excluído por limite/curadoria final do seminário ou menor aderência relativa."
            r["decisao_titulo_resumo"] = "EXCLUIR"
            r["decisao_texto_completo"] = "NAO_APLICAVEL"
            r["motivo_exclusao_texto_completo"] = "Excluído na triagem por título/resumo."
            r["prioridade_seminario"] = "EXCLUÍDA"
            r["ordem_curadoria"] = ""

    return headers, out


def write_csv(path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    clean_headers = [h for h in headers if not h.startswith("_")]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=clean_headers, extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            writer.writerow({h: r.get(h, "") for h in clean_headers})


def write_xlsx_if_possible(path: Path, headers: list[str], rows: list[dict[str, str]], tema: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.chart import BarChart, Reference
    except Exception as exc:
        log("WARN", f"openpyxl não disponível; XLSX não será gerado. Motivo: {exc}")
        return

    clean_headers = [h for h in headers if not h.startswith("_")]
    included = [r for r in rows if str(r.get("incluir_final", "")).strip().upper() == "SIM"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Curadoria IA de referências PRISMA"])
    ws.append(["Tema", tema])
    ws.append([])
    ws.append(["Indicador", "Valor"])
    ws.append(["Registros analisados", len(rows)])
    ws.append(["Referências incluídas", len(included)])
    ws.append(["Referências excluídas", len(rows) - len(included)])
    ws.append(["Base de decisão", "Título/resumo, metadados, heurística e IA quando ativada"])
    ws.append([])
    ws.append(["Eixo temático", "Incluídas"])
    counts = {}
    for r in included:
        k = r.get("eixo_seminario") or "Sem eixo"
        counts[k] = counts.get(k, 0) + 1
    for k, v in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])):
        ws.append([k, v])

    def style_header(row):
        fill = PatternFill("solid", fgColor="1F4E79")
        for cell in row:
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    style_header(ws[1])
    style_header(ws[4])
    style_header(ws[10])

    if counts:
        chart = BarChart()
        chart.title = "Referências incluídas por eixo"
        data = Reference(ws, min_col=2, min_row=10, max_row=10+len(counts))
        cats = Reference(ws, min_col=1, min_row=11, max_row=10+len(counts))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "D10")

    inc_headers = [
        "ordem_curadoria", "prioridade_seminario", "eixo_seminario", "titulo", "autores", "ano",
        "periodico", "doi", "url", "url_pdf_aberto", "decisao_titulo_resumo", "decisao_texto_completo",
        "incluir_final", "pais_contexto", "objetivo_estudo", "desenho_metodo", "amostra_base",
        "achados_principais", "limitacoes", "contribuicao_pergunta", "como_usar_no_seminario",
        "referencia_abnt_rascunho"
    ]
    ws_inc = wb.create_sheet("Referências incluídas")
    ws_inc.append(inc_headers)
    for r in sorted(included, key=lambda x: int(x.get("ordem_curadoria") or 999999)):
        ws_inc.append([r.get(h, "") for h in inc_headers])

    ws_tri = wb.create_sheet("Triagem completa")
    ws_tri.append(clean_headers)
    for r in rows:
        ws_tri.append([r.get(h, "") for h in clean_headers])

    ws_abnt = wb.create_sheet("Referências ABNT")
    ws_abnt.append(["Ordem", "Referência ABNT preliminar", "DOI", "URL"])
    for r in sorted(included, key=lambda x: int(x.get("ordem_curadoria") or 999999)):
        ws_abnt.append([r.get("ordem_curadoria", ""), r.get("referencia_abnt_rascunho", ""), r.get("doi", ""), r.get("url", "")])

    ws_crit = wb.create_sheet("Critérios")
    criteria = [
        ["Campo", "Conteúdo"],
        ["Critério geral", "Incluir referências diretamente úteis ao seminário ATESTMED/PMF."],
        ["Inclusão", "Benefício por incapacidade, avaliação médico-pericial, análise documental, teleperícia, filas/capacidade, qualidade, integridade e equidade."],
        ["Exclusão", "Telemedicina genérica, reabilitação clínica sem vínculo pericial, deficiência geral sem certificação/elegibilidade, tema periférico ou redundante."],
        ["Observação", "Antes de submissão formal, conferir texto completo dos incluídos."],
    ]
    for row in criteria:
        ws_crit.append(row)

    for sheet in wb.worksheets:
        if sheet.max_row >= 1:
            style_header(sheet[1])
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(1, min(sheet.max_column, 25) + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in sheet[letter][: min(sheet.max_row, 60)]:
                max_len = max(max_len, len(str(cell.value or "")))
            sheet.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)

    if ws_inc.max_row > 1:
        tab = Table(displayName="TabelaReferenciasIncluidas", ref=f"A1:{get_column_letter(ws_inc.max_column)}{ws_inc.max_row}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        ws_inc.add_table(tab)
    if ws_tri.max_row > 1:
        tab = Table(displayName="TabelaTriagemCompleta", ref=f"A1:{get_column_letter(ws_tri.max_column)}{ws_tri.max_row}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        ws_tri.add_table(tab)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_summary(path: Path, rows: list[dict[str, str]], input_src: str, tema: str) -> None:
    included = [r for r in rows if str(r.get("incluir_final", "")).strip().upper() == "SIM"]
    axes = {}
    for r in included:
        k = r.get("eixo_seminario", "") or "Sem eixo"
        axes[k] = axes.get(k, 0) + 1
    lines = [
        "Resumo da curadoria IA PRISMA",
        "=" * 40,
        f"Entrada: {input_src}",
        f"Tema: {tema}",
        f"Total de registros: {len(rows)}",
        f"Incluídos finais: {len(included)}",
        f"Excluídos finais: {len(rows) - len(included)}",
        "",
        "Incluídos por eixo:",
    ]
    for k, v in sorted(axes.items(), key=lambda kv: (-kv[1], kv[0])):
        lines.append(f"- {k}: {v}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default="app_bundle/projetos/prisma_fluxo_pmf/prisma_fluxo_pmf.toml")
    parser.add_argument("--input", default="")
    parser.add_argument("--out-dir", default=str(DEFAULT_OUT))
    parser.add_argument("--tema", default=DEFAULT_TEMA)
    parser.add_argument("--usar-ia", action="store_true")
    parser.add_argument("--top-n-candidatos", type=int, default=90)
    parser.add_argument("--max-incluir", type=int, default=27)
    parser.add_argument("--batch-size", type=int, default=6)
    parser.add_argument("--max-retries", type=int, default=3)
    parser.add_argument("--model", default="")
    parser.add_argument("--api-url", default="")
    args = parser.parse_args()

    load_dotenv()
    out_dir = Path(args.out_dir)
    input_path = Path(args.input) if args.input else find_input(out_dir)
    headers, rows, src = load_triage(input_path)
    if not rows:
        log("ERRO", f"Nenhum registro lido de {input_path}.")
        return 1

    for i, r in enumerate(rows, start=1):
        r["_row_id"] = str(i)
        r["_score"] = str(score_row(r))

    ai_decisions, ai_log = apply_ai(rows, args)
    headers, final_rows = finalize_rows(headers, rows, ai_decisions, args.max_incluir)

    out_dir.mkdir(parents=True, exist_ok=True)
    csv_triagem = out_dir / f"{PREFIX}.triagem_humana.csv"
    csv_incluidas = out_dir / f"{PREFIX}.referencias_incluidas_seminario.csv"
    xlsx_matriz = out_dir / f"{PREFIX}.curadoria_ia_referencias.xlsx"
    summary = out_dir / f"{PREFIX}.curadoria_ia_resumo.txt"
    log_json = out_dir / f"{PREFIX}.curadoria_ia_log.json"

    write_csv(csv_triagem, headers, final_rows)
    included = [r for r in final_rows if str(r.get("incluir_final", "")).strip().upper() == "SIM"]
    write_csv(csv_incluidas, headers, included)
    write_xlsx_if_possible(xlsx_matriz, headers, final_rows, args.tema)
    write_summary(summary, final_rows, src, args.tema)
    log_json.write_text(json.dumps({
        "created_at": dt.datetime.now().isoformat(),
        "input": src,
        "usar_ia": bool(args.usar_ia),
        "model": args.model or os.environ.get("OPENAI_MODEL") or "gpt-4.1-mini",
        "top_n_candidatos": args.top_n_candidatos,
        "max_incluir": args.max_incluir,
        "registros": len(final_rows),
        "incluidos": len(included),
        "ai_log": ai_log,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    log("OK", f"Triagem humana para pipeline: {csv_triagem}")
    log("OK", f"Referências incluídas: {csv_incluidas}")
    log("OK", f"Matriz XLSX: {xlsx_matriz}")
    log("OK", f"Resumo: {summary}")
    log("OK", f"Log IA: {log_json}")
    log("INFO", f"Registros: {len(final_rows)} | Incluídos: {len(included)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

