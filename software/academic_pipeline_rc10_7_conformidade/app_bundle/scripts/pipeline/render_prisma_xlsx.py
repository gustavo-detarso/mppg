#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

from pathlib import Path
# Compatibilidade temporária entre pacote oficial e execução direta.
if __package__:
    from .prisma_model import PrismaReport, StudyRecord
else:
    from prisma_model import PrismaReport, StudyRecord

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col = column_cells[0].column
        for cell in column_cells:
            value = str(cell.value or "")
            max_len = max(max_len, min(len(value), 60))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(max_len + 2, 50))


def _style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="808080"))
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _append_studies(ws, studies: list[StudyRecord]) -> None:
    ws.append(["bib_key", "titulo", "autores", "ano", "doi", "base", "fonte", "arquivo_local", "score_aderencia", "decisao", "motivo", "justificativa"])
    for s in studies:
        ws.append([
            s.bib_key, s.titulo, "; ".join(s.autores), s.ano, s.doi, s.base, s.fonte,
            s.arquivo_local, s.score_aderencia, s.decisao, s.motivo, s.justificativa,
        ])
    _style_header(ws)
    _autosize(ws)


def render_prisma_xlsx(report: PrismaReport, output_path: Path) -> Path:
    if Workbook is None:
        raise RuntimeError("openpyxl não está instalado para exportar XLSX.")
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Campo", "Valor"])
    meta = report.metadata
    rows = [
        ("Título", meta.titulo),
        ("Tema", meta.tema),
        ("Recorte", meta.recorte),
        ("Objetivo", meta.objetivo),
        ("Responsável", meta.responsavel),
        ("Data", meta.data_execucao),
        ("Identificados", report.flow.identificados),
        ("Duplicados removidos", report.flow.duplicados_removidos),
        ("Após deduplicação", report.flow.apos_deduplicacao),
        ("Triados", report.flow.triados_titulo_resumo),
        ("Excluídos na triagem", report.flow.excluidos_titulo_resumo),
        ("Avaliados em texto completo", report.flow.avaliados_texto_completo),
        ("Excluídos após texto completo", report.flow.excluidos_texto_completo),
        ("Incluídos", report.flow.incluidos),
    ]
    for row in rows:
        ws.append(list(row))
    _style_header(ws)
    _autosize(ws)

    ws_q = wb.create_sheet("Queries")
    ws_q.append(["base", "query", "resultados_brutos", "observacoes"])
    for q in report.search_strategy.queries:
        ws_q.append([q.base, q.query, q.resultados_brutos, q.observacoes])
    _style_header(ws_q); _autosize(ws_q)

    ws_i = wb.create_sheet("Incluidos")
    _append_studies(ws_i, report.included_studies)

    ws_e = wb.create_sheet("Excluidos")
    _append_studies(ws_e, report.excluded_studies)

    ws_a = wb.create_sheet("Todos")
    _append_studies(ws_a, report.all_records)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
