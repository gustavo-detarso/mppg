#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Busca bibliográfica externa e triagem PRISMA auditável.

O módulo consulta fontes bibliográficas públicas ou autenticadas, deduplica
metadados e gera uma planilha de triagem humana. A busca não baixa PDFs nem
decide inclusão/exclusão automaticamente. As credenciais ficam somente no
arquivo ``.env`` da raiz do Academic Pipeline; não são persistidas no TOML,
nem impressas em URLs, logs ou relatórios.
"""
from __future__ import annotations

import csv
import json
import os
import re
import time
import unicodedata
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

try:  # A dependência já é usada pelo pipeline principal.
    from dotenv import load_dotenv
except Exception:  # pragma: no cover - ambiente mínimo sem python-dotenv
    load_dotenv = None  # type: ignore[assignment]

Progress = Callable[[str], None] | None

# A planilha de triagem também funciona como matriz analítica após a leitura
# do texto completo. Os campos finais permanecem vazios na descoberta e são
# preenchidos apenas para os estudos efetivamente incluídos.
TRIAGE_HEADERS = [
    "id_registro", "titulo", "autores", "ano", "periodico", "doi", "url",
    "url_pdf_aberto", "resumo", "fontes", "pontuacao_relevancia",
    "decisao_titulo_resumo", "motivo_exclusao_titulo_resumo", "texto_completo_local",
    "decisao_texto_completo", "motivo_exclusao_texto_completo", "incluir_final", "observacoes",
    "pais_contexto", "objetivo_estudo", "desenho_metodo", "amostra_base",
    "achados_principais", "limitacoes", "contribuicao_pergunta",
]

MATRIX_ANALYTICAL_HEADERS = [
    "pais_contexto", "objetivo_estudo", "desenho_metodo", "amostra_base",
    "achados_principais", "limitacoes", "contribuicao_pergunta",
]

MATRIX_FIELD_GUIDE = [
    ("pais_contexto", "País, organização, população ou contexto institucional estudado."),
    ("objetivo_estudo", "Objetivo declarado pelo estudo incluído."),
    ("desenho_metodo", "Desenho, abordagem e método de análise."),
    ("amostra_base", "Amostra, período, base de dados ou material empírico."),
    ("achados_principais", "Principais resultados relevantes para a pergunta da revisão."),
    ("limitacoes", "Limitações metodológicas ou de escopo informadas/identificadas."),
    ("contribuicao_pergunta", "Contribuição específica do estudo para responder à pergunta de pesquisa."),
]

# Fontes de descoberta. Unpaywall é deliberadamente separado: ele não é um
# índice bibliográfico, mas um enriquecedor de links abertos por DOI.
PROVIDER_ORDER = [
    "crossref",
    "openalex",
    "semantic_scholar",
    "scopus",
    "wos",
    "pubmed",
    "europe_pmc",
    "scielo",
    "core",
]

PROVIDER_LABELS = {
    "crossref": "Crossref",
    "openalex": "OpenAlex",
    "semantic_scholar": "Semantic Scholar",
    "scopus": "Scopus",
    "wos": "Web of Science",
    "pubmed": "PubMed/NCBI",
    "europe_pmc": "Europe PMC",
    "scielo": "SciELO",
    "core": "CORE",
}

_PROVIDER_CREDENTIALS: dict[str, tuple[str, ...]] = {
    "crossref": ("CROSSREF_EMAIL",),
    "openalex": ("OPENALEX_API_KEY", "OPENALEX_EMAIL"),
    "semantic_scholar": ("SEMANTIC_SCHOLAR_API_KEY",),
    "scopus": ("SCOPUS_API_KEY", "SCOPUS_INSTTOKEN"),
    "wos": ("WOS_API_KEY",),
    "pubmed": ("NCBI_API_KEY", "NCBI_EMAIL"),
    "europe_pmc": ("EUROPEPMC_EMAIL",),
    "scielo": ("SCIELO_API_TOKEN",),
    "core": ("CORE_API_KEY", "CORE_EMAIL"),
    "unpaywall": ("UNPAYWALL_EMAIL",),
}

_REQUIRED_ENV: dict[str, tuple[str, ...]] = {
    "scopus": ("SCOPUS_API_KEY",),
    "wos": ("WOS_API_KEY",),
    "core": ("CORE_API_KEY",),
    "unpaywall": ("UNPAYWALL_EMAIL",),
}


# ---------------------------------------------------------------------------
# Ambiente, catálogo e utilitários
# ---------------------------------------------------------------------------


def _now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _project_root() -> Path:
    # .../app_bundle/scripts/pipeline/prisma_busca_externa.py -> raiz do projeto
    return Path(__file__).resolve().parents[3]


def load_search_environment() -> None:
    """Carrega ``.env`` da raiz sem sobrescrever variáveis já exportadas."""
    if load_dotenv is None:
        return
    env_path = _project_root() / ".env"
    if env_path.is_file():
        load_dotenv(dotenv_path=env_path, override=False)


def _has_env(name: str) -> bool:
    return bool(os.getenv(name, "").strip())


def provider_statuses() -> dict[str, dict[str, Any]]:
    """Retorna status de configuração sem revelar valores secretos."""
    load_search_environment()
    values: dict[str, dict[str, Any]] = {}
    for provider in PROVIDER_ORDER:
        names = _PROVIDER_CREDENTIALS.get(provider, ())
        detected = [name for name in names if _has_env(name)]
        required = _REQUIRED_ENV.get(provider, ())
        missing_required = [name for name in required if not _has_env(name)]
        if missing_required:
            status = "credencial obrigatória ausente"
            available = False
        elif detected:
            status = "credencial/e-mail detectado"
            available = True
        else:
            status = "acesso público; limites da fonte podem ser menores"
            available = True
        if provider == "scielo" and not detected:
            status = "acesso público; token opcional não detectado"
        values[provider] = {
            "key": provider,
            "label": PROVIDER_LABELS[provider],
            "available": available,
            "status": status,
            "credentials_detected": detected,
            "credentials_required_missing": missing_required,
        }
    values["unpaywall"] = {
        "key": "unpaywall",
        "label": "Unpaywall (enriquecimento de acesso aberto)",
        "available": _has_env("UNPAYWALL_EMAIL"),
        "status": "e-mail detectado" if _has_env("UNPAYWALL_EMAIL") else "UNPAYWALL_EMAIL ausente",
        "credentials_detected": ["UNPAYWALL_EMAIL"] if _has_env("UNPAYWALL_EMAIL") else [],
        "credentials_required_missing": [] if _has_env("UNPAYWALL_EMAIL") else ["UNPAYWALL_EMAIL"],
    }
    return values


def provider_selection_choices() -> list[tuple[str, str]]:
    """Opções para o wizard, incluindo o atalho explícito ``Todas``."""
    status = provider_statuses()
    values: list[tuple[str, str]] = [("todas", "[Todas] Selecionar todas as fontes de descoberta")]
    for provider in PROVIDER_ORDER:
        item = status[provider]
        suffix = item["status"]
        values.append((provider, f"{item['label']} — {suffix}"))
    return values


def expand_provider_selection(values: Iterable[Any]) -> list[str]:
    """Expande a escolha especial ``todas`` e normaliza aliases históricos."""
    aliases = {
        "all": "todas",
        "__todas__": "todas",
        "open_alex": "openalex",
        "cross_ref": "crossref",
        "semantic": "semantic_scholar",
        "semanticscholar": "semantic_scholar",
        "web_of_science": "wos",
        "webofscience": "wos",
        "ncbi": "pubmed",
        "europepmc": "europe_pmc",
        "europe_pmc": "europe_pmc",
    }
    requested = [aliases.get(str(item).strip().lower().replace("-", "_").replace(" ", "_"), str(item).strip().lower()) for item in values]
    if "todas" in requested:
        return list(PROVIDER_ORDER)
    return list(dict.fromkeys(item for item in requested if item in PROVIDER_ORDER))


def _section(cfg: dict[str, Any], name: str) -> dict[str, Any]:
    value = cfg.get(name, {})
    return value if isinstance(value, dict) else {}


def external_search_enabled(cfg: dict[str, Any]) -> bool:
    section = _section(cfg, "busca_prisma")
    return bool(section.get("ativo", False)) and str(section.get("modo") or "").strip().lower() == "busca_externa"


def _list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        return [str(item).strip() for item in value if str(item).strip()]
    return [item.strip() for item in re.split(r"[;,\n]+", str(value)) if item.strip()]


def _norm(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text.lower()).strip()


def _doi(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", text)
    text = re.sub(r"^doi:\s*", "", text)
    return text.rstrip(".,;)")


def _title_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _norm(value))


def _int(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _authors(value: Any) -> str:
    if isinstance(value, list):
        result: list[str] = []
        for item in value:
            if isinstance(item, dict):
                name = str(item.get("name") or item.get("display_name") or item.get("full_name") or item.get("family") or "").strip()
            else:
                name = str(item or "").strip()
            if name:
                result.append(name)
        return "; ".join(result)
    return str(value or "").strip()


def _openalex_abstract(value: Any) -> str:
    if not isinstance(value, dict):
        return ""
    parts: list[tuple[int, str]] = []
    for token, positions in value.items():
        if not isinstance(positions, list):
            continue
        for position in positions:
            try:
                parts.append((int(position), str(token)))
            except (TypeError, ValueError):
                pass
    return " ".join(token for _pos, token in sorted(parts))


def _safe_url(url: str) -> str:
    """Remove credenciais e e-mails antes de registrar qualquer URL."""
    return re.sub(
        r"([?&](?:apiKey|api_key|key|token|insttoken|mailto|email)=)[^&]+",
        r"\1REDACTED",
        str(url or ""),
        flags=re.I,
    )


def _request_text(url: str, *, headers: dict[str, str] | None = None, timeout: int = 40) -> str:
    request_headers = {
        "Accept": "application/json, application/xml, text/xml, text/plain;q=0.8, */*;q=0.2",
        "User-Agent": "AcademicPipeline-PRISMA/1.1",
    }
    request_headers.update(headers or {})
    request = Request(url, headers=request_headers, method="GET")
    try:
        with urlopen(request, timeout=timeout) as response:  # nosec B310: endpoints are fixed APIs.
            return response.read().decode("utf-8", errors="replace")
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")[:300] if getattr(exc, "fp", None) else ""
        raise RuntimeError(f"HTTP {exc.code}: {body}") from exc
    except URLError as exc:
        raise RuntimeError(f"Falha de rede: {exc.reason}") from exc
    except TimeoutError as exc:
        raise RuntimeError("Tempo esgotado na consulta bibliográfica.") from exc


def _request_json(url: str, *, headers: dict[str, str] | None = None, timeout: int = 40) -> Any:
    text = _request_text(url, headers=headers, timeout=timeout)
    try:
        return json.loads(text)
    except json.JSONDecodeError as exc:
        raise RuntimeError("A fonte não retornou JSON válido.") from exc


def _record(source: str, source_id: Any, *, title: Any = "", authors: Any = "", year: Any = "", venue: Any = "", doi: Any = "", url: Any = "", abstract: Any = "", pdf: Any = "", citations: Any = None, kind: Any = "") -> dict[str, Any]:
    return {
        "id_registro": f"{source}:{source_id}",
        "titulo": re.sub(r"\s+", " ", str(title or "")).strip(),
        "autores": _authors(authors),
        "ano": str(year or "").strip(),
        "periodico": re.sub(r"\s+", " ", str(venue or "")).strip(),
        "doi": _doi(doi),
        "url": str(url or "").strip(),
        "url_pdf_aberto": str(pdf or "").strip(),
        "resumo": re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", str(abstract or ""))).strip(),
        "fontes": [source],
        "pontuacao_relevancia": 0.0,
        "citacoes": _int(citations),
        "tipo_publicacao": str(kind or "").strip(),
    }


def _env(name: str) -> str:
    load_search_environment()
    return os.getenv(name, "").strip()


def _provider_email(provider: str, fallback: str = "") -> str:
    fields = {
        "crossref": ("CROSSREF_EMAIL",),
        "openalex": ("OPENALEX_EMAIL",),
        "pubmed": ("NCBI_EMAIL",),
        "europe_pmc": ("EUROPEPMC_EMAIL",),
        "core": ("CORE_EMAIL",),
    }
    for name in fields.get(provider, ()):  # prioriza e-mail específico da fonte
        value = _env(name)
        if value:
            return value
    return str(fallback or "").strip()


def _query_config(cfg: dict[str, Any]) -> dict[str, Any]:
    load_search_environment()
    search = _section(cfg, "busca_prisma")
    research = _section(cfg, "pesquisa")
    providers = expand_provider_selection(_list(search.get("bases")))
    if not providers:
        providers = ["crossref", "openalex", "semantic_scholar"]
    keywords = _list(research.get("palavras_chave"))
    query = str(search.get("consulta_geral") or "").strip() or " ".join(keywords).strip() or str(research.get("tema") or "").strip()
    if not query:
        raise RuntimeError("Defina consulta_geral, palavras_chave ou tema antes de executar a busca externa.")

    def clamp(value: Any, default: int, low: int, high: int) -> int:
        try:
            return max(low, min(int(value), high))
        except (TypeError, ValueError):
            return default

    fallback_email = str(search.get("email_contato") or "").strip()
    return {
        "providers": providers,
        "query": query,
        "keywords": keywords,
        "per_base": clamp(search.get("limite_por_base"), 100, 10, 200),
        "initial_limit": clamp(search.get("limite_triagem_inicial"), 250, 10, 1000),
        "target": clamp(search.get("meta_estudos_incluidos"), 15, 1, 500),
        "year_start": str(search.get("ano_inicio") or "").strip(),
        "year_end": str(search.get("ano_fim") or "").strip(),
        "email": fallback_email,
        "emails": {name: _provider_email(name, fallback_email) for name in ("crossref", "openalex", "pubmed", "europe_pmc", "core")},
        "inclusion": _list(search.get("criterios_inclusao") or _section(cfg, "relatorio_pesquisa").get("criterios_inclusao")),
        "exclusion": _list(search.get("criterios_exclusao") or _section(cfg, "relatorio_pesquisa").get("criterios_exclusao")),
        "corpus_local": str(search.get("corpus_local_opcional") or "").strip(),
        "enrich_unpaywall": bool(search.get("enriquecer_unpaywall", False)),
        "unpaywall_limit": clamp(search.get("limite_unpaywall"), 100, 0, 250),
    }


# ---------------------------------------------------------------------------
# Adaptadores das fontes
# ---------------------------------------------------------------------------


def _crossref(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    params: dict[str, Any] = {
        "query.bibliographic": query,
        "rows": config["per_base"],
        "select": "DOI,title,author,published-print,published-online,issued,container-title,abstract,type,URL",
    }
    filters: list[str] = []
    if config["year_start"]:
        filters.append(f"from-pub-date:{config['year_start']}-01-01")
    if config["year_end"]:
        filters.append(f"until-pub-date:{config['year_end']}-12-31")
    if filters:
        params["filter"] = ",".join(filters)
    if config["emails"]["crossref"]:
        params["mailto"] = config["emails"]["crossref"]
    url = "https://api.crossref.org/works?" + urlencode(params)
    items = ((_request_json(url) or {}).get("message") or {}).get("items") or []
    records: list[dict[str, Any]] = []
    for idx, item in enumerate(items, 1):
        titles = item.get("title") or []
        venues = item.get("container-title") or []
        dates = (item.get("published-print") or item.get("published-online") or item.get("issued") or {}).get("date-parts") or []
        year = dates[0][0] if dates and isinstance(dates[0], list) and dates[0] else ""
        records.append(_record("crossref", item.get("DOI") or idx, title=titles[0] if titles else "", authors=item.get("author"), year=year, venue=venues[0] if venues else "", doi=item.get("DOI"), url=item.get("URL"), abstract=item.get("abstract"), kind=item.get("type")))
    return records, {"provider": "crossref", "url": _safe_url(url), "retrieved": len(records), "error": ""}


def _openalex(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    params: dict[str, Any] = {"search": query, "per-page": config["per_base"]}
    filters: list[str] = []
    if config["year_start"]:
        filters.append(f"from_publication_date:{config['year_start']}-01-01")
    if config["year_end"]:
        filters.append(f"to_publication_date:{config['year_end']}-12-31")
    if filters:
        params["filter"] = ",".join(filters)
    if config["emails"]["openalex"]:
        params["mailto"] = config["emails"]["openalex"]
    if _env("OPENALEX_API_KEY"):
        params["api_key"] = _env("OPENALEX_API_KEY")
    url = "https://api.openalex.org/works?" + urlencode(params)
    items = (_request_json(url) or {}).get("results") or []
    records: list[dict[str, Any]] = []
    for idx, item in enumerate(items, 1):
        loc = item.get("primary_location") or {}
        source = loc.get("source") or {}
        open_access = item.get("open_access") or {}
        authors = []
        for item_author in item.get("authorships") or []:
            name = str((item_author.get("author") or {}).get("display_name") or "").strip()
            if name:
                authors.append(name)
        records.append(_record("openalex", item.get("id") or idx, title=item.get("title"), authors="; ".join(authors), year=item.get("publication_year"), venue=source.get("display_name"), doi=item.get("doi"), url=item.get("doi") or item.get("id"), abstract=_openalex_abstract(item.get("abstract_inverted_index")), pdf=loc.get("pdf_url") or open_access.get("oa_url"), citations=item.get("cited_by_count"), kind=item.get("type")))
    return records, {"provider": "openalex", "url": _safe_url(url), "retrieved": len(records), "error": ""}


def _semantic_scholar(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    params = {"query": query, "limit": config["per_base"], "fields": "title,abstract,authors,year,venue,externalIds,url,openAccessPdf,citationCount,publicationTypes"}
    url = "https://api.semanticscholar.org/graph/v1/paper/search?" + urlencode(params)
    headers: dict[str, str] = {}
    key = _env("SEMANTIC_SCHOLAR_API_KEY")
    if key:
        headers["x-api-key"] = key
    items = (_request_json(url, headers=headers) or {}).get("data") or []
    records: list[dict[str, Any]] = []
    for idx, item in enumerate(items, 1):
        identifiers = item.get("externalIds") or {}
        oa = item.get("openAccessPdf") or {}
        records.append(_record("semantic_scholar", item.get("paperId") or idx, title=item.get("title"), authors=item.get("authors"), year=item.get("year"), venue=item.get("venue"), doi=identifiers.get("DOI"), url=item.get("url"), abstract=item.get("abstract"), pdf=oa.get("url"), citations=item.get("citationCount"), kind="; ".join(_list(item.get("publicationTypes")))))
    return records, {"provider": "semantic_scholar", "url": _safe_url(url), "retrieved": len(records), "error": ""}


def _scopus(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    key = _env("SCOPUS_API_KEY")
    if not key:
        raise RuntimeError("SCOPUS_API_KEY não configurada; a fonte foi ignorada.")
    terms = [part.strip().replace('"', '') for part in re.split(r"\s*;\s*|\s+OR\s+", query, flags=re.I) if part.strip()]
    scopus_query = "TITLE-ABS-KEY(" + " OR ".join(f'"{term}"' for term in terms) + ")"
    params: dict[str, Any] = {"query": scopus_query, "count": config["per_base"], "httpAccept": "application/json"}
    url = "https://api.elsevier.com/content/search/scopus?" + urlencode(params)
    headers = {"X-ELS-APIKey": key}
    if _env("SCOPUS_INSTTOKEN"):
        headers["X-ELS-Insttoken"] = _env("SCOPUS_INSTTOKEN")
    entries = ((_request_json(url, headers=headers) or {}).get("search-results") or {}).get("entry") or []
    records = [_record("scopus", item.get("dc:identifier") or idx, title=item.get("dc:title"), authors=item.get("dc:creator"), year=str(item.get("prism:coverDate") or "")[:4], venue=item.get("prism:publicationName"), doi=item.get("prism:doi"), url=item.get("prism:url"), citations=item.get("citedby-count"), kind=item.get("subtypeDescription")) for idx, item in enumerate(entries, 1)]
    return records, {"provider": "scopus", "url": _safe_url(url), "retrieved": len(records), "error": ""}


def _wos(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    key = _env("WOS_API_KEY")
    if not key:
        raise RuntimeError("WOS_API_KEY não configurada; a fonte foi ignorada.")
    # Web of Science Starter API. A URL pode ser sobrescrita no ambiente em
    # caso de plano/tenant institucional com rota própria.
    base = _env("WOS_API_BASE") or "https://api.clarivate.com/apis/wos-starter/v1/documents"
    params: dict[str, Any] = {"q": query, "limit": config["per_base"], "page": 1}
    if config["year_start"]:
        params["fromYear"] = config["year_start"]
    if config["year_end"]:
        params["toYear"] = config["year_end"]
    url = base + ("&" if "?" in base else "?") + urlencode(params)
    payload = _request_json(url, headers={"X-ApiKey": key}) or {}
    items = payload.get("hits") or payload.get("documents") or payload.get("data") or []
    records: list[dict[str, Any]] = []
    for idx, item in enumerate(items, 1):
        source = item.get("source") or {}
        names = (item.get("names") or {}).get("authors") or {}
        authors = names.get("authors") if isinstance(names, dict) else item.get("authors")
        identifiers = item.get("identifiers") or {}
        doi = identifiers.get("doi") if isinstance(identifiers, dict) else item.get("doi")
        year = source.get("publishYear") or source.get("year") or item.get("year")
        venue = source.get("title") or item.get("sourceTitle") or item.get("journal")
        records.append(_record("wos", item.get("uid") or item.get("id") or idx, title=item.get("title"), authors=authors, year=year, venue=venue, doi=doi, url=item.get("url") or item.get("links", {}).get("record"), abstract=item.get("abstract"), citations=item.get("timesCited"), kind=item.get("documentType")))
    return records, {"provider": "wos", "url": _safe_url(url), "retrieved": len(records), "error": ""}


def _pubmed(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    params: dict[str, Any] = {"db": "pubmed", "term": query, "retmax": config["per_base"], "retmode": "json", "sort": "relevance", "tool": "AcademicPipelinePRISMA"}
    if config["year_start"] or config["year_end"]:
        params["datetype"] = "pdat"
        if config["year_start"]:
            params["mindate"] = config["year_start"] + "/01/01"
        if config["year_end"]:
            params["maxdate"] = config["year_end"] + "/12/31"
    email = config["emails"]["pubmed"]
    if email:
        params["email"] = email
    if _env("NCBI_API_KEY"):
        params["api_key"] = _env("NCBI_API_KEY")
    search_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?" + urlencode(params)
    result = (_request_json(search_url) or {}).get("esearchresult") or {}
    ids = [str(item) for item in result.get("idlist") or [] if str(item)]
    if not ids:
        return [], {"provider": "pubmed", "url": _safe_url(search_url), "retrieved": 0, "error": ""}
    summary_params: dict[str, Any] = {"db": "pubmed", "id": ",".join(ids), "retmode": "json", "tool": "AcademicPipelinePRISMA"}
    if email:
        summary_params["email"] = email
    if _env("NCBI_API_KEY"):
        summary_params["api_key"] = _env("NCBI_API_KEY")
    summary_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi?" + urlencode(summary_params)
    payload = _request_json(summary_url) or {}
    summary = payload.get("result") or {}
    records: list[dict[str, Any]] = []
    for idx, uid in enumerate(ids, 1):
        item = summary.get(uid) or {}
        article_ids = item.get("articleids") or []
        doi = next((value.get("value") for value in article_ids if isinstance(value, dict) and str(value.get("idtype") or "").lower() == "doi"), "")
        authors = item.get("authors") or []
        pubdate = str(item.get("pubdate") or item.get("sortpubdate") or "")
        year_match = re.search(r"\b(19|20)\d{2}\b", pubdate)
        records.append(_record("pubmed", uid, title=item.get("title"), authors=authors, year=year_match.group(0) if year_match else "", venue=item.get("fulljournalname") or item.get("source"), doi=doi, url=f"https://pubmed.ncbi.nlm.nih.gov/{uid}/", kind=item.get("pubtype")))
    return records, {"provider": "pubmed", "url": _safe_url(search_url), "retrieved": len(records), "error": ""}


def _europe_pmc(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    params: dict[str, Any] = {"query": query, "format": "json", "resultType": "core", "pageSize": config["per_base"]}
    if config["year_start"]:
        params["query"] = f"({query}) AND FIRST_PDATE:[{config['year_start']}-01-01 TO *]"
    if config["year_end"]:
        params["query"] = f"({params['query']}) AND FIRST_PDATE:[* TO {config['year_end']}-12-31]"
    url = "https://www.ebi.ac.uk/europepmc/webservices/rest/search?" + urlencode(params)
    results = ((_request_json(url) or {}).get("resultList") or {}).get("result") or []
    records: list[dict[str, Any]] = []
    for idx, item in enumerate(results, 1):
        pdf = ""
        fulltext = item.get("fullTextUrlList") or {}
        for candidate in fulltext.get("fullTextUrl") or []:
            if isinstance(candidate, dict) and candidate.get("documentStyle") in {"pdf", "PDF"}:
                pdf = str(candidate.get("url") or "")
                break
        records.append(_record("europe_pmc", item.get("id") or item.get("pmid") or item.get("pmcid") or idx, title=item.get("title"), authors=item.get("authorString"), year=item.get("pubYear"), venue=item.get("journalTitle"), doi=item.get("doi"), url=item.get("doi") and f"https://doi.org/{item.get('doi')}" or item.get("fullTextUrl"), abstract=item.get("abstractText"), pdf=pdf, citations=item.get("citedByCount"), kind=item.get("pubType")))
    return records, {"provider": "europe_pmc", "url": _safe_url(url), "retrieved": len(records), "error": ""}


def _xml_text(node: ET.Element | None, *names: str) -> str:
    if node is None:
        return ""
    for name in names:
        found = node.find(name)
        if found is not None and found.text:
            return found.text.strip()
    return ""


def _scielo(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    # O buscador SciELO expõe resposta summary em XML. O token, quando
    # fornecido, é encaminhado em cabeçalho sem aparecer no relatório.
    params: list[tuple[str, str]] = [
        ("q", query), ("output", "site"), ("lang", "pt"), ("from", "0"), ("sort", ""),
        ("format", "summary"), ("count", str(config["per_base"])), ("fb", ""), ("page", "1"),
        ("filter[in][]", "scl"), ("filter[type][]", "article"),
    ]
    url = "https://search.scielo.org/?" + urlencode(params)
    headers: dict[str, str] = {}
    token = _env("SCIELO_API_TOKEN")
    if token:
        headers["Authorization"] = f"Bearer {token}"
    raw = _request_text(url, headers=headers)
    records: list[dict[str, Any]] = []
    try:
        root = ET.fromstring(raw)
        items = root.findall(".//item") or root.findall(".//doc")
        for idx, item in enumerate(items, 1):
            title = _xml_text(item, "title", "titulo")
            link = _xml_text(item, "link", "url")
            description = _xml_text(item, "description", "abstract", "resumo")
            creator = _xml_text(item, "creator", "author", "autor")
            date = _xml_text(item, "pubDate", "date", "year", "ano")
            doi_match = re.search(r"10\.\d{4,9}/[-._;()/:a-z0-9]+", raw if not link else link + " " + description, flags=re.I)
            records.append(_record("scielo", idx, title=title, authors=creator, year=(re.search(r"\b(19|20)\d{2}\b", date) or [""])[0], url=link, abstract=description, doi=doi_match.group(0) if doi_match else "", kind="article"))
    except ET.ParseError:
        # Fallback conservador para resposta HTML/sumário não XML: não tenta
        # inventar metadados; comunica a incompatibilidade de formato.
        raise RuntimeError("A busca SciELO não retornou um sumário XML compatível. Tente novamente ou configure uma rota institucional em SCIELO_SEARCH_URL.")
    return records, {"provider": "scielo", "url": _safe_url(url), "retrieved": len(records), "error": ""}


def _core(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    key = _env("CORE_API_KEY")
    if not key:
        raise RuntimeError("CORE_API_KEY não configurada; a fonte foi ignorada.")
    base = _env("CORE_API_BASE") or "https://api.core.ac.uk/v3/search/works"
    params: dict[str, Any] = {"q": query, "limit": config["per_base"]}
    if config["year_start"]:
        params["fromYear"] = config["year_start"]
    if config["year_end"]:
        params["toYear"] = config["year_end"]
    url = base + ("&" if "?" in base else "?") + urlencode(params)
    payload = _request_json(url, headers={"Authorization": f"Bearer {key}"}) or {}
    items = payload.get("results") or payload.get("data") or []
    records: list[dict[str, Any]] = []
    for idx, item in enumerate(items, 1):
        authors = item.get("authors") or item.get("authorsString") or []
        doi = item.get("doi") or item.get("identifiers", {}).get("doi") if isinstance(item.get("identifiers"), dict) else item.get("doi")
        records.append(_record("core", item.get("id") or idx, title=item.get("title"), authors=authors, year=item.get("publishedDate") or item.get("yearPublished"), venue=item.get("journals") or item.get("publisher"), doi=doi, url=item.get("downloadUrl") or item.get("sourceFulltextUrls") or item.get("url"), abstract=item.get("abstract"), pdf=item.get("downloadUrl"), citations=item.get("citedByCount"), kind=item.get("documentType")))
    return records, {"provider": "core", "url": _safe_url(url), "retrieved": len(records), "error": ""}


# ---------------------------------------------------------------------------
# Consolidação, Unpaywall e saídas
# ---------------------------------------------------------------------------


def _merge(first: dict[str, Any], second: dict[str, Any]) -> None:
    for key in ("titulo", "autores", "ano", "periodico", "doi", "url", "url_pdf_aberto", "resumo", "tipo_publicacao"):
        if not first.get(key) and second.get(key):
            first[key] = second[key]
    first["fontes"] = sorted(set(_list(first.get("fontes")) + _list(second.get("fontes"))))
    first["citacoes"] = max(_int(first.get("citacoes")) or 0, _int(second.get("citacoes")) or 0) or None


def _deduplicate(records: Iterable[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    out: list[dict[str, Any]] = []
    by_doi: dict[str, dict[str, Any]] = {}
    by_title: dict[str, dict[str, Any]] = {}
    removed = 0
    for record in records:
        doi = _doi(record.get("doi"))
        title = _title_key(record.get("titulo"))
        existing = by_doi.get(doi) if doi else by_title.get(title)
        if existing:
            _merge(existing, record)
            removed += 1
            continue
        if doi:
            by_doi[doi] = record
        if title:
            by_title[title] = record
        out.append(record)
    return out, removed


def _score(record: dict[str, Any], keywords: list[str]) -> float:
    haystack = _norm(" ".join(str(record.get(key) or "") for key in ("titulo", "resumo", "periodico")))
    if not haystack:
        return 0.0
    score = 0.0
    for keyword in keywords:
        token = _norm(keyword)
        if token and token in haystack:
            score += 1.0
    return score


def _enrich_with_unpaywall(records: list[dict[str, Any]], config: dict[str, Any], *, progress: Progress = None) -> dict[str, Any]:
    email = _env("UNPAYWALL_EMAIL")
    if not config.get("enrich_unpaywall"):
        return {"enabled": False, "checked": 0, "enriched": 0, "reason": "desativado"}
    if not email:
        return {"enabled": True, "checked": 0, "enriched": 0, "reason": "UNPAYWALL_EMAIL ausente"}
    candidates = [item for item in records if item.get("doi") and not item.get("url_pdf_aberto")]
    checked = 0
    enriched = 0
    for item in candidates[: int(config.get("unpaywall_limit") or 0)]:
        if progress:
            progress(f"Consultando Unpaywall para acesso aberto ({checked + 1}/{min(len(candidates), int(config.get('unpaywall_limit') or 0))})")
        url = f"https://api.unpaywall.org/v2/{_doi(item.get('doi'))}?" + urlencode({"email": email})
        try:
            payload = _request_json(url) or {}
        except Exception:
            checked += 1
            time.sleep(0.12)
            continue
        locations = [payload.get("best_oa_location"), *(payload.get("oa_locations") or [])]
        for location in locations:
            if not isinstance(location, dict):
                continue
            pdf = str(location.get("url_for_pdf") or "").strip()
            landing = str(location.get("url") or "").strip()
            if pdf or landing:
                item["url_pdf_aberto"] = pdf or landing
                enriched += 1
                break
        checked += 1
        time.sleep(0.12)
    return {"enabled": True, "checked": checked, "enriched": enriched, "reason": ""}


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _write_csv(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=TRIAGE_HEADERS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            prepared = dict(row)
            prepared["fontes"] = "; ".join(_list(prepared.get("fontes")))
            writer.writerow({field: prepared.get(field, "") for field in TRIAGE_HEADERS})


def _write_xlsx(path: Path, rows: list[dict[str, Any]]) -> str | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        return None
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Triagem e matriz"
    sheet.append(TRIAGE_HEADERS)
    for row in rows:
        sheet.append(["; ".join(_list(row.get(field))) if field == "fontes" else row.get(field, "") for field in TRIAGE_HEADERS])
    sheet.freeze_panes = "A2"
    header_fill = PatternFill(fill_type="solid", fgColor="003E7E")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column in sheet.columns:
        letter = column[0].column_letter
        header = str(column[0].value or "")
        width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 48)
        if header in MATRIX_ANALYTICAL_HEADERS:
            width = max(width, 28)
        sheet.column_dimensions[letter].width = width

    guide = workbook.create_sheet("Guia da matriz")
    guide.append(["Campo", "Preenchimento esperado"])
    for field, description in MATRIX_FIELD_GUIDE:
        guide.append([field, description])
    for cell in guide[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    guide.column_dimensions["A"].width = 30
    guide.column_dimensions["B"].width = 94
    guide.freeze_panes = "A2"
    workbook.save(path)
    return str(path)


def _protocol_markdown(config: dict[str, Any], logs: list[dict[str, Any]], counts: dict[str, int], unpaywall: dict[str, Any]) -> str:
    lines = [
        "# Protocolo e registro de busca PRISMA", "",
        f"- Execução: {_now()}", f"- Consulta geral: `{config['query']}`",
        f"- Bases selecionadas: {', '.join(config['providers'])}",
        f"- Meta de estudos incluídos: {config['target']}",
        f"- Unpaywall: {'ativado' if unpaywall.get('enabled') else 'desativado'}; registros enriquecidos: {unpaywall.get('enriched', 0)}", "",
        "## Critérios de inclusão", *[f"- {item}" for item in (config["inclusion"] or ["Não informado."])], "",
        "## Critérios de exclusão", *[f"- {item}" for item in (config["exclusion"] or ["Não informado."])], "",
        "## Contagens", *[f"- {name.replace('_', ' ')}: {value}" for name, value in counts.items()], "",
        "## Registro por fonte",
    ]
    for log in logs:
        lines.append(f"- **{log['provider']}** — recuperados: {log.get('retrieved', 0)}")
        if log.get("url"):
            lines.append(f"  - consulta: `{log['url']}`")
        if log.get("error"):
            lines.append(f"  - aviso/erro: {log['error']}")
    lines += ["", "## Próxima etapa obrigatória", "Preencha a planilha de triagem. O sistema não define automaticamente a inclusão ou a exclusão de estudos."]
    return "\n".join(lines).rstrip() + "\n"


def run_external_prisma_search(cfg: dict[str, Any], out_dir: Path, prefix: str, *, progress: Progress = None) -> dict[str, Any]:
    config = _query_config(cfg)
    fetchers = {
        "crossref": _crossref,
        "openalex": _openalex,
        "semantic_scholar": _semantic_scholar,
        "scopus": _scopus,
        "wos": _wos,
        "pubmed": _pubmed,
        "europe_pmc": _europe_pmc,
        "scielo": _scielo,
        "core": _core,
    }
    raw: list[dict[str, Any]] = []
    logs: list[dict[str, Any]] = []
    successes = 0
    for source in config["providers"]:
        if progress:
            progress(f"Consultando {PROVIDER_LABELS[source]} para descoberta bibliográfica")
        try:
            values, log = fetchers[source](config["query"], config)
            raw.extend(values)
            logs.append(log)
            successes += 1
        except Exception as exc:
            logs.append({"provider": source, "url": "", "retrieved": 0, "error": str(exc)})
        time.sleep(0.15)
    if not successes:
        errors = "; ".join(f"{item['provider']}: {item['error']}" for item in logs)
        raise RuntimeError("Nenhuma fonte bibliográfica respondeu com sucesso. " + errors)
    if progress:
        progress("Deduplicando registros por DOI e título normalizado")
    deduplicated, removed = _deduplicate(raw)
    for item in deduplicated:
        item["pontuacao_relevancia"] = _score(item, config["keywords"])
    deduplicated.sort(key=lambda item: (-float(item.get("pontuacao_relevancia") or 0), str(item.get("titulo") or "").lower()))
    triage = deduplicated[:config["initial_limit"]]
    unpaywall = _enrich_with_unpaywall(triage, config, progress=progress)
    for item in triage:
        item.update({"decisao_titulo_resumo": "PENDENTE", "motivo_exclusao_titulo_resumo": "", "texto_completo_local": "", "decisao_texto_completo": "NAO_INICIADO", "motivo_exclusao_texto_completo": "", "incluir_final": "PENDENTE", "observacoes": ""})
    out_dir.mkdir(parents=True, exist_ok=True)
    raw_path = out_dir / f"{prefix}.candidatos_brutos.json"
    dedup_path = out_dir / f"{prefix}.candidatos_deduplicados.json"
    triage_csv = out_dir / f"{prefix}.triagem_titulo_resumo.csv"
    triage_xlsx = out_dir / f"{prefix}.triagem_titulo_resumo.xlsx"
    protocol = out_dir / f"{prefix}.protocolo_busca_prisma.md"
    log_path = out_dir / f"{prefix}.busca_prisma_log.json"
    report_path = out_dir / f"{prefix}.prisma_report.json"
    counts = {"registros_identificados": len(raw), "duplicatas_removidas": removed, "registros_apos_deduplicacao": len(deduplicated), "registros_enviados_para_triagem": len(triage), "triagem_titulo_resumo_concluida": 0, "textos_completos_avaliados": 0, "estudos_incluidos": 0}
    _write_json(raw_path, raw)
    _write_json(dedup_path, deduplicated)
    _write_csv(triage_csv, triage)
    xlsx = _write_xlsx(triage_xlsx, triage)
    _write_json(log_path, {"gerado_em": _now(), "configuracao": {"providers": config["providers"], "query": config["query"], "per_base": config["per_base"], "initial_limit": config["initial_limit"], "target": config["target"], "year_start": config["year_start"], "year_end": config["year_end"], "enrich_unpaywall": config["enrich_unpaywall"]}, "fontes": logs, "unpaywall": unpaywall})
    protocol.write_text(_protocol_markdown(config, logs, counts, unpaywall), encoding="utf-8")
    output = {"schema_version": "1.1", "tipo": "prisma_busca_externa", "gerado_em": _now(), "status": "triagem_titulo_resumo_pendente", "meta_estudos_incluidos": config["target"], "consulta_geral": config["query"], "palavras_chave": config["keywords"], "bases": config["providers"], "criterios_inclusao": config["inclusion"], "criterios_exclusao": config["exclusion"], "contagens": counts, "fontes": logs, "unpaywall": unpaywall, "artefatos": {"candidatos_brutos": str(raw_path), "candidatos_deduplicados": str(dedup_path), "planilha_triagem_csv": str(triage_csv), "protocolo": str(protocol), "log": str(log_path)}}
    if xlsx:
        output["artefatos"]["planilha_triagem_xlsx"] = xlsx
    _write_json(report_path, output)
    output["artefatos"]["prisma_report_json"] = str(report_path)
    return output



# ---------------------------------------------------------------------------
# Relatório PRISMA em ORG (prévio e consolidado)
# ---------------------------------------------------------------------------


def _org_text(value: Any) -> str:
    """Normaliza texto de metadados para uso seguro em linhas ORG."""
    raw = "" if value is None else str(value)
    text = " ".join(raw.replace("\u00a0", " ").split())
    return text.replace("|", "/")


def _org_table(headers: list[str], rows: list[list[Any]]) -> list[str]:
    """Monta uma tabela ORG compacta, sem depender do renderizador de paper."""
    lines = ["| " + " | ".join(_org_text(item) for item in headers) + " |"]
    lines.append("|-" + "-+-".join("-" * max(3, len(_org_text(item))) for item in headers) + "-|")
    for row in rows:
        lines.append("| " + " | ".join(_org_text(item) for item in row) + " |")
    return lines


def _org_reference(item: dict[str, Any]) -> str:
    """Formata uma referência descritiva do estudo incluído sem exigir .bib."""
    authors = _org_text(item.get("autores") or "Autor não informado")
    year = _org_text(item.get("ano") or "s.d.")
    title = _org_text(item.get("titulo") or "Título não informado")
    venue = _org_text(item.get("periodico") or "")
    doi = _doi(item.get("doi"))
    parts = [f"{authors} ({year}). {title}."]
    if venue:
        parts.append(venue + ".")
    if doi:
        parts.append(f"DOI: https://doi.org/{doi}.")
    return " ".join(parts)


def _matrix_cell(item: dict[str, Any], key: str) -> str:
    """Converte campos da planilha para células legíveis no anexo da matriz."""
    value = _org_text(item.get(key) or "")
    return value if value else "—"


def _matrix_study_label(item: dict[str, Any]) -> str:
    authors = _matrix_cell(item, "autores")
    year = _matrix_cell(item, "ano")
    title = _matrix_cell(item, "titulo")
    return f"{authors} ({year}) — {title}"


def _append_matrix_annex_org(lines: list[str], included: list[dict[str, Any]]) -> None:
    """Acrescenta uma matriz completa em duas longtables no anexo do relatório.

    A separação evita uma tabela excessivamente larga: a primeira concentra
    identificação, contexto e método; a segunda reúne achados, limitações e
    contribuição para a pergunta da revisão. As tabelas ficam em paisagem,
    podem continuar em várias páginas e repetem o cabeçalho no PDF.
    """
    lines += [
        "",
        "* Anexo A — Matriz dos estudos incluídos",
        "A matriz abaixo é a versão documental da planilha de triagem. Os campos analíticos devem ser preenchidos após a leitura do texto completo dos estudos incluídos. A versão editável permanece disponível em CSV/XLSX.",
        "",
        "** A.1 Identificação, contexto e método",
        "#+BEGIN_EXPORT latex",
        "\\begin{landscape}",
        "\\footnotesize",
        "#+END_EXPORT",
        "#+ATTR_LATEX: :environment longtable :align p{0.22\\linewidth}p{0.11\\linewidth}p{0.17\\linewidth}p{0.17\\linewidth}p{0.16\\linewidth}",
    ]
    part_one = [
        [
            _matrix_study_label(item),
            _matrix_cell(item, "pais_contexto"),
            _matrix_cell(item, "objetivo_estudo"),
            _matrix_cell(item, "desenho_metodo"),
            _matrix_cell(item, "amostra_base"),
        ]
        for item in included
        if isinstance(item, dict)
    ]
    lines.extend(_org_table(["Estudo", "Contexto/país", "Objetivo", "Desenho/método", "Amostra/base"], part_one))
    lines += [
        "#+BEGIN_EXPORT latex",
        "\\normalsize",
        "\\end{landscape}",
        "#+END_EXPORT",
        "",
        "** A.2 Achados, limitações e contribuição",
        "#+BEGIN_EXPORT latex",
        "\\begin{landscape}",
        "\\footnotesize",
        "#+END_EXPORT",
        "#+ATTR_LATEX: :environment longtable :align p{0.20\\linewidth}p{0.26\\linewidth}p{0.19\\linewidth}p{0.19\\linewidth}",
    ]
    part_two = [
        [
            _matrix_study_label(item),
            _matrix_cell(item, "achados_principais"),
            _matrix_cell(item, "limitacoes"),
            _matrix_cell(item, "contribuicao_pergunta"),
        ]
        for item in included
        if isinstance(item, dict)
    ]
    lines.extend(_org_table(["Estudo", "Achados principais", "Limitações", "Contribuição para a pergunta"], part_two))
    lines += [
        "#+BEGIN_EXPORT latex",
        "\\normalsize",
        "\\end{landscape}",
        "#+END_EXPORT",
    ]


def render_external_prisma_org_report(
    cfg: dict[str, Any],
    out_dir: Path,
    prefix: str,
    payload: dict[str, Any],
    *,
    phase: str,
) -> Path:
    """Gera ORG FGV para a busca inicial ou a consolidação após triagem.

    O relatório é deliberadamente determinístico: ele reproduz protocolo,
    fontes, critérios, contagens e decisões informadas na planilha. Nenhum
    estudo é incluído/excluído pelo gerador.
    """
    phase_norm = str(phase or "").strip().lower()
    if phase_norm not in {"preliminar", "final"}:
        raise ValueError("phase deve ser 'preliminar' ou 'final'.")
    project = _section(cfg, "projeto")
    research = _section(cfg, "pesquisa")
    report = _section(cfg, "relatorio_pesquisa")
    document = _section(cfg, "documento")
    activity = _section(cfg, "atividade")
    title = _org_text(report.get("titulo") or document.get("titulo_trabalho") or "Relatório de Pesquisa PRISMA")
    author = _org_text(document.get("autor") or activity.get("aluno") or "")
    date_value = _org_text(document.get("data") or activity.get("data") or "")
    latex_class = _org_text(document.get("classe_latex") or "fgv-paper")
    layout = _org_text(document.get("layout") or "")
    suffix = "relatorio_prisma_preliminar" if phase_norm == "preliminar" else "relatorio_prisma_final"
    target = out_dir / f"{prefix}.{suffix}.org"

    counts_raw = payload.get("contagens", {})
    counts = counts_raw if isinstance(counts_raw, dict) else {}
    criteria_inclusion = payload.get("criterios_inclusao") or report.get("criterios_inclusao") or []
    criteria_exclusion = payload.get("criterios_exclusao") or report.get("criterios_exclusao") or []
    providers = payload.get("bases") or []
    source_logs = payload.get("fontes") or []
    artifacts = payload.get("artefatos") or {}

    lines = [
        f"#+TITLE: {title}",
        f"#+AUTHOR: {author}",
        f"#+DATE: {date_value}",
        "#+LANGUAGE: pt-BR",
        "#+OPTIONS: toc:t num:t",
        f"#+LATEX_CLASS: {latex_class}",
        "#+LATEX_HEADER: \\usepackage{longtable}",
        "#+LATEX_HEADER: \\usepackage{booktabs}",
        "#+LATEX_HEADER: \\usepackage{pdflscape}",
        "#+LATEX_HEADER: \\usepackage{array}",
        "",
        "* Identificação",
        f"- Perfil: =relatorio_prisma_busca_orientada_fgv=.",
        f"- Etapa: {'busca e triagem pendente' if phase_norm == 'preliminar' else 'consolidação após triagem humana'}.",
        f"- Layout institucional: {layout or 'não informado'}.",
        f"- Projeto: {_org_text(project.get('nome') or '')}.",
        "",
        "* Questão e escopo da revisão",
        f"- Tema: {_org_text(research.get('tema') or '')}",
        f"- Recorte: {_org_text(research.get('recorte') or '')}",
        f"- Objetivo: {_org_text(research.get('objetivo') or '')}",
        f"- Pergunta de pesquisa: {_org_text(research.get('pergunta_pesquisa') or '')}",
        f"- Tipo de estudo: {_org_text(research.get('tipo_estudo') or '')}",
        "",
        "* Protocolo de busca",
        f"- Consulta geral: ={_org_text(payload.get('consulta_geral') or '')}=",
        f"- Palavras-chave: {_org_text('; '.join(str(item) for item in (payload.get('palavras_chave') or research.get('palavras_chave') or [])))}",
        f"- Idiomas: {_org_text('; '.join(str(item) for item in (research.get('idiomas') or [])))}",
        f"- Bases selecionadas: {_org_text('; '.join(PROVIDER_LABELS.get(str(item), str(item)) for item in providers))}",
        "",
        "** Critérios de inclusão",
    ]
    inclusion_values = criteria_inclusion if isinstance(criteria_inclusion, list) else [criteria_inclusion]
    lines.extend(f"- {_org_text(item)}" for item in inclusion_values if _org_text(item))
    if not any(_org_text(item) for item in inclusion_values):
        lines.append("- Não informado.")
    lines += ["", "** Critérios de exclusão"]
    exclusion_values = criteria_exclusion if isinstance(criteria_exclusion, list) else [criteria_exclusion]
    lines.extend(f"- {_org_text(item)}" for item in exclusion_values if _org_text(item))
    if not any(_org_text(item) for item in exclusion_values):
        lines.append("- Não informado.")

    lines += ["", "* Registro por fonte"]
    if isinstance(source_logs, list) and source_logs:
        rows: list[list[Any]] = []
        for source in source_logs:
            if not isinstance(source, dict):
                continue
            provider = str(source.get("provider") or "")
            status = "erro" if source.get("error") else "ok"
            rows.append([
                PROVIDER_LABELS.get(provider, provider or "fonte"),
                source.get("retrieved", 0),
                status,
            ])
        lines.extend(_org_table(["Base", "Registros recuperados", "Situação"], rows) if rows else ["- Nenhum registro de fonte disponível."])
    else:
        lines.append("- Nenhum registro de fonte disponível.")

    lines += ["", "* Fluxo de seleção"]
    if counts:
        lines.extend(_org_table(["Etapa", "Quantidade"], [[str(key).replace("_", " "), value] for key, value in counts.items()]))
    else:
        lines.append("- As contagens ainda não foram registradas.")

    if phase_norm == "preliminar":
        lines += [
            "",
            "* Situação da triagem",
            "A busca foi concluída, mas a inclusão e a exclusão de estudos permanecem pendentes de decisão humana na planilha de triagem.",
        ]
        triage_path = _org_text(artifacts.get("planilha_triagem_xlsx") or artifacts.get("planilha_triagem_csv") or "")
        if triage_path:
            lines.append(f"- Planilha de triagem: ={triage_path}=")
        lines += [
            "- Após preencher a planilha, importe-a com =--prisma-importar-triagem=. O pipeline então produzirá a versão final deste relatório no mesmo layout e com a mesma engine PDF configurada no TOML.",
            "",
            "* Nota metodológica",
            "Este documento registra a estratégia de descoberta e a situação da triagem. Ele não apresenta a busca como revisão concluída nem substitui a avaliação humana dos títulos, resumos e textos completos.",
        ]
    else:
        included = payload.get("estudos_incluidos") or []
        lines += ["", "* Estudos incluídos"]
        if isinstance(included, list) and included:
            for index, item in enumerate(included, start=1):
                if isinstance(item, dict):
                    lines.append(f"{index}. {_org_reference(item)}")
        else:
            lines.append("Nenhum estudo foi marcado como incluído. Revise a planilha de triagem antes de interpretar o relatório como concluído.")
        matrix_path = _org_text(artifacts.get("matriz_estudos_incluidos_xlsx") or artifacts.get("matriz_estudos_incluidos_csv") or "")
        if matrix_path:
            lines += ["", f"- Matriz editável (CSV/XLSX): ={matrix_path}="]
        if isinstance(included, list) and included:
            _append_matrix_annex_org(lines, [item for item in included if isinstance(item, dict)])
        lines += [
            "",
            "* Nota metodológica",
            "As contagens e os estudos incluídos foram consolidados exclusivamente a partir da planilha de triagem preenchida pela pessoa responsável. O sistema não inferiu decisões de elegibilidade.",
        ]

    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return target

# ---------------------------------------------------------------------------
# Importação de triagem manual
# ---------------------------------------------------------------------------


def _decision(value: Any) -> str:
    text = _norm(value).replace(" ", "_")
    aliases = {"sim": "INCLUIR", "s": "INCLUIR", "incluir": "INCLUIR", "incluido": "INCLUIR", "incluida": "INCLUIR", "nao": "EXCLUIR", "n": "EXCLUIR", "excluir": "EXCLUIR", "excluido": "EXCLUIR", "excluida": "EXCLUIR", "pendente": "PENDENTE", "nao_iniciado": "PENDENTE", "": "PENDENTE"}
    return aliases.get(text, str(value or "").strip().upper() or "PENDENTE")


def _read_triage(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise RuntimeError("A planilha de triagem não contém cabeçalho.")
        required = {"titulo", "decisao_titulo_resumo", "decisao_texto_completo", "incluir_final"}
        missing = sorted(required - set(reader.fieldnames))
        if missing:
            raise RuntimeError("A planilha não contém as colunas obrigatórias: " + ", ".join(missing))
        return [dict(item) for item in reader]


def _final_markdown(config: dict[str, Any], source: Path, counts: dict[str, int], included: list[dict[str, Any]]) -> str:
    lines = ["# Relatório PRISMA — consolidação após triagem humana", "", f"- Planilha importada: `{source}`", f"- Consolidação: {_now()}", f"- Meta de estudos incluídos: {config['target']}", f"- Estudos incluídos: {counts['estudos_incluidos']}", "", "## Fluxo de seleção", *[f"- {key.replace('_', ' ')}: {value}" for key, value in counts.items()], "", "## Estudos incluídos"]
    if included:
        for index, item in enumerate(included, 1):
            lines.append(f"{index}. {item.get('autores') or 'Autor não informado'} ({item.get('ano') or 's.d.'}). {item.get('titulo') or 'Título não informado'}. {item.get('periodico') or ''}. DOI: {item.get('doi') or 'não informado'}")
    else:
        lines.append("Nenhum estudo foi marcado como incluído. Revise a planilha antes de considerar a revisão concluída.")
    lines += ["", "## Observação metodológica", "As decisões foram importadas da planilha preenchida pela pessoa responsável. O sistema consolida registros e contagens; não infere inclusão ou exclusão."]
    return "\n".join(lines).rstrip() + "\n"


def import_manual_prisma_triage(cfg: dict[str, Any], out_dir: Path, prefix: str, triage_path: Path) -> dict[str, Any]:
    config = _query_config(cfg)
    source = triage_path.expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(f"Planilha de triagem não encontrada: {source}")
    rows = _read_triage(source)
    screened = [item for item in rows if _decision(item.get("decisao_titulo_resumo")) != "PENDENTE"]
    excluded_ta = [item for item in rows if _decision(item.get("decisao_titulo_resumo")) == "EXCLUIR"]
    full_assessed = [item for item in rows if _decision(item.get("decisao_texto_completo")) != "PENDENTE"]
    full_excluded = [item for item in rows if _decision(item.get("decisao_texto_completo")) == "EXCLUIR"]
    included = [item for item in rows if _decision(item.get("incluir_final")) == "INCLUIR"]
    out_dir.mkdir(parents=True, exist_ok=True)
    matrix_csv = out_dir / f"{prefix}.matriz_estudos_incluidos.csv"
    matrix_xlsx = out_dir / f"{prefix}.matriz_estudos_incluidos.xlsx"
    report_md = out_dir / f"{prefix}.relatorio_prisma_final.md"
    report_json = out_dir / f"{prefix}.prisma_report_final.json"
    # Preserva a etapa de descoberta registrada na primeira execução para que
    # o relatório consolidado mantenha o fluxo PRISMA completo.
    initial_path = out_dir / f"{prefix}.prisma_report.json"
    initial: dict[str, Any] = {}
    if initial_path.is_file():
        try:
            loaded = json.loads(initial_path.read_text(encoding="utf-8"))
            initial = loaded if isinstance(loaded, dict) else {}
        except Exception:
            initial = {}
    initial_counts = initial.get("contagens", {}) if isinstance(initial.get("contagens"), dict) else {}
    counts = dict(initial_counts)
    counts.update({"registros_na_planilha": len(rows), "registros_com_triagem_titulo_resumo_concluida": len(screened), "registros_excluidos_titulo_resumo": len(excluded_ta), "textos_completos_avaliados": len(full_assessed), "textos_completos_excluidos": len(full_excluded), "estudos_incluidos": len(included)})
    _write_csv(matrix_csv, included)
    xlsx = _write_xlsx(matrix_xlsx, included)
    report_md.write_text(_final_markdown(config, source, counts, included), encoding="utf-8")
    blank_matrix_fields = [
        field for field in MATRIX_ANALYTICAL_HEADERS
        if any(not str(item.get(field) or "").strip() for item in included)
    ]
    notices = [f"Meta configurada: {config['target']}; estudos incluídos na planilha: {len(included)}."] if len(included) != config["target"] else []
    if blank_matrix_fields:
        notices.append(
            "A matriz analítica contém campos sem preenchimento em pelo menos um estudo incluído: "
            + ", ".join(blank_matrix_fields) + "."
        )
    output = {"schema_version": "1.2", "tipo": "prisma_busca_externa_consolidado", "gerado_em": _now(), "status": "triagem_importada", "meta_estudos_incluidos": config["target"], "consulta_geral": config["query"], "palavras_chave": config["keywords"], "bases": config["providers"], "criterios_inclusao": config["inclusion"], "criterios_exclusao": config["exclusion"], "contagens": counts, "fontes": initial.get("fontes", []) if isinstance(initial.get("fontes", []), list) else [], "estudos_incluidos": included, "avisos": notices, "artefatos": {"triagem_importada": str(source), "matriz_estudos_incluidos_csv": str(matrix_csv), "relatorio_markdown": str(report_md), "prisma_report_busca": str(initial_path) if initial_path.is_file() else ""}}
    if xlsx:
        output["artefatos"]["matriz_estudos_incluidos_xlsx"] = xlsx
    _write_json(report_json, output)
    output["artefatos"]["prisma_report_json"] = str(report_json)
    return output
