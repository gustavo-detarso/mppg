#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Busca bibliográfica externa e triagem PRISMA auditável.

O módulo consulta fontes bibliográficas públicas ou autenticadas, deduplica
metadados e gera uma planilha de triagem humana. A busca não baixa PDFs nem
decide inclusão/exclusão automaticamente. As credenciais ficam somente no
arquivo ``.env`` da raiz do Academic Pipeline; não são persistidas no TOML,
nem impressas em URLs, logs ou relatórios.
"""
from __future__ import annotations

import csv
import json
import os
import re
import sys
import time
import unicodedata
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Any, Callable, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

try:  # A dependência já é usada pelo pipeline principal.
    from dotenv import load_dotenv
except Exception:  # pragma: no cover - ambiente mínimo sem python-dotenv
    load_dotenv = None  # type: ignore[assignment]

Progress = Callable[[str], None] | None

# A API do Semantic Scholar aprovada para este projeto aceita no máximo uma
# requisição por segundo, somando todos os endpoints. Mantemos margem acima de
# um segundo para não depender de arredondamentos de relógio/rede.
# SEMANTIC_SCHOLAR_RETRY_429_V2
# PRISMA_SOURCE_RESILIENCE_V1
# PRISMA_CORE_RECOVERY_V2
SEMANTIC_SCHOLAR_DEFAULT_MIN_INTERVAL = 1.05
SEMANTIC_SCHOLAR_DEFAULT_MAX_RETRIES = 4
SEMANTIC_SCHOLAR_DEFAULT_RETRY_BASE_SECONDS = 2.0
SEMANTIC_SCHOLAR_DEFAULT_RETRY_MAX_SECONDS = 45.0
_SEMANTIC_SCHOLAR_LAST_REQUEST_AT = 0.0

# Erros 5xx e 429 podem ser transitórios em APIs bibliográficas. Cada
# adaptador ativa a retentativa explicitamente; não há repetição global.
RETRYABLE_HTTP_STATUS_CODES = frozenset({429, 500, 502, 503, 504})
CROSSREF_DEFAULT_MAX_RETRIES = 3
CROSSREF_DEFAULT_RETRY_BASE_SECONDS = 1.5
CROSSREF_DEFAULT_RETRY_MAX_SECONDS = 20.0

# CORE publica um limite de cinco requisições isoladas a cada dez segundos.
# A margem de 2,15 s evita atingir exatamente o limiar, mesmo com pequenas
# variações de relógio/rede. Quando o serviço de busca permanece congestionado,
# a segunda tentativa reduz somente o tamanho da página; a expressão de busca
# e o registro da proveniência são preservados.
CORE_DEFAULT_MAX_RETRIES = 3
CORE_DEFAULT_RETRY_BASE_SECONDS = 4.0
CORE_DEFAULT_RETRY_MAX_SECONDS = 45.0
CORE_DEFAULT_MIN_INTERVAL = 2.15
CORE_DEFAULT_TIMEOUT = 35
CORE_DEFAULT_DEGRADED_LIMIT = 10
CORE_DEFAULT_DEGRADED_COOLDOWN_SECONDS = 12.0
CORE_DEFAULT_DEGRADED_MAX_RETRIES = 1
_CORE_LAST_REQUEST_AT = 0.0

# A planilha de triagem também funciona como matriz analítica após a leitura
# do texto completo. Os campos finais permanecem vazios na descoberta e são
# preenchidos apenas para os estudos efetivamente incluídos.
TRIAGE_HEADERS = [
    "id_registro", "titulo", "autores", "ano", "periodico", "doi", "url",
    "url_pdf_aberto", "resumo", "fontes", "consultas_busca", "pontuacao_relevancia",
    "status_pre_triagem_ia", "escore_aderencia_ia", "recomendacao_ia", "confianca_ia",
    "bloco_tematico_ia", "criterio_inclusao_ia", "criterio_exclusao_ia", "justificativa_ia",
    "decisao_titulo_resumo", "motivo_exclusao_titulo_resumo", "texto_completo_local",
    "decisao_texto_completo", "motivo_exclusao_texto_completo", "incluir_final", "observacoes",
    "pais_contexto", "objetivo_estudo", "desenho_metodo", "amostra_base",
    "achados_principais", "limitacoes", "contribuicao_pergunta",
]

MATRIX_ANALYTICAL_HEADERS = [
    "pais_contexto", "objetivo_estudo", "desenho_metodo", "amostra_base",
    "achados_principais", "limitacoes", "contribuicao_pergunta",
]

MATRIX_FIELD_GUIDE = [
    ("pais_contexto", "País, organização, população ou contexto institucional estudado."),
    ("objetivo_estudo", "Objetivo declarado pelo estudo incluído."),
    ("desenho_metodo", "Desenho, abordagem e método de análise."),
    ("amostra_base", "Amostra, período, base de dados ou material empírico."),
    ("achados_principais", "Principais resultados relevantes para a pergunta da revisão."),
    ("limitacoes", "Limitações metodológicas ou de escopo informadas/identificadas."),
    ("contribuicao_pergunta", "Contribuição específica do estudo para responder à pergunta de pesquisa."),
]

# Fontes de descoberta. Unpaywall é deliberadamente separado: ele não é um
# índice bibliográfico, mas um enriquecedor de links abertos por DOI.
PROVIDER_ORDER = [
    "crossref",
    "openalex",
    "semantic_scholar",
    "scopus",
    "wos",
    "pubmed",
    "europe_pmc",
    "scielo",
    "core",
]

PROVIDER_LABELS = {
    "crossref": "Crossref",
    "openalex": "OpenAlex",
    "semantic_scholar": "Semantic Scholar",
    "scopus": "Scopus",
    "wos": "Web of Science",
    "pubmed": "PubMed/NCBI",
    "europe_pmc": "Europe PMC",
    "scielo": "SciELO",
    "core": "CORE",
}

_PROVIDER_CREDENTIALS: dict[str, tuple[str, ...]] = {
    "crossref": ("CROSSREF_EMAIL",),
    "openalex": ("OPENALEX_API_KEY", "OPENALEX_EMAIL"),
    "semantic_scholar": ("SEMANTIC_SCHOLAR_API_KEY", "S2_API_KEY"),
    "scopus": ("SCOPUS_API_KEY", "SCOPUS_INSTTOKEN"),
    "wos": ("WOS_API_KEY",),
    "pubmed": ("NCBI_API_KEY", "NCBI_EMAIL"),
    "europe_pmc": ("EUROPEPMC_EMAIL",),
    "scielo": ("SCIELO_API_TOKEN",),
    "core": ("CORE_API_KEY", "CORE_EMAIL"),
    "unpaywall": ("UNPAYWALL_EMAIL",),
}

_REQUIRED_ENV: dict[str, tuple[str, ...]] = {
    "scopus": ("SCOPUS_API_KEY",),
    "wos": ("WOS_API_KEY",),
    "core": ("CORE_API_KEY",),
    "unpaywall": ("UNPAYWALL_EMAIL",),
}


# ---------------------------------------------------------------------------
# Ambiente, catálogo e utilitários
# ---------------------------------------------------------------------------


def _now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _project_root() -> Path:
    # .../app_bundle/scripts/pipeline/prisma_busca_externa.py -> raiz do projeto
    return Path(__file__).resolve().parents[3]


def load_search_environment() -> None:
    """Carrega ``.env`` da raiz sem sobrescrever variáveis já exportadas."""
    if load_dotenv is None:
        return
    env_path = _project_root() / ".env"
    if env_path.is_file():
        load_dotenv(dotenv_path=env_path, override=False)


def _has_env(name: str) -> bool:
    return bool(os.getenv(name, "").strip())


def provider_statuses() -> dict[str, dict[str, Any]]:
    """Retorna status de configuração sem revelar valores secretos."""
    load_search_environment()
    values: dict[str, dict[str, Any]] = {}
    for provider in PROVIDER_ORDER:
        names = _PROVIDER_CREDENTIALS.get(provider, ())
        detected = [name for name in names if _has_env(name)]
        required = _REQUIRED_ENV.get(provider, ())
        missing_required = [name for name in required if not _has_env(name)]
        if missing_required:
            status = "credencial obrigatória ausente"
            available = False
        elif detected:
            status = "credencial/e-mail detectado"
            available = True
        else:
            status = "acesso público; limites da fonte podem ser menores"
            available = True
        if provider == "scielo" and not detected:
            status = "acesso público; token opcional não detectado"
        values[provider] = {
            "key": provider,
            "label": PROVIDER_LABELS[provider],
            "available": available,
            "status": status,
            "credentials_detected": detected,
            "credentials_required_missing": missing_required,
        }
    values["unpaywall"] = {
        "key": "unpaywall",
        "label": "Unpaywall (enriquecimento de acesso aberto)",
        "available": _has_env("UNPAYWALL_EMAIL"),
        "status": "e-mail detectado" if _has_env("UNPAYWALL_EMAIL") else "UNPAYWALL_EMAIL ausente",
        "credentials_detected": ["UNPAYWALL_EMAIL"] if _has_env("UNPAYWALL_EMAIL") else [],
        "credentials_required_missing": [] if _has_env("UNPAYWALL_EMAIL") else ["UNPAYWALL_EMAIL"],
    }
    return values


def provider_selection_choices() -> list[tuple[str, str]]:
    """Opções para o wizard, incluindo o atalho explícito ``Todas``."""
    status = provider_statuses()
    values: list[tuple[str, str]] = [("todas", "[Todas] Selecionar todas as fontes de descoberta")]
    for provider in PROVIDER_ORDER:
        item = status[provider]
        suffix = item["status"]
        values.append((provider, f"{item['label']} — {suffix}"))
    return values


def expand_provider_selection(values: Iterable[Any]) -> list[str]:
    """Expande a escolha especial ``todas`` e normaliza aliases históricos."""
    aliases = {
        "all": "todas",
        "__todas__": "todas",
        "open_alex": "openalex",
        "cross_ref": "crossref",
        "semantic": "semantic_scholar",
        "semanticscholar": "semantic_scholar",
        "web_of_science": "wos",
        "webofscience": "wos",
        "ncbi": "pubmed",
        "europepmc": "europe_pmc",
        "europe_pmc": "europe_pmc",
    }
    requested = [aliases.get(str(item).strip().lower().replace("-", "_").replace(" ", "_"), str(item).strip().lower()) for item in values]
    if "todas" in requested:
        return list(PROVIDER_ORDER)
    return list(dict.fromkeys(item for item in requested if item in PROVIDER_ORDER))


def _section(cfg: dict[str, Any], name: str) -> dict[str, Any]:
    value = cfg.get(name, {})
    return value if isinstance(value, dict) else {}


def external_search_enabled(cfg: dict[str, Any]) -> bool:
    section = _section(cfg, "busca_prisma")
    return bool(section.get("ativo", False)) and str(section.get("modo") or "").strip().lower() == "busca_externa"


def _list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        return [str(item).strip() for item in value if str(item).strip()]
    return [item.strip() for item in re.split(r"[;,\n]+", str(value)) if item.strip()]


def _norm(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text.lower()).strip()


def _doi(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", text)
    text = re.sub(r"^doi:\s*", "", text)
    return text.rstrip(".,;)")


def _title_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _norm(value))


def _int(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _authors(value: Any) -> str:
    if isinstance(value, list):
        result: list[str] = []
        for item in value:
            if isinstance(item, dict):
                name = str(item.get("name") or item.get("display_name") or item.get("full_name") or item.get("family") or "").strip()
            else:
                name = str(item or "").strip()
            if name:
                result.append(name)
        return "; ".join(result)
    return str(value or "").strip()


def _openalex_abstract(value: Any) -> str:
    if not isinstance(value, dict):
        return ""
    parts: list[tuple[int, str]] = []
    for token, positions in value.items():
        if not isinstance(positions, list):
            continue
        for position in positions:
            try:
                parts.append((int(position), str(token)))
            except (TypeError, ValueError):
                pass
    return " ".join(token for _pos, token in sorted(parts))


def _safe_url(url: str) -> str:
    """Remove credenciais e e-mails antes de registrar qualquer URL."""
    return re.sub(
        r"([?&](?:apiKey|api_key|key|token|insttoken|mailto|email)=)[^&]+",
        r"\1REDACTED",
        str(url or ""),
        flags=re.I,
    )


def _retry_after_seconds(value: Any) -> float | None:
    """Converte Retry-After em segundos, aceitando segundos ou data HTTP."""
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return max(0.0, float(raw))
    except ValueError:
        pass
    try:
        parsed = parsedate_to_datetime(raw)
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return max(0.0, (parsed - datetime.now(timezone.utc)).total_seconds())
    except (TypeError, ValueError, IndexError, OverflowError):
        return None


def _semantic_scholar_debug(message: str) -> None:
    """Emite detalhes de espera somente quando o diagnóstico os solicitar."""
    enabled = str(os.getenv("SEMANTIC_SCHOLAR_RATE_LIMIT_DEBUG") or "").strip().lower()
    if enabled in {"1", "true", "yes", "sim", "on"}:
        print(f"[Semantic Scholar] {message}", file=sys.stderr)


def _provider_retry_debug(provider: str, message: str) -> None:
    """Emite detalhes de retentativa apenas quando explicitamente solicitado."""
    enabled = str(os.getenv("PRISMA_HTTP_RETRY_DEBUG") or "").strip().lower()
    if enabled in {"1", "true", "yes", "sim", "on"}:
        label = PROVIDER_LABELS.get(provider, provider)
        print(f"[{label}] {message}", file=sys.stderr)
    if provider == "semantic_scholar":
        _semantic_scholar_debug(message)


class ProviderUnavailableError(RuntimeError):
    """Sinaliza que uma fonte deve ser interrompida sem repetir todos os blocos."""

    def __init__(self, provider: str, message: str) -> None:
        super().__init__(message)
        self.provider = provider


def _request_text(
    url: str,
    *,
    headers: dict[str, str] | None = None,
    timeout: int = 40,
    before_attempt: Callable[[], None] | None = None,
    retryable_http_statuses: Iterable[int] = (),
    http_retries: int = 0,
    retry_base_seconds: float = 2.0,
    retry_max_seconds: float = 45.0,
    retry_provider: str = "",
) -> str:
    """Obtém texto de uma API com retentativas explicitamente opt-in.

    Retenta apenas códigos HTTP transitórios solicitados pelo adaptador, além
    de timeouts/erros de rede. ``before_attempt`` é executado em todas as
    tentativas, preservando limites de taxa como o do Semantic Scholar.
    """
    request_headers = {
        "Accept": "application/json, application/xml, text/xml, text/plain;q=0.8, */*;q=0.2",
        "User-Agent": "AcademicPipeline-PRISMA/1.1",
    }
    request_headers.update(headers or {})
    request = Request(url, headers=request_headers, method="GET")
    retries = max(0, int(http_retries))
    retry_base = max(0.25, float(retry_base_seconds))
    retry_cap = max(retry_base, float(retry_max_seconds))
    retry_codes = {int(code) for code in retryable_http_statuses}

    def wait_before_retry(reason: str, attempt: int, retry_after: float | None = None) -> None:
        backoff = min(retry_cap, retry_base * (2 ** attempt))
        delay = max(backoff, retry_after if retry_after is not None else 0.0)
        prefix = f"{reason}; " if reason else ""
        _provider_retry_debug(
            retry_provider,
            f"{prefix}aguardando {delay:.2f}s antes da tentativa {attempt + 2}/{retries + 1}.",
        )
        time.sleep(delay)

    for attempt in range(retries + 1):
        if before_attempt is not None:
            before_attempt()
        try:
            with urlopen(request, timeout=timeout) as response:  # nosec B310: endpoints are fixed APIs.
                return response.read().decode("utf-8", errors="replace")
        except HTTPError as exc:
            if exc.code in retry_codes and attempt < retries:
                retry_after = _retry_after_seconds(exc.headers.get("Retry-After"))
                try:
                    exc.close()
                except Exception:
                    pass
                wait_before_retry(f"HTTP {exc.code} recebido", attempt, retry_after)
                continue
            body = exc.read().decode("utf-8", errors="replace")[:300] if getattr(exc, "fp", None) else ""
            raise RuntimeError(f"HTTP {exc.code}: {body}") from exc
        except URLError as exc:
            if attempt < retries:
                wait_before_retry(f"Falha de rede: {exc.reason}", attempt)
                continue
            raise RuntimeError(f"Falha de rede: {exc.reason}") from exc
        except TimeoutError as exc:
            if attempt < retries:
                wait_before_retry("Tempo esgotado na consulta bibliográfica", attempt)
                continue
            raise RuntimeError("Tempo esgotado na consulta bibliográfica.") from exc

    raise RuntimeError("A consulta bibliográfica terminou sem resposta ou exceção.")


def _request_json(
    url: str,
    *,
    headers: dict[str, str] | None = None,
    timeout: int = 40,
    before_attempt: Callable[[], None] | None = None,
    retryable_http_statuses: Iterable[int] = (),
    http_retries: int = 0,
    retry_base_seconds: float = 2.0,
    retry_max_seconds: float = 45.0,
    retry_provider: str = "",
) -> Any:
    text = _request_text(
        url,
        headers=headers,
        timeout=timeout,
        before_attempt=before_attempt,
        retryable_http_statuses=retryable_http_statuses,
        http_retries=http_retries,
        retry_base_seconds=retry_base_seconds,
        retry_max_seconds=retry_max_seconds,
        retry_provider=retry_provider,
    )
    try:
        return json.loads(text)
    except json.JSONDecodeError as exc:
        raise RuntimeError("A fonte não retornou JSON válido.") from exc

def _record(source: str, source_id: Any, *, title: Any = "", authors: Any = "", year: Any = "", venue: Any = "", doi: Any = "", url: Any = "", abstract: Any = "", pdf: Any = "", citations: Any = None, kind: Any = "") -> dict[str, Any]:
    return {
        "id_registro": f"{source}:{source_id}",
        "titulo": re.sub(r"\s+", " ", str(title or "")).strip(),
        "autores": _authors(authors),
        "ano": str(year or "").strip(),
        "periodico": re.sub(r"\s+", " ", str(venue or "")).strip(),
        "doi": _doi(doi),
        "url": str(url or "").strip(),
        "url_pdf_aberto": str(pdf or "").strip(),
        "resumo": re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", str(abstract or ""))).strip(),
        "fontes": [source],
        "pontuacao_relevancia": 0.0,
        "citacoes": _int(citations),
        "tipo_publicacao": str(kind or "").strip(),
    }


def _env(name: str) -> str:
    load_search_environment()
    return os.getenv(name, "").strip()


def _first_env(*names: str) -> str:
    """Retorna a primeira variável não vazia sem expor seu conteúdo."""
    for name in names:
        value = _env(name)
        if value:
            return value
    return ""


def _semantic_scholar_min_interval(config: dict[str, Any]) -> float:
    try:
        value = float(config.get("semantic_scholar_min_interval") or SEMANTIC_SCHOLAR_DEFAULT_MIN_INTERVAL)
    except (TypeError, ValueError):
        value = SEMANTIC_SCHOLAR_DEFAULT_MIN_INTERVAL
    # Nunca permite intervalo inferior a um segundo para a chave aprovada.
    return max(1.0, min(value, 10.0))


def _semantic_scholar_max_retries(config: dict[str, Any]) -> int:
    raw = config.get("semantic_scholar_max_retries")
    try:
        value = SEMANTIC_SCHOLAR_DEFAULT_MAX_RETRIES if raw in (None, "") else int(raw)
    except (TypeError, ValueError):
        value = SEMANTIC_SCHOLAR_DEFAULT_MAX_RETRIES
    return max(0, min(value, 8))


def _semantic_scholar_retry_base_seconds(config: dict[str, Any]) -> float:
    try:
        value = float(
            config.get("semantic_scholar_retry_base_seconds")
            or SEMANTIC_SCHOLAR_DEFAULT_RETRY_BASE_SECONDS
        )
    except (TypeError, ValueError):
        value = SEMANTIC_SCHOLAR_DEFAULT_RETRY_BASE_SECONDS
    return max(0.25, min(value, 60.0))


def _semantic_scholar_retry_max_seconds(config: dict[str, Any]) -> float:
    try:
        value = float(
            config.get("semantic_scholar_retry_max_seconds")
            or SEMANTIC_SCHOLAR_DEFAULT_RETRY_MAX_SECONDS
        )
    except (TypeError, ValueError):
        value = SEMANTIC_SCHOLAR_DEFAULT_RETRY_MAX_SECONDS
    return max(_semantic_scholar_retry_base_seconds(config), min(value, 300.0))


def _bounded_int(value: Any, default: int, low: int, high: int) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        parsed = default
    return max(low, min(parsed, high))


def _bounded_float(value: Any, default: float, low: float, high: float) -> float:
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        parsed = default
    return max(low, min(parsed, high))


def _build_provider_retry_settings(search: dict[str, Any], provider: str) -> dict[str, Any]:
    """Normaliza políticas de retentativa sem exigir novas perguntas no TUI."""
    if provider == "semantic_scholar":
        return {
            "max_retries": _semantic_scholar_max_retries(search),
            "retry_base_seconds": _semantic_scholar_retry_base_seconds(search),
            "retry_max_seconds": _semantic_scholar_retry_max_seconds(search),
        }

    defaults = {
        "crossref": (
            CROSSREF_DEFAULT_MAX_RETRIES,
            CROSSREF_DEFAULT_RETRY_BASE_SECONDS,
            CROSSREF_DEFAULT_RETRY_MAX_SECONDS,
        ),
        "core": (
            CORE_DEFAULT_MAX_RETRIES,
            CORE_DEFAULT_RETRY_BASE_SECONDS,
            CORE_DEFAULT_RETRY_MAX_SECONDS,
        ),
    }
    max_default, base_default, cap_default = defaults[provider]
    prefix = provider.replace("-", "_")
    return {
        "max_retries": _bounded_int(search.get(f"{prefix}_max_retries"), max_default, 0, 8),
        "retry_base_seconds": _bounded_float(
            search.get(f"{prefix}_retry_base_seconds"),
            base_default,
            0.25,
            60.0,
        ),
        "retry_max_seconds": _bounded_float(
            search.get(f"{prefix}_retry_max_seconds"),
            cap_default,
            base_default,
            300.0,
        ),
    }


def _provider_retry_settings(config: dict[str, Any], provider: str) -> dict[str, Any]:
    configured = config.get("provider_retry_settings")
    if isinstance(configured, dict):
        value = configured.get(provider)
        if isinstance(value, dict):
            return value
    return _build_provider_retry_settings({}, provider)


def _core_min_interval(config: dict[str, Any]) -> float:
    settings = _provider_retry_settings(config, "core")
    return _bounded_float(
        settings.get("min_interval_seconds"),
        CORE_DEFAULT_MIN_INTERVAL,
        0.25,
        10.0,
    )


def _core_timeout(config: dict[str, Any]) -> int:
    settings = _provider_retry_settings(config, "core")
    return _bounded_int(settings.get("timeout_seconds"), CORE_DEFAULT_TIMEOUT, 10, 90)


def _core_degraded_limit(config: dict[str, Any]) -> int:
    settings = _provider_retry_settings(config, "core")
    return _bounded_int(
        settings.get("degraded_limit"),
        CORE_DEFAULT_DEGRADED_LIMIT,
        1,
        100,
    )


def _core_degraded_cooldown_seconds(config: dict[str, Any]) -> float:
    settings = _provider_retry_settings(config, "core")
    return _bounded_float(
        settings.get("degraded_cooldown_seconds"),
        CORE_DEFAULT_DEGRADED_COOLDOWN_SECONDS,
        1.0,
        180.0,
    )


def _core_degraded_max_retries(config: dict[str, Any]) -> int:
    settings = _provider_retry_settings(config, "core")
    return _bounded_int(
        settings.get("degraded_max_retries"),
        CORE_DEFAULT_DEGRADED_MAX_RETRIES,
        0,
        4,
    )


def _is_retryable_core_failure(exc: Exception) -> bool:
    """Distingue indisponibilidade transitória de erro de credencial/formato."""
    message = str(exc or "").casefold()
    if "tempo esgotado" in message or "falha de rede" in message:
        return True
    return any(f"http {status}" in message for status in RETRYABLE_HTTP_STATUS_CODES)


def _wait_for_core(config: dict[str, Any]) -> None:
    """Evita sequência agressiva contra a API CORE durante consultas em blocos."""
    global _CORE_LAST_REQUEST_AT
    interval = _core_min_interval(config)
    elapsed = time.monotonic() - _CORE_LAST_REQUEST_AT
    if _CORE_LAST_REQUEST_AT and elapsed < interval:
        wait = interval - elapsed
        _provider_retry_debug("core", f"aguardando {wait:.2f}s para espaçar chamadas.")
        time.sleep(wait)
    _CORE_LAST_REQUEST_AT = time.monotonic()


def _wait_for_semantic_scholar(config: dict[str, Any]) -> None:
    """Aplica a cota cumulativa de uma requisição/segundo do Semantic Scholar."""
    global _SEMANTIC_SCHOLAR_LAST_REQUEST_AT
    interval = _semantic_scholar_min_interval(config)
    elapsed = time.monotonic() - _SEMANTIC_SCHOLAR_LAST_REQUEST_AT
    if _SEMANTIC_SCHOLAR_LAST_REQUEST_AT and elapsed < interval:
        wait = interval - elapsed
        _semantic_scholar_debug(f"aguardando {wait:.2f}s para respeitar o intervalo mínimo.")
        time.sleep(wait)
    _SEMANTIC_SCHOLAR_LAST_REQUEST_AT = time.monotonic()


def _provider_email(provider: str, fallback: str = "") -> str:
    fields = {
        "crossref": ("CROSSREF_EMAIL",),
        "openalex": ("OPENALEX_EMAIL",),
        "pubmed": ("NCBI_EMAIL",),
        "europe_pmc": ("EUROPEPMC_EMAIL",),
        "core": ("CORE_EMAIL",),
    }
    for name in fields.get(provider, ()):  # prioriza e-mail específico da fonte
        value = _env(name)
        if value:
            return value
    return str(fallback or "").strip()


def _safe_strategy_id(value: Any, fallback: str) -> str:
    token = re.sub(r"[^a-z0-9_]+", "_", _norm(value).replace(" ", "_")).strip("_")
    return token or fallback


def _strategy_rows(search: dict[str, Any], *, fallback_query: str) -> list[dict[str, str]]:
    """Normaliza os blocos e executa cada consulta curta independentemente.

    ``consultas`` é uma lista dentro de cada tabela de estratégia. ``consulta``
    continua aceito para TOMLs experimentais/anteriores que usem valor único.
    """
    raw = search.get("estrategias")
    rows: list[dict[str, str]] = []
    if isinstance(raw, list):
        for block_index, item in enumerate(raw, 1):
            if not isinstance(item, dict):
                continue
            label = str(item.get("rotulo") or item.get("label") or f"Bloco temático {block_index}").strip()
            block_id = _safe_strategy_id(item.get("id") or label, f"bloco_{block_index}")
            queries = [str(value).strip() for value in _list(item.get("consultas")) if str(value).strip()]
            if not queries:
                single = str(item.get("consulta") or item.get("query") or "").strip()
                # A versão experimental inicial usava ; para separar alternativas.
                queries = [value.strip() for value in single.split(";") if value.strip()]
            for query_index, query in enumerate(queries, 1):
                ident = block_id if len(queries) == 1 else f"{block_id}_{query_index}"
                rows.append({
                    "id": ident,
                    "bloco_id": block_id,
                    "bloco_rotulo": label or f"Bloco temático {block_index}",
                    "rotulo": label or f"Bloco temático {block_index}",
                    "consulta": query,
                })
    if rows:
        seen: set[str] = set()
        normalized: list[dict[str, str]] = []
        for index, row in enumerate(rows, 1):
            ident = row["id"]
            if ident in seen:
                ident = f"{ident}_{index}"
            seen.add(ident)
            normalized.append({**row, "id": ident})
        return normalized
    return [{
        "id": "consulta_unica",
        "bloco_id": "consulta_unica",
        "bloco_rotulo": "Consulta única",
        "rotulo": "Consulta única",
        "consulta": fallback_query,
    }]


def _query_config(cfg: dict[str, Any]) -> dict[str, Any]:
    load_search_environment()
    search = _section(cfg, "busca_prisma")
    research = _section(cfg, "pesquisa")
    providers = expand_provider_selection(_list(search.get("bases")))
    if not providers:
        providers = ["crossref", "openalex", "semantic_scholar"]
    keywords = _list(research.get("palavras_chave"))
    fallback_query = str(search.get("consulta_geral") or "").strip() or " ".join(keywords).strip() or str(research.get("tema") or "").strip()
    if not fallback_query and not isinstance(search.get("estrategias"), list):
        raise RuntimeError("Defina ao menos uma estratégia de busca, consulta_geral, palavras_chave ou tema antes de executar a busca externa.")

    def clamp(value: Any, default: int, low: int, high: int) -> int:
        try:
            return max(low, min(int(value), high))
        except (TypeError, ValueError):
            return default

    strategy_rows = _strategy_rows(search, fallback_query=fallback_query)
    if not strategy_rows or not any(item["consulta"].strip() for item in strategy_rows):
        raise RuntimeError("A estratégia de busca não contém consultas utilizáveis.")
    strategy_mode = str(search.get("estrategia") or ("blocos_tematicos" if len(strategy_rows) > 1 else "consulta_unica")).strip().lower()
    fallback_email = str(search.get("email_contato") or "").strip()
    per_base = clamp(search.get("limite_por_base"), 100, 1, 200)
    return {
        "providers": providers,
        "query": fallback_query or " | ".join(item["consulta"] for item in strategy_rows),
        "query_summary": " | ".join(item["consulta"] for item in strategy_rows),
        "strategy_mode": strategy_mode,
        "query_plans": strategy_rows,
        "keywords": keywords,
        "research_theme": str(research.get("tema") or "").strip(),
        "research_scope": str(research.get("recorte") or "").strip(),
        "research_objective": str(research.get("objetivo") or "").strip(),
        "research_question": str(research.get("pergunta_pesquisa") or "").strip(),
        "per_base": per_base,
        "scopus_per_query": clamp(search.get("limite_scopus_por_consulta"), min(per_base, 10), 1, 25),
        "initial_limit": clamp(search.get("limite_triagem_inicial"), 250, 10, 1000),
        "target": clamp(search.get("meta_estudos_incluidos"), 15, 1, 500),
        "year_start": str(search.get("ano_inicio") or "").strip(),
        "year_end": str(search.get("ano_fim") or "").strip(),
        "email": fallback_email,
        "emails": {name: _provider_email(name, fallback_email) for name in ("crossref", "openalex", "pubmed", "europe_pmc", "core")},
        "inclusion": _list(search.get("criterios_inclusao") or _section(cfg, "relatorio_pesquisa").get("criterios_inclusao")),
        "exclusion": _list(search.get("criterios_exclusao") or _section(cfg, "relatorio_pesquisa").get("criterios_exclusao")),
        "corpus_local": str(search.get("corpus_local_opcional") or "").strip(),
        "enrich_unpaywall": bool(search.get("enriquecer_unpaywall", False)),
        "unpaywall_limit": clamp(search.get("limite_unpaywall"), 100, 0, 250),
        # Pré-triagem assistida por IA. Ela só prioriza a planilha: não decide
        # inclusão/exclusão e permanece desativada em TOMLs anteriores.
        "ai_screen_enabled": bool(search.get("pre_triagem_ia", False)),
        "ai_screen_model": str(search.get("pre_triagem_ia_modelo") or "").strip(),
        "ai_screen_batch_size": clamp(search.get("pre_triagem_ia_lote"), 20, 5, 30),
        "ai_screen_max_records": clamp(search.get("pre_triagem_ia_max_registros"), 1500, 10, 5000),
        "ai_screen_review_reserve": clamp(search.get("pre_triagem_ia_reserva_incertos"), 40, 0, 1000),
        "ai_screen_min_confidence": clamp(search.get("pre_triagem_ia_min_confianca"), 55, 0, 100),
        "ai_screen_max_abstract_chars": clamp(search.get("pre_triagem_ia_max_chars_resumo"), 700, 200, 1500),
        "semantic_scholar_min_interval": _semantic_scholar_min_interval(search),
        "semantic_scholar_max_retries": _semantic_scholar_max_retries(search),
        "semantic_scholar_retry_base_seconds": _semantic_scholar_retry_base_seconds(search),
        "semantic_scholar_retry_max_seconds": _semantic_scholar_retry_max_seconds(search),
        "provider_retry_settings": {
            "crossref": _build_provider_retry_settings(search, "crossref"),
            "semantic_scholar": _build_provider_retry_settings(search, "semantic_scholar"),
            "core": {
                **_build_provider_retry_settings(search, "core"),
                "min_interval_seconds": _bounded_float(
                    search.get("core_min_interval_seconds"),
                    CORE_DEFAULT_MIN_INTERVAL,
                    0.25,
                    10.0,
                ),
                "timeout_seconds": _bounded_int(
                    search.get("core_timeout_seconds"),
                    CORE_DEFAULT_TIMEOUT,
                    10,
                    90,
                ),
                "degraded_limit": _bounded_int(
                    search.get("core_degraded_limit"),
                    CORE_DEFAULT_DEGRADED_LIMIT,
                    1,
                    100,
                ),
                "degraded_cooldown_seconds": _bounded_float(
                    search.get("core_degraded_cooldown_seconds"),
                    CORE_DEFAULT_DEGRADED_COOLDOWN_SECONDS,
                    1.0,
                    180.0,
                ),
                "degraded_max_retries": _bounded_int(
                    search.get("core_degraded_max_retries"),
                    CORE_DEFAULT_DEGRADED_MAX_RETRIES,
                    0,
                    4,
                ),
            },
        },
    }


# ---------------------------------------------------------------------------
# Adaptadores das fontes
# ---------------------------------------------------------------------------


def _crossref(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    params: dict[str, Any] = {
        "query.bibliographic": query,
        "rows": config["per_base"],
        "select": "DOI,title,author,published-print,published-online,issued,container-title,abstract,type,URL",
    }
    filters: list[str] = []
    if config["year_start"]:
        filters.append(f"from-pub-date:{config['year_start']}-01-01")
    if config["year_end"]:
        filters.append(f"until-pub-date:{config['year_end']}-12-31")
    if filters:
        params["filter"] = ",".join(filters)
    if config["emails"]["crossref"]:
        params["mailto"] = config["emails"]["crossref"]
    url = "https://api.crossref.org/works?" + urlencode(params)
    retry = _provider_retry_settings(config, "crossref")
    items = (
        (_request_json(
            url,
            retryable_http_statuses=RETRYABLE_HTTP_STATUS_CODES,
            http_retries=int(retry.get("max_retries") or 0),
            retry_base_seconds=float(retry.get("retry_base_seconds") or CROSSREF_DEFAULT_RETRY_BASE_SECONDS),
            retry_max_seconds=float(retry.get("retry_max_seconds") or CROSSREF_DEFAULT_RETRY_MAX_SECONDS),
            retry_provider="crossref",
        ) or {}).get("message") or {}
    ).get("items") or []
    records: list[dict[str, Any]] = []
    for idx, item in enumerate(items, 1):
        titles = item.get("title") or []
        venues = item.get("container-title") or []
        dates = (item.get("published-print") or item.get("published-online") or item.get("issued") or {}).get("date-parts") or []
        year = dates[0][0] if dates and isinstance(dates[0], list) and dates[0] else ""
        records.append(_record("crossref", item.get("DOI") or idx, title=titles[0] if titles else "", authors=item.get("author"), year=year, venue=venues[0] if venues else "", doi=item.get("DOI"), url=item.get("URL"), abstract=item.get("abstract"), kind=item.get("type")))
    return records, {"provider": "crossref", "url": _safe_url(url), "retrieved": len(records), "error": ""}


def _openalex(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    params: dict[str, Any] = {"search": query, "per-page": config["per_base"]}
    filters: list[str] = []
    if config["year_start"]:
        filters.append(f"from_publication_date:{config['year_start']}-01-01")
    if config["year_end"]:
        filters.append(f"to_publication_date:{config['year_end']}-12-31")
    if filters:
        params["filter"] = ",".join(filters)
    if config["emails"]["openalex"]:
        params["mailto"] = config["emails"]["openalex"]
    if _env("OPENALEX_API_KEY"):
        params["api_key"] = _env("OPENALEX_API_KEY")
    url = "https://api.openalex.org/works?" + urlencode(params)
    items = (_request_json(url) or {}).get("results") or []
    records: list[dict[str, Any]] = []
    for idx, item in enumerate(items, 1):
        loc = item.get("primary_location") or {}
        source = loc.get("source") or {}
        open_access = item.get("open_access") or {}
        authors = []
        for item_author in item.get("authorships") or []:
            name = str((item_author.get("author") or {}).get("display_name") or "").strip()
            if name:
                authors.append(name)
        records.append(_record("openalex", item.get("id") or idx, title=item.get("title"), authors="; ".join(authors), year=item.get("publication_year"), venue=source.get("display_name"), doi=item.get("doi"), url=item.get("doi") or item.get("id"), abstract=_openalex_abstract(item.get("abstract_inverted_index")), pdf=loc.get("pdf_url") or open_access.get("oa_url"), citations=item.get("cited_by_count"), kind=item.get("type")))
    return records, {"provider": "openalex", "url": _safe_url(url), "retrieved": len(records), "error": ""}


def _semantic_scholar(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    params = {"query": query, "limit": config["per_base"], "fields": "title,abstract,authors,year,venue,externalIds,url,openAccessPdf,citationCount,publicationTypes"}
    url = "https://api.semanticscholar.org/graph/v1/paper/search?" + urlencode(params)
    headers: dict[str, str] = {}
    # A documentação da chave S2 exige o envio no cabeçalho x-api-key. Aceita
    # ambos os nomes usuais no .env para facilitar a migração sem expor valores.
    key = _first_env("SEMANTIC_SCHOLAR_API_KEY", "S2_API_KEY")
    if key:
        headers["x-api-key"] = key
    retry = _provider_retry_settings(config, "semantic_scholar")
    items = (
        _request_json(
            url,
            headers=headers,
            before_attempt=lambda: _wait_for_semantic_scholar(config),
            retryable_http_statuses=RETRYABLE_HTTP_STATUS_CODES,
            http_retries=int(retry.get("max_retries") or 0),
            retry_base_seconds=float(retry.get("retry_base_seconds") or SEMANTIC_SCHOLAR_DEFAULT_RETRY_BASE_SECONDS),
            retry_max_seconds=float(retry.get("retry_max_seconds") or SEMANTIC_SCHOLAR_DEFAULT_RETRY_MAX_SECONDS),
            retry_provider="semantic_scholar",
        )
        or {}
    ).get("data") or []
    records: list[dict[str, Any]] = []
    for idx, item in enumerate(items, 1):
        identifiers = item.get("externalIds") or {}
        oa = item.get("openAccessPdf") or {}
        records.append(_record("semantic_scholar", item.get("paperId") or idx, title=item.get("title"), authors=item.get("authors"), year=item.get("year"), venue=item.get("venue"), doi=identifiers.get("DOI"), url=item.get("url"), abstract=item.get("abstract"), pdf=oa.get("url"), citations=item.get("citationCount"), kind="; ".join(_list(item.get("publicationTypes")))))
    return records, {
        "provider": "semantic_scholar",
        "url": _safe_url(url),
        "retrieved": len(records),
        "error": "",
        "min_interval_seconds": _semantic_scholar_min_interval(config),
        "max_retries_on_transient_errors": int(retry.get("max_retries") or 0),
        "retry_base_seconds": float(retry.get("retry_base_seconds") or 0.0),
    }


def _scopus(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    key = _env("SCOPUS_API_KEY")
    if not key:
        raise RuntimeError("SCOPUS_API_KEY não configurada; a fonte foi ignorada.")
    terms = [part.strip().replace('"', '') for part in re.split(r"\s*;\s*|\s+OR\s+", query, flags=re.I) if part.strip()]
    scopus_query = "TITLE-ABS-KEY(" + " OR ".join(f'"{term}"' for term in terms) + ")"
    params: dict[str, Any] = {"query": scopus_query, "count": min(int(config["per_base"]), int(config.get("scopus_per_query") or 10)), "httpAccept": "application/json"}
    url = "https://api.elsevier.com/content/search/scopus?" + urlencode(params)
    headers = {"X-ELS-APIKey": key}
    if _env("SCOPUS_INSTTOKEN"):
        headers["X-ELS-Insttoken"] = _env("SCOPUS_INSTTOKEN")
    entries = ((_request_json(url, headers=headers) or {}).get("search-results") or {}).get("entry") or []
    records = [_record("scopus", item.get("dc:identifier") or idx, title=item.get("dc:title"), authors=item.get("dc:creator"), year=str(item.get("prism:coverDate") or "")[:4], venue=item.get("prism:publicationName"), doi=item.get("prism:doi"), url=item.get("prism:url"), citations=item.get("citedby-count"), kind=item.get("subtypeDescription")) for idx, item in enumerate(entries, 1)]
    return records, {"provider": "scopus", "url": _safe_url(url), "retrieved": len(records), "error": ""}


def _wos(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    key = _env("WOS_API_KEY")
    if not key:
        raise RuntimeError("WOS_API_KEY não configurada; a fonte foi ignorada.")
    # Web of Science Starter API. A URL pode ser sobrescrita no ambiente em
    # caso de plano/tenant institucional com rota própria.
    base = _env("WOS_API_BASE") or "https://api.clarivate.com/apis/wos-starter/v1/documents"
    params: dict[str, Any] = {"q": query, "limit": config["per_base"], "page": 1}
    if config["year_start"]:
        params["fromYear"] = config["year_start"]
    if config["year_end"]:
        params["toYear"] = config["year_end"]
    url = base + ("&" if "?" in base else "?") + urlencode(params)
    payload = _request_json(url, headers={"X-ApiKey": key}) or {}
    items = payload.get("hits") or payload.get("documents") or payload.get("data") or []
    records: list[dict[str, Any]] = []
    for idx, item in enumerate(items, 1):
        source = item.get("source") or {}
        names = (item.get("names") or {}).get("authors") or {}
        authors = names.get("authors") if isinstance(names, dict) else item.get("authors")
        identifiers = item.get("identifiers") or {}
        doi = identifiers.get("doi") if isinstance(identifiers, dict) else item.get("doi")
        year = source.get("publishYear") or source.get("year") or item.get("year")
        venue = source.get("title") or item.get("sourceTitle") or item.get("journal")
        records.append(_record("wos", item.get("uid") or item.get("id") or idx, title=item.get("title"), authors=authors, year=year, venue=venue, doi=doi, url=item.get("url") or item.get("links", {}).get("record"), abstract=item.get("abstract"), citations=item.get("timesCited"), kind=item.get("documentType")))
    return records, {"provider": "wos", "url": _safe_url(url), "retrieved": len(records), "error": ""}


def _pubmed(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    params: dict[str, Any] = {"db": "pubmed", "term": query, "retmax": config["per_base"], "retmode": "json", "sort": "relevance", "tool": "AcademicPipelinePRISMA"}
    if config["year_start"] or config["year_end"]:
        params["datetype"] = "pdat"
        if config["year_start"]:
            params["mindate"] = config["year_start"] + "/01/01"
        if config["year_end"]:
            params["maxdate"] = config["year_end"] + "/12/31"
    email = config["emails"]["pubmed"]
    if email:
        params["email"] = email
    if _env("NCBI_API_KEY"):
        params["api_key"] = _env("NCBI_API_KEY")
    search_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?" + urlencode(params)
    result = (_request_json(search_url) or {}).get("esearchresult") or {}
    ids = [str(item) for item in result.get("idlist") or [] if str(item)]
    if not ids:
        return [], {"provider": "pubmed", "url": _safe_url(search_url), "retrieved": 0, "error": ""}
    summary_params: dict[str, Any] = {"db": "pubmed", "id": ",".join(ids), "retmode": "json", "tool": "AcademicPipelinePRISMA"}
    if email:
        summary_params["email"] = email
    if _env("NCBI_API_KEY"):
        summary_params["api_key"] = _env("NCBI_API_KEY")
    summary_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi?" + urlencode(summary_params)
    payload = _request_json(summary_url) or {}
    summary = payload.get("result") or {}
    records: list[dict[str, Any]] = []
    for idx, uid in enumerate(ids, 1):
        item = summary.get(uid) or {}
        article_ids = item.get("articleids") or []
        doi = next((value.get("value") for value in article_ids if isinstance(value, dict) and str(value.get("idtype") or "").lower() == "doi"), "")
        authors = item.get("authors") or []
        pubdate = str(item.get("pubdate") or item.get("sortpubdate") or "")
        year_match = re.search(r"\b(19|20)\d{2}\b", pubdate)
        records.append(_record("pubmed", uid, title=item.get("title"), authors=authors, year=year_match.group(0) if year_match else "", venue=item.get("fulljournalname") or item.get("source"), doi=doi, url=f"https://pubmed.ncbi.nlm.nih.gov/{uid}/", kind=item.get("pubtype")))
    return records, {"provider": "pubmed", "url": _safe_url(search_url), "retrieved": len(records), "error": ""}


def _europe_pmc(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    params: dict[str, Any] = {"query": query, "format": "json", "resultType": "core", "pageSize": config["per_base"]}
    if config["year_start"]:
        params["query"] = f"({query}) AND FIRST_PDATE:[{config['year_start']}-01-01 TO *]"
    if config["year_end"]:
        params["query"] = f"({params['query']}) AND FIRST_PDATE:[* TO {config['year_end']}-12-31]"
    url = "https://www.ebi.ac.uk/europepmc/webservices/rest/search?" + urlencode(params)
    results = ((_request_json(url) or {}).get("resultList") or {}).get("result") or []
    records: list[dict[str, Any]] = []
    for idx, item in enumerate(results, 1):
        pdf = ""
        fulltext = item.get("fullTextUrlList") or {}
        for candidate in fulltext.get("fullTextUrl") or []:
            if isinstance(candidate, dict) and candidate.get("documentStyle") in {"pdf", "PDF"}:
                pdf = str(candidate.get("url") or "")
                break
        records.append(_record("europe_pmc", item.get("id") or item.get("pmid") or item.get("pmcid") or idx, title=item.get("title"), authors=item.get("authorString"), year=item.get("pubYear"), venue=item.get("journalTitle"), doi=item.get("doi"), url=item.get("doi") and f"https://doi.org/{item.get('doi')}" or item.get("fullTextUrl"), abstract=item.get("abstractText"), pdf=pdf, citations=item.get("citedByCount"), kind=item.get("pubType")))
    return records, {"provider": "europe_pmc", "url": _safe_url(url), "retrieved": len(records), "error": ""}


def _xml_text(node: ET.Element | None, *names: str) -> str:
    if node is None:
        return ""
    for name in names:
        found = node.find(name)
        if found is not None and found.text:
            return found.text.strip()
    return ""


def _looks_like_scielo_automation_challenge(value: Any) -> bool:
    text = str(value or "").casefold()
    markers = (
        "bunny-shield",
        "bunny shield",
        "shield-challenge",
        "establishing a secure connection",
    )
    return any(marker in text for marker in markers)


def _scielo(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    # O buscador SciELO expõe resposta summary em XML. O token, quando
    # fornecido, é encaminhado em cabeçalho sem aparecer no relatório.
    params: list[tuple[str, str]] = [
        ("q", query), ("output", "site"), ("lang", "pt"), ("from", "0"), ("sort", ""),
        ("format", "summary"), ("count", str(config["per_base"])), ("fb", ""), ("page", "1"),
        ("filter[in][]", "scl"), ("filter[type][]", "article"),
    ]
    base = _env("SCIELO_SEARCH_URL") or "https://search.scielo.org/"
    url = base + ("&" if "?" in base else "?") + urlencode(params)
    headers: dict[str, str] = {}
    token = _env("SCIELO_API_TOKEN")
    if token:
        headers["Authorization"] = f"Bearer {token}"
    try:
        raw = _request_text(url, headers=headers)
    except RuntimeError as exc:
        if _looks_like_scielo_automation_challenge(exc):
            raise ProviderUnavailableError(
                "scielo",
                "SciELO bloqueou a automação com Bunny Shield; a fonte foi interrompida após a primeira tentativa.",
            ) from exc
        raise
    if _looks_like_scielo_automation_challenge(raw):
        raise ProviderUnavailableError(
            "scielo",
            "SciELO bloqueou a automação com Bunny Shield; a fonte foi interrompida após a primeira tentativa.",
        )

    records: list[dict[str, Any]] = []
    try:
        root = ET.fromstring(raw)
        items = root.findall(".//item") or root.findall(".//doc")
        for idx, item in enumerate(items, 1):
            title = _xml_text(item, "title", "titulo")
            link = _xml_text(item, "link", "url")
            description = _xml_text(item, "description", "abstract", "resumo")
            creator = _xml_text(item, "creator", "author", "autor")
            date = _xml_text(item, "pubDate", "date", "year", "ano")
            doi_match = re.search(r"10\.\d{4,9}/[-._;()/:a-z0-9]+", raw if not link else link + " " + description, flags=re.I)
            records.append(_record("scielo", idx, title=title, authors=creator, year=(re.search(r"\b(19|20)\d{2}\b", date) or [""])[0], url=link, abstract=description, doi=doi_match.group(0) if doi_match else "", kind="article"))
    except ET.ParseError:
        # Fallback conservador para resposta HTML/sumário não XML: não tenta
        # inventar metadados; comunica a incompatibilidade de formato.
        raise RuntimeError("A busca SciELO não retornou um sumário XML compatível. Tente novamente ou configure uma rota institucional em SCIELO_SEARCH_URL.")
    return records, {"provider": "scielo", "url": _safe_url(url), "retrieved": len(records), "error": ""}


def _core(query: str, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    key = _env("CORE_API_KEY")
    if not key:
        raise RuntimeError("CORE_API_KEY não configurada; a fonte foi ignorada.")
    base = _env("CORE_API_BASE") or "https://api.core.ac.uk/v3/search/works"
    retry = _provider_retry_settings(config, "core")
    requested_limit = _bounded_int(config.get("per_base"), 20, 1, 100)

    def build_url(limit: int) -> str:
        params: dict[str, Any] = {"q": query, "limit": limit}
        if config["year_start"]:
            params["fromYear"] = config["year_start"]
        if config["year_end"]:
            params["toYear"] = config["year_end"]
        return base + ("&" if "?" in base else "?") + urlencode(params)

    def fetch(url: str, *, retries: int) -> dict[str, Any]:
        return _request_json(
            url,
            headers={"Authorization": f"Bearer {key}"},
            timeout=_core_timeout(config),
            before_attempt=lambda: _wait_for_core(config),
            retryable_http_statuses=RETRYABLE_HTTP_STATUS_CODES,
            http_retries=retries,
            retry_base_seconds=float(retry.get("retry_base_seconds") or CORE_DEFAULT_RETRY_BASE_SECONDS),
            retry_max_seconds=float(retry.get("retry_max_seconds") or CORE_DEFAULT_RETRY_MAX_SECONDS),
            retry_provider="core",
        ) or {}

    url = build_url(requested_limit)
    recovery: dict[str, Any] = {
        "applied": False,
        "requested_limit": requested_limit,
        "effective_limit": requested_limit,
    }
    try:
        payload = fetch(url, retries=int(retry.get("max_retries") or 0))
    except RuntimeError as primary_exc:
        if not _is_retryable_core_failure(primary_exc):
            raise
        degraded_limit = min(requested_limit, _core_degraded_limit(config))
        if degraded_limit >= requested_limit:
            raise
        cooldown = _core_degraded_cooldown_seconds(config)
        _provider_retry_debug(
            "core",
            "consulta principal indisponível; "
            f"aguardando {cooldown:.2f}s para recuperar com limite reduzido de {degraded_limit}.",
        )
        time.sleep(cooldown)
        recovery_url = build_url(degraded_limit)
        try:
            payload = fetch(recovery_url, retries=_core_degraded_max_retries(config))
        except RuntimeError as recovery_exc:
            raise RuntimeError(
                "CORE não respondeu após recuperação com limite reduzido "
                f"({degraded_limit} resultados): {recovery_exc}"
            ) from recovery_exc
        recovery = {
            "applied": True,
            "requested_limit": requested_limit,
            "effective_limit": degraded_limit,
            "primary_error": str(primary_exc)[:300],
            "cooldown_seconds": cooldown,
            "url": _safe_url(recovery_url),
        }
        url = recovery_url

    items = payload.get("results") or payload.get("data") or []
    records: list[dict[str, Any]] = []
    for idx, item in enumerate(items, 1):
        authors = item.get("authors") or item.get("authorsString") or []
        doi = item.get("doi") or item.get("identifiers", {}).get("doi") if isinstance(item.get("identifiers"), dict) else item.get("doi")
        records.append(_record("core", item.get("id") or idx, title=item.get("title"), authors=authors, year=item.get("publishedDate") or item.get("yearPublished"), venue=item.get("journals") or item.get("publisher"), doi=doi, url=item.get("downloadUrl") or item.get("sourceFulltextUrls") or item.get("url"), abstract=item.get("abstract"), pdf=item.get("downloadUrl"), citations=item.get("citedByCount"), kind=item.get("documentType")))
    return records, {
        "provider": "core",
        "url": _safe_url(url),
        "retrieved": len(records),
        "error": "",
        "min_interval_seconds": _core_min_interval(config),
        "max_retries_on_transient_errors": int(retry.get("max_retries") or 0),
        "recovery": recovery,
    }


# ---------------------------------------------------------------------------
# Consolidação, Unpaywall e saídas
# ---------------------------------------------------------------------------


def _merge(first: dict[str, Any], second: dict[str, Any]) -> None:
    for key in ("titulo", "autores", "ano", "periodico", "doi", "url", "url_pdf_aberto", "resumo", "tipo_publicacao"):
        if not first.get(key) and second.get(key):
            first[key] = second[key]
    first["fontes"] = sorted(set(_list(first.get("fontes")) + _list(second.get("fontes"))))
    first["consultas_busca"] = sorted(set(_list(first.get("consultas_busca")) + _list(second.get("consultas_busca"))))
    first["citacoes"] = max(_int(first.get("citacoes")) or 0, _int(second.get("citacoes")) or 0) or None


def _deduplicate(records: Iterable[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    out: list[dict[str, Any]] = []
    by_doi: dict[str, dict[str, Any]] = {}
    by_title: dict[str, dict[str, Any]] = {}
    removed = 0
    for record in records:
        doi = _doi(record.get("doi"))
        title = _title_key(record.get("titulo"))
        existing = by_doi.get(doi) if doi else by_title.get(title)
        if existing:
            _merge(existing, record)
            removed += 1
            continue
        if doi:
            by_doi[doi] = record
        if title:
            by_title[title] = record
        out.append(record)
    return out, removed


def _score(record: dict[str, Any], keywords: list[str]) -> float:
    haystack = _norm(" ".join(str(record.get(key) or "") for key in ("titulo", "resumo", "periodico")))
    if not haystack:
        return 0.0
    score = 0.0
    for keyword in keywords:
        token = _norm(keyword)
        if token and token in haystack:
            score += 1.0
    return score



_AI_RECOMMENDATIONS = {
    "PRIORIDADE_ALTA",
    "REVISAR_HUMANO",
    "PROVAVEL_EXCLUSAO",
    "INCERTO_METADADOS",
}


def _truncate_text(value: Any, limit: int) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if len(text) <= limit:
        return text
    return text[: max(0, limit - 1)].rstrip() + "…"


def _chunk_rows(rows: list[dict[str, Any]], size: int) -> list[list[dict[str, Any]]]:
    return [rows[index:index + size] for index in range(0, len(rows), size)]


def _extract_ai_content(response: Any) -> str:
    try:
        content = response.choices[0].message.content
    except Exception as exc:  # pragma: no cover - cliente remoto
        raise RuntimeError("A API de IA não devolveu conteúdo utilizável para a pré-triagem.") from exc
    if isinstance(content, list):
        content = "".join(
            str(getattr(item, "text", "") or (item.get("text", "") if isinstance(item, dict) else ""))
            for item in content
        )
    return str(content or "").strip()


def _parse_ai_json(content: Any) -> dict[str, Any]:
    text = str(content or "").strip()
    text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text, flags=re.IGNORECASE | re.DOTALL).strip()
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        raise RuntimeError("A IA devolveu pré-triagem fora do JSON solicitado.") from exc
    if not isinstance(payload, dict):
        raise RuntimeError("A IA não devolveu o objeto JSON esperado na pré-triagem.")
    return payload


def _float_between(value: Any, *, default: float = 0.0, low: float = 0.0, high: float = 100.0) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        number = default
    return max(low, min(number, high))


def _normalize_ai_recommendation(value: Any, score: float, *, metadata_poor: bool) -> str:
    token = _norm(value).replace(" ", "_")
    aliases = {
        "prioridade_alta": "PRIORIDADE_ALTA",
        "prioridade_texto_completo": "PRIORIDADE_ALTA",
        "incluir_para_texto_completo": "PRIORIDADE_ALTA",
        "alta_aderencia": "PRIORIDADE_ALTA",
        "revisar_humano": "REVISAR_HUMANO",
        "revisar": "REVISAR_HUMANO",
        "pendente": "REVISAR_HUMANO",
        "provavel_exclusao": "PROVAVEL_EXCLUSAO",
        "provavel_excluir": "PROVAVEL_EXCLUSAO",
        "baixa_aderencia": "PROVAVEL_EXCLUSAO",
        "incerto_metadados": "INCERTO_METADADOS",
        "metadados_insuficientes": "INCERTO_METADADOS",
        "incerto": "INCERTO_METADADOS",
    }
    normalized = aliases.get(token, "")
    if normalized in _AI_RECOMMENDATIONS:
        return normalized
    if metadata_poor:
        return "INCERTO_METADADOS"
    if score >= 80:
        return "PRIORIDADE_ALTA"
    if score >= 45:
        return "REVISAR_HUMANO"
    return "PROVAVEL_EXCLUSAO"


def _pretriage_prompt(config: dict[str, Any], rows: list[dict[str, Any]]) -> tuple[str, str]:
    criteria_inclusion = "\n".join(f"- {item}" for item in (config.get("inclusion") or ["Não informado."]))
    criteria_exclusion = "\n".join(f"- {item}" for item in (config.get("exclusion") or ["Não informado."]))
    blocks = sorted({str(item.get("bloco_rotulo") or "").strip() for item in config.get("query_plans", []) if str(item.get("bloco_rotulo") or "").strip()})
    block_text = "; ".join(blocks) if blocks else "Outro/indefinido"
    compact_rows = []
    for row in rows:
        compact_rows.append({
            "id_registro": str(row.get("id_registro") or ""),
            "titulo": _truncate_text(row.get("titulo"), 360),
            "resumo": _truncate_text(row.get("resumo"), int(config["ai_screen_max_abstract_chars"])),
            "periodico": _truncate_text(row.get("periodico"), 180),
            "ano": str(row.get("ano") or ""),
            "fontes": _list(row.get("fontes")),
            "consultas_busca": _list(row.get("consultas_busca")),
        })
    system = (
        "Você é um assistente metodológico de pré-triagem em revisão estruturada. "
        "Avalie exclusivamente a aderência aparente em título, resumo e metadados. "
        "Não tome decisão final de inclusão ou exclusão, não invente dados e não avalie qualidade metodológica quando ela não estiver explícita. "
        "Use `INCERTO_METADADOS` quando título/resumo forem insuficientes. "
        "Retorne apenas JSON válido no formato "
        '{"registros":[{"id_registro":"...","escore_aderencia_ia":0,"recomendacao_ia":"PRIORIDADE_ALTA|REVISAR_HUMANO|PROVAVEL_EXCLUSAO|INCERTO_METADADOS","confianca_ia":0,"bloco_tematico_ia":"...","criterio_inclusao_ia":"...","criterio_exclusao_ia":"...","justificativa_ia":"..."}]}. '
        "Cada id fornecido deve aparecer uma única vez."
    )
    user = (
        "Contexto da revisão:\n"
        f"Tema: {config.get('research_theme', '')}\n"
        f"Recorte: {config.get('research_scope', '')}\n"
        f"Objetivo: {config.get('research_objective', '')}\n"
        f"Pergunta: {config.get('research_question', '')}\n\n"
        "Critérios de inclusão:\n" + criteria_inclusion + "\n\n"
        "Critérios de exclusão:\n" + criteria_exclusion + "\n\n"
        f"Blocos temáticos possíveis: {block_text}.\n\n"
        "Registros para pré-triagem:\n" + json.dumps({"registros": compact_rows}, ensure_ascii=False)
    )
    return system, user


def _safe_ai_text(value: Any, limit: int = 900) -> str:
    return _truncate_text(value, limit).replace("\x00", "")


def _mark_pretriage_failure(record: dict[str, Any], status: str, detail: str = "") -> None:
    record.update({
        "status_pre_triagem_ia": status,
        "escore_aderencia_ia": "",
        "recomendacao_ia": "INCERTO_METADADOS",
        "confianca_ia": "",
        "bloco_tematico_ia": "",
        "criterio_inclusao_ia": "",
        "criterio_exclusao_ia": "",
        "justificativa_ia": _safe_ai_text(detail, 280),
    })


def _apply_ai_pretriage(record: dict[str, Any], value: dict[str, Any]) -> None:
    score = _float_between(value.get("escore_aderencia_ia", value.get("escore", 0)))
    confidence = _float_between(value.get("confianca_ia", value.get("confianca", 0)))
    metadata_poor = not str(record.get("titulo") or "").strip() or not str(record.get("resumo") or "").strip()
    recommendation = _normalize_ai_recommendation(
        value.get("recomendacao_ia", value.get("recomendacao", "")),
        score,
        metadata_poor=metadata_poor,
    )
    record.update({
        "status_pre_triagem_ia": "CONCLUIDA",
        "escore_aderencia_ia": round(score, 2),
        "recomendacao_ia": recommendation,
        "confianca_ia": round(confidence, 2),
        "bloco_tematico_ia": _safe_ai_text(value.get("bloco_tematico_ia", value.get("bloco_tematico", "")), 160),
        "criterio_inclusao_ia": _safe_ai_text(value.get("criterio_inclusao_ia", ""), 280),
        "criterio_exclusao_ia": _safe_ai_text(value.get("criterio_exclusao_ia", ""), 280),
        "justificativa_ia": _safe_ai_text(value.get("justificativa_ia", value.get("justificativa", "")), 900),
    })


def _pretriage_with_ai(
    records: list[dict[str, Any]],
    config: dict[str, Any],
    *,
    client: Any | None,
    model: str | None,
    progress: Progress = None,
) -> dict[str, Any]:
    """Prioriza registros antes do corte da planilha, sem decidir elegibilidade."""
    enabled = bool(config.get("ai_screen_enabled"))
    audit: dict[str, Any] = {
        "enabled": enabled,
        "decision_final": "humana_obrigatoria",
        "records_available": len(records),
        "records_evaluated": 0,
        "records_not_evaluated_limit": 0,
        "batches_planned": 0,
        "batches_completed": 0,
        "batches_failed": 0,
        "errors": [],
        "model": str(model or config.get("ai_screen_model") or ""),
    }
    if not enabled:
        for record in records:
            record.update({
                "status_pre_triagem_ia": "DESATIVADA",
                "escore_aderencia_ia": "",
                "recomendacao_ia": "",
                "confianca_ia": "",
                "bloco_tematico_ia": "",
                "criterio_inclusao_ia": "",
                "criterio_exclusao_ia": "",
                "justificativa_ia": "",
            })
        return audit
    if client is None:
        raise RuntimeError("A pré-triagem por IA está ativa, mas o cliente OpenAI não foi inicializado. Configure OPENAI_API_KEY ou desative pre_triagem_ia.")
    effective_model = str(model or config.get("ai_screen_model") or "").strip()
    if not effective_model:
        raise RuntimeError("A pré-triagem por IA está ativa, mas nenhum modelo foi definido em [openai].model, OPENAI_MODEL ou pre_triagem_ia_modelo.")
    audit["model"] = effective_model
    maximum = min(len(records), int(config["ai_screen_max_records"]))
    candidate_records = records[:maximum]
    remaining_records = records[maximum:]
    for record in remaining_records:
        _mark_pretriage_failure(record, "NAO_AVALIADA_LIMITE", "Registro preservado para revisão humana: excede o limite configurado de pré-triagem por IA.")
    audit["records_not_evaluated_limit"] = len(remaining_records)
    batches = _chunk_rows(candidate_records, int(config["ai_screen_batch_size"]))
    audit["batches_planned"] = len(batches)
    for index, batch in enumerate(batches, 1):
        if progress:
            progress(f"Pré-triagem assistida por IA ({index}/{len(batches)}): avaliando {len(batch)} registro(s)")
        system, user = _pretriage_prompt(config, batch)
        try:
            try:
                response = client.chat.completions.create(
                    model=effective_model,
                    messages=[{"role": "system", "content": system}, {"role": "user", "content": user}],
                    temperature=0.0,
                    response_format={"type": "json_object"},
                )
            except Exception:
                response = client.chat.completions.create(
                    model=effective_model,
                    messages=[{"role": "system", "content": system}, {"role": "user", "content": user}],
                    temperature=0.0,
                )
            payload = _parse_ai_json(_extract_ai_content(response))
            values = payload.get("registros", payload.get("records", payload.get("itens", [])))
            if not isinstance(values, list):
                raise RuntimeError("A IA não devolveu a lista de registros esperada.")
            by_id = {
                str(item.get("id_registro") or item.get("id") or "").strip(): item
                for item in values
                if isinstance(item, dict) and str(item.get("id_registro") or item.get("id") or "").strip()
            }
            for record in batch:
                row = by_id.get(str(record.get("id_registro") or "").strip())
                if row is None:
                    _mark_pretriage_failure(record, "FALHA_RESPOSTA", "A IA não devolveu avaliação para este registro; revisão humana obrigatória.")
                else:
                    _apply_ai_pretriage(record, row)
            audit["batches_completed"] += 1
        except Exception as exc:
            audit["batches_failed"] += 1
            audit["errors"].append(f"lote {index}: {str(exc)[:500]}")
            for record in batch:
                _mark_pretriage_failure(record, "FALHA_LOTE", "Falha na pré-triagem por IA; revisão humana obrigatória.")
        # Evita sequência agressiva contra o endpoint de IA, sem assumir um
        # limite específico além das políticas do cliente/provedor.
        time.sleep(0.05)
    audit["records_evaluated"] = sum(1 for record in candidate_records if record.get("status_pre_triagem_ia") == "CONCLUIDA")
    recommendation_counts: dict[str, int] = {}
    for record in records:
        value = str(record.get("recomendacao_ia") or "SEM_CLASSIFICACAO")
        recommendation_counts[value] = recommendation_counts.get(value, 0) + 1
    audit["recommendation_counts"] = recommendation_counts
    return audit


def _pretriage_sort_key(record: dict[str, Any]) -> tuple[int, float, float, float, str]:
    recommendation = str(record.get("recomendacao_ia") or "").strip().upper()
    priority = {
        "PRIORIDADE_ALTA": 0,
        "REVISAR_HUMANO": 1,
        "INCERTO_METADADOS": 2,
        "PROVAVEL_EXCLUSAO": 3,
    }.get(recommendation, 2)
    return (
        priority,
        -_float_between(record.get("escore_aderencia_ia"), default=0.0),
        -_float_between(record.get("confianca_ia"), default=0.0),
        -_float_between(record.get("pontuacao_relevancia"), default=0.0),
        str(record.get("titulo") or "").casefold(),
    )


def _select_triage_records(records: list[dict[str, Any]], config: dict[str, Any]) -> list[dict[str, Any]]:
    """Seleciona a planilha priorizada e preserva uma reserva de incerteza."""
    limit = int(config["initial_limit"])
    ordered = sorted(records, key=_pretriage_sort_key)
    if not config.get("ai_screen_enabled"):
        return ordered[:limit]
    reserve_size = min(int(config["ai_screen_review_reserve"]), limit)
    confidence_floor = float(config["ai_screen_min_confidence"])
    mandatory = [
        item for item in ordered
        if str(item.get("status_pre_triagem_ia") or "") != "CONCLUIDA"
        or str(item.get("recomendacao_ia") or "") == "INCERTO_METADADOS"
        or _float_between(item.get("confianca_ia"), default=0.0) < confidence_floor
    ]
    selected = mandatory[:reserve_size]
    selected_ids = {str(item.get("id_registro") or "") for item in selected}
    for item in ordered:
        if len(selected) >= limit:
            break
        ident = str(item.get("id_registro") or "")
        if ident not in selected_ids:
            selected.append(item)
            selected_ids.add(ident)
    # A reserva garante presença de itens incertos; a planilha continua ordenada
    # por aderência para que a leitura comece pelos registros mais promissores.
    return sorted(selected, key=_pretriage_sort_key)


def _enrich_with_unpaywall(records: list[dict[str, Any]], config: dict[str, Any], *, progress: Progress = None) -> dict[str, Any]:
    email = _env("UNPAYWALL_EMAIL")
    if not config.get("enrich_unpaywall"):
        return {"enabled": False, "checked": 0, "enriched": 0, "reason": "desativado"}
    if not email:
        return {"enabled": True, "checked": 0, "enriched": 0, "reason": "UNPAYWALL_EMAIL ausente"}
    candidates = [item for item in records if item.get("doi") and not item.get("url_pdf_aberto")]
    checked = 0
    enriched = 0
    for item in candidates[: int(config.get("unpaywall_limit") or 0)]:
        if progress:
            progress(f"Consultando Unpaywall para acesso aberto ({checked + 1}/{min(len(candidates), int(config.get('unpaywall_limit') or 0))})")
        url = f"https://api.unpaywall.org/v2/{_doi(item.get('doi'))}?" + urlencode({"email": email})
        try:
            payload = _request_json(url) or {}
        except Exception:
            checked += 1
            time.sleep(0.12)
            continue
        locations = [payload.get("best_oa_location"), *(payload.get("oa_locations") or [])]
        for location in locations:
            if not isinstance(location, dict):
                continue
            pdf = str(location.get("url_for_pdf") or "").strip()
            landing = str(location.get("url") or "").strip()
            if pdf or landing:
                item["url_pdf_aberto"] = pdf or landing
                enriched += 1
                break
        checked += 1
        time.sleep(0.12)
    return {"enabled": True, "checked": checked, "enriched": enriched, "reason": ""}


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _write_csv(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=TRIAGE_HEADERS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            prepared = dict(row)
            for list_field in ("fontes", "consultas_busca"):
                prepared[list_field] = "; ".join(_list(prepared.get(list_field)))
            writer.writerow({field: prepared.get(field, "") for field in TRIAGE_HEADERS})


def _write_xlsx(path: Path, rows: list[dict[str, Any]]) -> str | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        return None
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Triagem e matriz"
    sheet.append(TRIAGE_HEADERS)
    for row in rows:
        sheet.append(["; ".join(_list(row.get(field))) if field in {"fontes", "consultas_busca"} else row.get(field, "") for field in TRIAGE_HEADERS])
    sheet.freeze_panes = "A2"
    header_fill = PatternFill(fill_type="solid", fgColor="003E7E")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column in sheet.columns:
        letter = column[0].column_letter
        header = str(column[0].value or "")
        width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 48)
        if header in MATRIX_ANALYTICAL_HEADERS:
            width = max(width, 28)
        sheet.column_dimensions[letter].width = width

    guide = workbook.create_sheet("Guia da matriz")
    guide.append(["Campo", "Preenchimento esperado"])
    for field, description in MATRIX_FIELD_GUIDE:
        guide.append([field, description])
    for cell in guide[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    guide.column_dimensions["A"].width = 30
    guide.column_dimensions["B"].width = 94
    guide.freeze_panes = "A2"
    workbook.save(path)
    return str(path)


def _protocol_markdown(
    config: dict[str, Any],
    logs: list[dict[str, Any]],
    counts: dict[str, int],
    unpaywall: dict[str, Any],
    pretriage: dict[str, Any],
) -> str:
    lines = [
        "# Protocolo e registro de busca PRISMA", "",
        f"- Execução: {_now()}",
        f"- Estratégia: {config['strategy_mode']}",
        f"- Bases selecionadas: {', '.join(config['providers'])}",
        f"- Meta de estudos incluídos: {config['target']}",
        f"- Limite por bloco e base: {config['per_base']}",
        f"- Limite seguro do Scopus por bloco: {config['scopus_per_query']}",
        f"- Unpaywall: {'ativado' if unpaywall.get('enabled') else 'desativado'}; registros enriquecidos: {unpaywall.get('enriched', 0)}",
        f"- Pré-triagem por IA: {'ativada' if pretriage.get('enabled') else 'desativada'}; decisão final de elegibilidade permanece humana.",
        "",
        "## Blocos de busca",
    ]
    for plan in config["query_plans"]:
        lines.append(f"- **{plan['bloco_rotulo']}** (`{plan['bloco_id']}`): `{plan['consulta']}`")
    lines += ["", "## Critérios de inclusão", *[f"- {item}" for item in (config["inclusion"] or ["Não informado."])], "",
              "## Critérios de exclusão", *[f"- {item}" for item in (config["exclusion"] or ["Não informado."])], "",
              "## Contagens", *[f"- {name.replace('_', ' ')}: {value}" for name, value in counts.items()], "",
              "## Registro por fonte e bloco"]
    for source_log in logs:
        provider = str(source_log.get("provider") or "fonte")
        lines.append(f"- **{PROVIDER_LABELS.get(provider, provider)}** — recuperados: {source_log.get('retrieved', 0)}")
        for query_log in source_log.get("consultas", []) if isinstance(source_log.get("consultas"), list) else []:
            lines.append(f"  - {query_log.get('bloco_rotulo') or query_log.get('rotulo') or query_log.get('id') or 'consulta'} — recuperados: {query_log.get('retrieved', 0)}")
            lines.append(f"    - expressão: `{query_log.get('consulta') or ''}`")
            if query_log.get("url"):
                lines.append(f"    - URL: `{query_log['url']}`")
            if query_log.get("error"):
                lines.append(f"    - aviso/erro: {query_log['error']}")
        source_status = str(source_log.get("status") or "").strip()
        source_detail = str(source_log.get("status_detail") or "").strip()
        status_labels = {
            "nao_executada_credencial_ausente": "não executada: credencial ausente",
            "bloqueada_anti_automacao": "bloqueada por anti-automação",
            "indisponivel": "indisponível",
        }
        if source_status in status_labels:
            lines.append(f"  - situação: {status_labels[source_status]}" + (f" — {source_detail}" if source_detail else ""))
        if source_log.get("error"):
            lines.append(f"  - resumo de falhas: {source_log['error']}")
    if pretriage.get("enabled"):
        lines += [
            "",
            "## Pré-triagem assistida por IA",
            f"- Modelo: {pretriage.get('model') or 'não informado'}.",
            f"- Registros avaliados: {pretriage.get('records_evaluated', 0)} de {pretriage.get('records_available', 0)}.",
            f"- Lotes concluídos/falhos: {pretriage.get('batches_completed', 0)}/{pretriage.get('batches_failed', 0)}.",
            "- A IA apenas prioriza e justifica a ordem de revisão; ela não inclui nem exclui estudos definitivamente.",
        ]
        if pretriage.get("errors"):
            lines.append(f"- Avisos de processamento: {len(pretriage.get('errors') or [])} lote(s) exigem revisão humana obrigatória.")
    lines += ["", "## Próxima etapa obrigatória", "Revise a planilha ordenada por aderência. A decisão humana continua obrigatória para inclusão, exclusão e elegibilidade de texto completo."]
    return "\n".join(lines).rstrip() + "\n"


def _tag_query_provenance(records: list[dict[str, Any]], source: str, plan: dict[str, str]) -> list[dict[str, Any]]:
    marker = f"{source}:{plan['bloco_id']}:{plan['id']}"
    for record in records:
        record["consultas_busca"] = sorted(set(_list(record.get("consultas_busca")) + [marker]))
    return records


def _missing_required_credentials(provider: str) -> list[str]:
    return [name for name in _REQUIRED_ENV.get(provider, ()) if not _has_env(name)]


def _source_log(
    provider: str,
    query_logs: list[dict[str, Any]],
    *,
    status: str = "",
    status_detail: str = "",
) -> dict[str, Any]:
    errors = [f"{item.get('rotulo') or item.get('id')}: {item.get('error')}" for item in query_logs if item.get("error")]
    if not status:
        if errors and len(errors) >= max(1, len(query_logs)):
            status = "erro"
        elif errors:
            status = "parcial"
        else:
            status = "ok"
    return {
        "provider": provider,
        "retrieved": sum(int(item.get("retrieved") or 0) for item in query_logs),
        "error": "; ".join(errors),
        "status": status,
        "status_detail": status_detail,
        "consultas": query_logs,
    }


def run_external_prisma_search(
    cfg: dict[str, Any],
    out_dir: Path,
    prefix: str,
    *,
    progress: Progress = None,
    client: Any | None = None,
    model: str | None = None,
) -> dict[str, Any]:
    config = _query_config(cfg)
    fetchers = {
        "crossref": _crossref,
        "openalex": _openalex,
        "semantic_scholar": _semantic_scholar,
        "scopus": _scopus,
        "wos": _wos,
        "pubmed": _pubmed,
        "europe_pmc": _europe_pmc,
        "scielo": _scielo,
        "core": _core,
    }
    raw: list[dict[str, Any]] = []
    logs: list[dict[str, Any]] = []
    successes = 0
    for source in config["providers"]:
        missing_credentials = _missing_required_credentials(source)
        if missing_credentials:
            detail = "Credencial obrigatória ausente: " + ", ".join(missing_credentials) + "."
            if progress:
                progress(f"{PROVIDER_LABELS[source]} não executada — {detail}")
            logs.append(
                _source_log(
                    source,
                    [],
                    status="nao_executada_credencial_ausente",
                    status_detail=detail,
                )
            )
            continue

        source_queries: list[dict[str, Any]] = []
        source_success = False
        interrupted_status = ""
        interrupted_detail = ""
        for plan in config["query_plans"]:
            if progress:
                progress(f"Consultando {PROVIDER_LABELS[source]} — {plan['rotulo']}")
            effective_config = dict(config)
            effective_config["per_base"] = min(int(config["per_base"]), int(config["scopus_per_query"])) if source == "scopus" else int(config["per_base"])
            try:
                values, query_log = fetchers[source](plan["consulta"], effective_config)
                _tag_query_provenance(values, source, plan)
                query_log.update({"id": plan["id"], "bloco_id": plan["bloco_id"], "bloco_rotulo": plan["bloco_rotulo"], "rotulo": plan["rotulo"], "consulta": plan["consulta"]})
                raw.extend(values)
                source_queries.append(query_log)
                source_success = True
            except ProviderUnavailableError as exc:
                interrupted_status = "bloqueada_anti_automacao" if source == "scielo" else "indisponivel"
                interrupted_detail = str(exc)
                source_queries.append({
                    "id": plan["id"],
                    "bloco_id": plan["bloco_id"],
                    "bloco_rotulo": plan["bloco_rotulo"],
                    "rotulo": plan["rotulo"],
                    "consulta": plan["consulta"],
                    "url": "",
                    "retrieved": 0,
                    "error": interrupted_detail,
                })
                if progress:
                    progress(f"{PROVIDER_LABELS[source]} interrompida — {interrupted_detail}")
                break
            except Exception as exc:
                source_queries.append({"id": plan["id"], "bloco_id": plan["bloco_id"], "bloco_rotulo": plan["bloco_rotulo"], "rotulo": plan["rotulo"], "consulta": plan["consulta"], "url": "", "retrieved": 0, "error": str(exc)})
            time.sleep(0.20)

        status = interrupted_status if interrupted_status and not source_success else ""
        logs.append(_source_log(source, source_queries, status=status, status_detail=interrupted_detail))
        if source_success:
            successes += 1
    if not successes:
        errors = "; ".join(
            f"{item['provider']}: {item.get('error') or item.get('status_detail') or item.get('status') or 'sem detalhe'}"
            for item in logs
        )
        raise RuntimeError("Nenhuma fonte bibliográfica respondeu com sucesso. " + errors)
    if progress:
        progress("Deduplicando registros por DOI e título normalizado")
    deduplicated, removed = _deduplicate(raw)
    for item in deduplicated:
        item["pontuacao_relevancia"] = _score(item, config["keywords"])
    if config.get("ai_screen_enabled") and progress:
        progress("Inicializando pré-triagem assistida por IA antes do corte da planilha")
    pretriage = _pretriage_with_ai(deduplicated, config, client=client, model=model, progress=progress)
    triage = _select_triage_records(deduplicated, config)
    unpaywall = _enrich_with_unpaywall(triage, config, progress=progress)
    for item in triage:
        item.update({"decisao_titulo_resumo": "PENDENTE", "motivo_exclusao_titulo_resumo": "", "texto_completo_local": "", "decisao_texto_completo": "NAO_INICIADO", "motivo_exclusao_texto_completo": "", "incluir_final": "PENDENTE", "observacoes": ""})
    out_dir.mkdir(parents=True, exist_ok=True)
    raw_path = out_dir / f"{prefix}.candidatos_brutos.json"
    dedup_path = out_dir / f"{prefix}.candidatos_deduplicados.json"
    triage_csv = out_dir / f"{prefix}.triagem_titulo_resumo.csv"
    triage_xlsx = out_dir / f"{prefix}.triagem_titulo_resumo.xlsx"
    protocol = out_dir / f"{prefix}.protocolo_busca_prisma.md"
    log_path = out_dir / f"{prefix}.busca_prisma_log.json"
    report_path = out_dir / f"{prefix}.prisma_report.json"
    pretriage_path = out_dir / f"{prefix}.pre_triagem_ia.json"
    counts = {"registros_identificados": len(raw), "duplicatas_removidas": removed, "registros_apos_deduplicacao": len(deduplicated), "registros_pre_triagem_ia_avaliados": int(pretriage.get("records_evaluated") or 0), "registros_enviados_para_triagem": len(triage), "triagem_titulo_resumo_concluida": 0, "textos_completos_avaliados": 0, "estudos_incluidos": 0}
    _write_json(raw_path, raw)
    _write_json(dedup_path, deduplicated)
    _write_json(pretriage_path, pretriage)
    _write_csv(triage_csv, triage)
    xlsx = _write_xlsx(triage_xlsx, triage)
    search_audit = {
        "providers": config["providers"],
        "strategy_mode": config["strategy_mode"],
        "query_plans": config["query_plans"],
        "per_base": config["per_base"],
        "scopus_per_query": config["scopus_per_query"],
        "initial_limit": config["initial_limit"],
        "target": config["target"],
        "year_start": config["year_start"],
        "year_end": config["year_end"],
        "enrich_unpaywall": config["enrich_unpaywall"],
        "pre_triagem_ia": {
            "enabled": config["ai_screen_enabled"],
            "model": pretriage.get("model") or "",
            "batch_size": config["ai_screen_batch_size"],
            "max_records": config["ai_screen_max_records"],
            "review_reserve": config["ai_screen_review_reserve"],
            "min_confidence": config["ai_screen_min_confidence"],
        },
        "semantic_scholar_min_interval": config["semantic_scholar_min_interval"],
        "provider_retry_settings": config.get("provider_retry_settings", {}),
    }
    _write_json(log_path, {"gerado_em": _now(), "configuracao": search_audit, "fontes": logs, "unpaywall": unpaywall, "pre_triagem_ia": pretriage})
    protocol.write_text(_protocol_markdown(config, logs, counts, unpaywall, pretriage), encoding="utf-8")
    output = {
        "schema_version": "1.4",
        "tipo": "prisma_busca_externa",
        "gerado_em": _now(),
        "status": "triagem_titulo_resumo_pendente",
        "meta_estudos_incluidos": config["target"],
        "consulta_geral": config["query_summary"],
        "estrategia_busca": {"modo": config["strategy_mode"], "blocos": config["query_plans"]},
        "palavras_chave": config["keywords"],
        "bases": config["providers"],
        "criterios_inclusao": config["inclusion"],
        "criterios_exclusao": config["exclusion"],
        "contagens": counts,
        "fontes": logs,
        "unpaywall": unpaywall,
        "pre_triagem_ia": pretriage,
        "artefatos": {"candidatos_brutos": str(raw_path), "candidatos_deduplicados": str(dedup_path), "pre_triagem_ia": str(pretriage_path), "planilha_triagem_csv": str(triage_csv), "protocolo": str(protocol), "log": str(log_path)},
    }
    if xlsx:
        output["artefatos"]["planilha_triagem_xlsx"] = xlsx
    _write_json(report_path, output)
    output["artefatos"]["prisma_report_json"] = str(report_path)
    return output

# ---------------------------------------------------------------------------
# Relatório PRISMA em ORG (prévio e consolidado)
# Restaurado pela correção v15.2.
# ---------------------------------------------------------------------------


def _prisma_org_text(value: Any) -> str:
    """Normaliza texto de metadados para uso seguro em linhas ORG."""
    raw = "" if value is None else str(value)
    text = " ".join(raw.replace("\u00a0", " ").split())
    return text.replace("|", "/")


def _prisma_org_table(headers: list[str], rows: list[list[Any]]) -> list[str]:
    """Monta tabela ORG compacta, sem depender do renderizador do paper."""
    lines = ["| " + " | ".join(_prisma_org_text(item) for item in headers) + " |"]
    lines.append("|-" + "-+-".join("-" * max(3, len(_prisma_org_text(item))) for item in headers) + "-|")
    for row in rows:
        lines.append("| " + " | ".join(_prisma_org_text(item) for item in row) + " |")
    return lines


def _prisma_doi(value: Any) -> str:
    raw = _prisma_org_text(value).strip()
    for prefix in ("https://doi.org/", "http://doi.org/", "doi:"):
        if raw.casefold().startswith(prefix):
            raw = raw[len(prefix):]
    return raw.strip()


def _prisma_reference(item: dict[str, Any]) -> str:
    authors = _prisma_org_text(item.get("autores") or "Autor não informado")
    year = _prisma_org_text(item.get("ano") or "s.d.")
    title = _prisma_org_text(item.get("titulo") or "Título não informado")
    venue = _prisma_org_text(item.get("periodico") or "")
    doi = _prisma_doi(item.get("doi"))
    parts = [f"{authors} ({year}). {title}."]
    if venue:
        parts.append(venue + ".")
    if doi:
        parts.append(f"DOI: https://doi.org/{doi}.")
    return " ".join(parts)


def _prisma_matrix_cell(item: dict[str, Any], key: str) -> str:
    value = _prisma_org_text(item.get(key) or "")
    return value if value else "—"


def _prisma_matrix_study_label(item: dict[str, Any]) -> str:
    return (
        f"{_prisma_matrix_cell(item, 'autores')} "
        f"({_prisma_matrix_cell(item, 'ano')}) — "
        f"{_prisma_matrix_cell(item, 'titulo')}"
    )


def _prisma_append_matrix_annex_org(lines: list[str], included: list[dict[str, Any]]) -> None:
    """Acrescenta matriz documental em duas longtables no anexo do relatório."""
    lines += [
        "",
        "* Anexo A — Matriz dos estudos incluídos",
        "A matriz abaixo é a versão documental da planilha de triagem. Os campos analíticos devem ser preenchidos após a leitura do texto completo dos estudos incluídos. A versão editável permanece disponível em CSV/XLSX.",
        "",
        "** A.1 Identificação, contexto e método",
        "#+BEGIN_EXPORT latex",
        "\\begin{landscape}",
        "\\footnotesize",
        "#+END_EXPORT",
        "#+ATTR_LATEX: :environment longtable :align p{0.22\\linewidth}p{0.11\\linewidth}p{0.17\\linewidth}p{0.17\\linewidth}p{0.16\\linewidth}",
    ]
    part_one = [
        [
            _prisma_matrix_study_label(item),
            _prisma_matrix_cell(item, "pais_contexto"),
            _prisma_matrix_cell(item, "objetivo_estudo"),
            _prisma_matrix_cell(item, "desenho_metodo"),
            _prisma_matrix_cell(item, "amostra_base"),
        ]
        for item in included
        if isinstance(item, dict)
    ]
    lines.extend(
        _prisma_org_table(
            ["Estudo", "Contexto/país", "Objetivo", "Desenho/método", "Amostra/base"],
            part_one,
        )
    )
    lines += [
        "#+BEGIN_EXPORT latex",
        "\\normalsize",
        "\\end{landscape}",
        "#+END_EXPORT",
        "",
        "** A.2 Achados, limitações e contribuição",
        "#+BEGIN_EXPORT latex",
        "\\begin{landscape}",
        "\\footnotesize",
        "#+END_EXPORT",
        "#+ATTR_LATEX: :environment longtable :align p{0.20\\linewidth}p{0.26\\linewidth}p{0.19\\linewidth}p{0.19\\linewidth}",
    ]
    part_two = [
        [
            _prisma_matrix_study_label(item),
            _prisma_matrix_cell(item, "achados_principais"),
            _prisma_matrix_cell(item, "limitacoes"),
            _prisma_matrix_cell(item, "contribuicao_pergunta"),
        ]
        for item in included
        if isinstance(item, dict)
    ]
    lines.extend(
        _prisma_org_table(
            ["Estudo", "Achados principais", "Limitações", "Contribuição para a pergunta"],
            part_two,
        )
    )
    lines += [
        "#+BEGIN_EXPORT latex",
        "\\normalsize",
        "\\end{landscape}",
        "#+END_EXPORT",
    ]


def _prisma_source_summary(source: dict[str, Any]) -> tuple[Any, str]:
    """Resume logs de fonte dos formatos antigo e novo."""
    queries = source.get("consultas") or source.get("queries") or []
    if not isinstance(queries, list):
        queries = []
    retrieved = source.get("retrieved")
    if retrieved is None:
        retrieved = sum(
            int(item.get("retrieved") or 0)
            for item in queries
            if isinstance(item, dict)
        )

    explicit_status = str(source.get("status") or "").strip()
    status_labels = {
        "nao_executada_credencial_ausente": "não executada: credencial ausente",
        "bloqueada_anti_automacao": "bloqueada por anti-automação",
        "indisponivel": "indisponível",
    }
    if explicit_status in status_labels:
        return retrieved, status_labels[explicit_status]
    if explicit_status in {"ok", "parcial", "erro"}:
        return retrieved, explicit_status

    errors = []
    top_error = _prisma_org_text(source.get("error") or "")
    if top_error:
        errors.append(top_error)
    for item in queries:
        if isinstance(item, dict) and _prisma_org_text(item.get("error") or ""):
            errors.append(_prisma_org_text(item.get("error")))
    if errors and len(errors) >= max(1, len(queries)):
        return retrieved, "erro"
    if errors:
        return retrieved, "parcial"
    return retrieved, "ok"


def _prisma_strategy_org_lines(payload: dict[str, Any]) -> list[str]:
    strategy = payload.get("estrategia_busca")
    if not isinstance(strategy, dict):
        query = _prisma_org_text(payload.get("consulta_geral") or "")
        return [f"- Consulta geral: ={query}="]
    mode = _prisma_org_text(strategy.get("modo") or "consulta_unica")
    rows = strategy.get("blocos") or []
    lines = [f"- Estratégia: ={mode}=."]
    if not isinstance(rows, list) or not rows:
        query = _prisma_org_text(payload.get("consulta_geral") or "")
        lines.append(f"- Consulta geral: ={query}=")
        return lines
    lines.append("** Blocos e consultas executadas")
    for row in rows:
        if not isinstance(row, dict):
            continue
        label = _prisma_org_text(row.get("bloco_rotulo") or row.get("rotulo") or row.get("id") or "Bloco")
        query = _prisma_org_text(row.get("consulta") or "")
        if query:
            lines.append(f"- {label}: ={query}=")
    return lines


def render_external_prisma_org_report(
    cfg: dict[str, Any],
    out_dir: Path,
    prefix: str,
    payload: dict[str, Any],
    *,
    phase: str,
) -> Path:
    """Gera ORG FGV para a busca inicial ou a consolidação após triagem."""
    phase_norm = str(phase or "").strip().lower()
    if phase_norm not in {"preliminar", "final"}:
        raise ValueError("phase deve ser 'preliminar' ou 'final'.")

    def section(name: str) -> dict[str, Any]:
        value = cfg.get(name, {})
        return value if isinstance(value, dict) else {}

    project = section("projeto")
    research = section("pesquisa")
    report = section("relatorio_pesquisa")
    document = section("documento")
    activity = section("atividade")
    title = _prisma_org_text(report.get("titulo") or document.get("titulo_trabalho") or "Relatório de Pesquisa PRISMA")
    author = _prisma_org_text(document.get("autor") or activity.get("aluno") or "")
    date_value = _prisma_org_text(document.get("data") or activity.get("data") or "")
    latex_class = _prisma_org_text(document.get("classe_latex") or "fgv-paper")
    layout = _prisma_org_text(document.get("layout") or "")
    suffix = "relatorio_prisma_preliminar" if phase_norm == "preliminar" else "relatorio_prisma_final"
    target = out_dir / f"{prefix}.{suffix}.org"

    counts = payload.get("contagens") if isinstance(payload.get("contagens"), dict) else {}
    criteria_inclusion = payload.get("criterios_inclusao") or report.get("criterios_inclusao") or []
    criteria_exclusion = payload.get("criterios_exclusao") or report.get("criterios_exclusao") or []
    providers = payload.get("bases") or []
    source_logs = payload.get("fontes") or []
    artifacts = payload.get("artefatos") if isinstance(payload.get("artefatos"), dict) else {}
    labels = globals().get("PROVIDER_LABELS", {})

    lines = [
        f"#+TITLE: {title}",
        f"#+AUTHOR: {author}",
        f"#+DATE: {date_value}",
        "#+LANGUAGE: pt-BR",
        "#+OPTIONS: toc:t num:t",
        f"#+LATEX_CLASS: {latex_class}",
        "#+LATEX_HEADER: \\usepackage{longtable}",
        "#+LATEX_HEADER: \\usepackage{booktabs}",
        "#+LATEX_HEADER: \\usepackage{pdflscape}",
        "#+LATEX_HEADER: \\usepackage{array}",
        "",
        "* Identificação",
        "- Perfil: =relatorio_prisma_busca_orientada_fgv=.",
        f"- Etapa: {'busca e triagem pendente' if phase_norm == 'preliminar' else 'consolidação após triagem humana'}.",
        f"- Layout institucional: {layout or 'não informado'}.",
        f"- Projeto: {_prisma_org_text(project.get('nome') or '')}.",
        "",
        "* Questão e escopo da revisão",
        f"- Tema: {_prisma_org_text(research.get('tema') or '')}",
        f"- Recorte: {_prisma_org_text(research.get('recorte') or '')}",
        f"- Objetivo: {_prisma_org_text(research.get('objetivo') or '')}",
        f"- Pergunta de pesquisa: {_prisma_org_text(research.get('pergunta_pesquisa') or '')}",
        f"- Tipo de estudo: {_prisma_org_text(research.get('tipo_estudo') or '')}",
        "",
        "* Protocolo de busca",
    ]
    lines.extend(_prisma_strategy_org_lines(payload))
    keywords = payload.get("palavras_chave") or research.get("palavras_chave") or []
    languages = research.get("idiomas") or []
    provider_labels = [labels.get(str(item), str(item)) for item in providers]
    lines += [
        f"- Palavras-chave: {_prisma_org_text('; '.join(str(item) for item in keywords))}",
        f"- Idiomas: {_prisma_org_text('; '.join(str(item) for item in languages))}",
        f"- Bases selecionadas: {_prisma_org_text('; '.join(provider_labels))}",
        "",
        "** Critérios de inclusão",
    ]
    inclusion_values = criteria_inclusion if isinstance(criteria_inclusion, list) else [criteria_inclusion]
    lines.extend(f"- {_prisma_org_text(item)}" for item in inclusion_values if _prisma_org_text(item))
    if not any(_prisma_org_text(item) for item in inclusion_values):
        lines.append("- Não informado.")
    lines += ["", "** Critérios de exclusão"]
    exclusion_values = criteria_exclusion if isinstance(criteria_exclusion, list) else [criteria_exclusion]
    lines.extend(f"- {_prisma_org_text(item)}" for item in exclusion_values if _prisma_org_text(item))
    if not any(_prisma_org_text(item) for item in exclusion_values):
        lines.append("- Não informado.")

    lines += ["", "* Registro por fonte"]
    if isinstance(source_logs, list) and source_logs:
        rows: list[list[Any]] = []
        for source in source_logs:
            if not isinstance(source, dict):
                continue
            provider = str(source.get("provider") or "")
            retrieved, status = _prisma_source_summary(source)
            rows.append([labels.get(provider, provider or "fonte"), retrieved, status])
        lines.extend(_prisma_org_table(["Base", "Registros recuperados", "Situação"], rows) if rows else ["- Nenhum registro de fonte disponível."])
    else:
        lines.append("- Nenhum registro de fonte disponível.")

    lines += ["", "* Fluxo de seleção"]
    if counts:
        rows = [[str(key).replace("_", " "), value] for key, value in counts.items()]
        lines.extend(_prisma_org_table(["Etapa", "Quantidade"], rows))
    else:
        lines.append("- As contagens ainda não foram registradas.")

    if phase_norm == "preliminar":
        lines += [
            "",
            "* Situação da triagem",
            "A busca foi concluída, mas a inclusão e a exclusão de estudos permanecem pendentes de decisão humana na planilha de triagem.",
        ]
        pretriage_payload = payload.get("pre_triagem_ia") if isinstance(payload.get("pre_triagem_ia"), dict) else {}
        if pretriage_payload.get("enabled"):
            lines += [
                "- A planilha foi ordenada por pré-triagem assistida por IA antes do limite inicial de registros.",
                "- A recomendação e a justificativa da IA servem apenas para priorização; elas não substituem a decisão humana.",
            ]
        triage_path = _prisma_org_text(artifacts.get("planilha_triagem_xlsx") or artifacts.get("planilha_triagem_csv") or "")
        if triage_path:
            lines.append(f"- Planilha de triagem: ={triage_path}=")
        lines += [
            "- Após preencher a planilha, importe-a com =--prisma-importar-triagem=. O pipeline então produzirá a versão final deste relatório no mesmo layout e com a mesma engine PDF configurada no TOML.",
            "",
            "* Nota metodológica",
            "Este documento registra a estratégia de descoberta e a situação da triagem. Ele não apresenta a busca como revisão concluída nem substitui a avaliação humana dos títulos, resumos e textos completos.",
        ]
    else:
        included = payload.get("estudos_incluidos") or []
        lines += ["", "* Estudos incluídos"]
        if isinstance(included, list) and included:
            for index, item in enumerate(included, start=1):
                if isinstance(item, dict):
                    lines.append(f"{index}. {_prisma_reference(item)}")
        else:
            lines.append("Nenhum estudo foi marcado como incluído. Revise a planilha de triagem antes de interpretar o relatório como concluído.")
        matrix_path = _prisma_org_text(artifacts.get("matriz_estudos_incluidos_xlsx") or artifacts.get("matriz_estudos_incluidos_csv") or "")
        if matrix_path:
            lines += ["", f"- Matriz editável (CSV/XLSX): ={matrix_path}="]
        if isinstance(included, list) and included:
            _prisma_append_matrix_annex_org(lines, [item for item in included if isinstance(item, dict)])
        lines += [
            "",
            "* Nota metodológica",
            "As contagens e os estudos incluídos foram consolidados exclusivamente a partir da planilha de triagem preenchida pela pessoa responsável. O sistema não inferiu decisões de elegibilidade.",
        ]

    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return target


def _decision(value: Any) -> str:
    text = _norm(value).replace(" ", "_")
    aliases = {"sim": "INCLUIR", "s": "INCLUIR", "incluir": "INCLUIR", "incluido": "INCLUIR", "incluida": "INCLUIR", "nao": "EXCLUIR", "n": "EXCLUIR", "excluir": "EXCLUIR", "excluido": "EXCLUIR", "excluida": "EXCLUIR", "pendente": "PENDENTE", "nao_iniciado": "PENDENTE", "": "PENDENTE"}
    return aliases.get(text, str(value or "").strip().upper() or "PENDENTE")


def _read_triage(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise RuntimeError("A planilha de triagem não contém cabeçalho.")
        required = {"titulo", "decisao_titulo_resumo", "decisao_texto_completo", "incluir_final"}
        missing = sorted(required - set(reader.fieldnames))
        if missing:
            raise RuntimeError("A planilha não contém as colunas obrigatórias: " + ", ".join(missing))
        return [dict(item) for item in reader]


def _final_markdown(config: dict[str, Any], source: Path, counts: dict[str, int], included: list[dict[str, Any]]) -> str:
    lines = ["# Relatório PRISMA — consolidação após triagem humana", "", f"- Planilha importada: `{source}`", f"- Consolidação: {_now()}", f"- Meta de estudos incluídos: {config['target']}", f"- Estudos incluídos: {counts['estudos_incluidos']}", "", "## Fluxo de seleção", *[f"- {key.replace('_', ' ')}: {value}" for key, value in counts.items()], "", "## Estudos incluídos"]
    if included:
        for index, item in enumerate(included, 1):
            lines.append(f"{index}. {item.get('autores') or 'Autor não informado'} ({item.get('ano') or 's.d.'}). {item.get('titulo') or 'Título não informado'}. {item.get('periodico') or ''}. DOI: {item.get('doi') or 'não informado'}")
    else:
        lines.append("Nenhum estudo foi marcado como incluído. Revise a planilha antes de considerar a revisão concluída.")
    lines += ["", "## Observação metodológica", "As decisões foram importadas da planilha preenchida pela pessoa responsável. O sistema consolida registros e contagens; não infere inclusão ou exclusão."]
    return "\n".join(lines).rstrip() + "\n"


def import_manual_prisma_triage(cfg: dict[str, Any], out_dir: Path, prefix: str, triage_path: Path) -> dict[str, Any]:
    config = _query_config(cfg)
    source = triage_path.expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(f"Planilha de triagem não encontrada: {source}")
    rows = _read_triage(source)
    screened = [item for item in rows if _decision(item.get("decisao_titulo_resumo")) != "PENDENTE"]
    excluded_ta = [item for item in rows if _decision(item.get("decisao_titulo_resumo")) == "EXCLUIR"]
    full_assessed = [item for item in rows if _decision(item.get("decisao_texto_completo")) != "PENDENTE"]
    full_excluded = [item for item in rows if _decision(item.get("decisao_texto_completo")) == "EXCLUIR"]
    included = [item for item in rows if _decision(item.get("incluir_final")) == "INCLUIR"]
    out_dir.mkdir(parents=True, exist_ok=True)
    matrix_csv = out_dir / f"{prefix}.matriz_estudos_incluidos.csv"
    matrix_xlsx = out_dir / f"{prefix}.matriz_estudos_incluidos.xlsx"
    report_md = out_dir / f"{prefix}.relatorio_prisma_final.md"
    report_json = out_dir / f"{prefix}.prisma_report_final.json"
    # Preserva a etapa de descoberta registrada na primeira execução para que
    # o relatório consolidado mantenha o fluxo PRISMA completo.
    initial_path = out_dir / f"{prefix}.prisma_report.json"
    initial: dict[str, Any] = {}
    if initial_path.is_file():
        try:
            loaded = json.loads(initial_path.read_text(encoding="utf-8"))
            initial = loaded if isinstance(loaded, dict) else {}
        except Exception:
            initial = {}
    initial_counts = initial.get("contagens", {}) if isinstance(initial.get("contagens"), dict) else {}
    counts = dict(initial_counts)
    counts.update({"registros_na_planilha": len(rows), "registros_com_triagem_titulo_resumo_concluida": len(screened), "registros_excluidos_titulo_resumo": len(excluded_ta), "textos_completos_avaliados": len(full_assessed), "textos_completos_excluidos": len(full_excluded), "estudos_incluidos": len(included)})
    _write_csv(matrix_csv, included)
    xlsx = _write_xlsx(matrix_xlsx, included)
    report_md.write_text(_final_markdown(config, source, counts, included), encoding="utf-8")
    blank_matrix_fields = [
        field for field in MATRIX_ANALYTICAL_HEADERS
        if any(not str(item.get(field) or "").strip() for item in included)
    ]
    notices = [f"Meta configurada: {config['target']}; estudos incluídos na planilha: {len(included)}."] if len(included) != config["target"] else []
    if blank_matrix_fields:
        notices.append(
            "A matriz analítica contém campos sem preenchimento em pelo menos um estudo incluído: "
            + ", ".join(blank_matrix_fields) + "."
        )
    output = {"schema_version": "1.3", "tipo": "prisma_busca_externa_consolidado", "gerado_em": _now(), "status": "triagem_importada", "meta_estudos_incluidos": config["target"], "consulta_geral": config["query_summary"], "estrategia_busca": {"modo": config["strategy_mode"], "blocos": config["query_plans"]}, "palavras_chave": config["keywords"], "bases": config["providers"], "criterios_inclusao": config["inclusion"], "criterios_exclusao": config["exclusion"], "contagens": counts, "fontes": initial.get("fontes", []) if isinstance(initial.get("fontes", []), list) else [], "estudos_incluidos": included, "avisos": notices, "artefatos": {"triagem_importada": str(source), "matriz_estudos_incluidos_csv": str(matrix_csv), "relatorio_markdown": str(report_md), "prisma_report_busca": str(initial_path) if initial_path.is_file() else ""}}
    if xlsx:
        output["artefatos"]["matriz_estudos_incluidos_xlsx"] = xlsx
    _write_json(report_json, output)
    output["artefatos"]["prisma_report_json"] = str(report_json)
    return output
