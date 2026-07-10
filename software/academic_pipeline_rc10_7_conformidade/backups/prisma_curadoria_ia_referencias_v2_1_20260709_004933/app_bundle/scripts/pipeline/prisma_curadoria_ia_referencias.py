#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import re
import subprocess
import sys
import time
import unicodedata
import urllib.request
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path


NS = {
    "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}

PREFIX = "relatorio_prisma_prisma_fluxo_pmf"
DEFAULT_OUT = Path("app_bundle/projetos/prisma_fluxo_pmf/output_pesquisa/relatorio_prisma_prisma_fluxo_pmf")

DEFAULT_TEMA = (
    "Redesenho do fluxo de análise dos benefícios por incapacidade na Perícia Médica Federal, "
    "com foco em IA, automação documental, teleperícia, decisão baseada em evidências, "
    "auditabilidade, qualidade decisória, fila e equidade."
)

DEFAULT_WEIGHTS = {
    "ia_automacao": 35,
    "telepericia_avaliacao_remota": 25,
    "decisao_baseada_em_evidencias": 20,
    "auditoria_qualidade_integridade": 10,
    "fila_capacidade_equidade": 5,
    "avaliacao_medico_pericial": 5,
}

DEFAULT_MINIMUMS = {
    "ia_automacao": 6,
    "telepericia_avaliacao_remota": 4,
    "decisao_baseada_em_evidencias": 5,
    "auditoria_qualidade_integridade": 3,
    "fila_capacidade_equidade": 2,
}

CATEGORY_LABELS = {
    "ia_automacao": "IA e automação documental",
    "telepericia_avaliacao_remota": "Teleperícia e avaliação remota",
    "decisao_baseada_em_evidencias": "Decisão baseada em evidências",
    "auditoria_qualidade_integridade": "Auditoria, qualidade e integridade",
    "fila_capacidade_equidade": "Fila, capacidade e equidade",
    "avaliacao_medico_pericial": "Avaliação médico-pericial e benefícios",
}

TERMS = {
    "ia_automacao": [
        "artificial intelligence", "inteligencia artificial", "inteligência artificial", "machine learning",
        "deep learning", "algorithm", "algoritmo", "automated", "automation", "automacao", "automação",
        "decision support", "clinical decision support", "support system", "natural language processing",
        "nlp", "predictive", "prediction", "classifier", "classification", "data mining",
        "digital triage", "triagem automatizada", "automated decision", "ai", "model",
    ],
    "telepericia_avaliacao_remota": [
        "telehealth", "telemedicine", "telemedicina", "telepericia", "teleperícia", "remote assessment",
        "remote medical assessment", "avaliacao remota", "avaliação remota", "teleassessment",
        "video", "videoconference", "virtual assessment", "digital certificate", "digitization",
        "digitalization", "digital certification", "certificacao digital", "certificação digital",
    ],
    "decisao_baseada_em_evidencias": [
        "evidence", "evidence-based", "baseada em evidencias", "baseada em evidências", "guideline",
        "protocol", "protocolo", "structured", "instrument", "scale", "validation", "validacao", "validação",
        "reliability", "interrater", "inter-rater", "icf", "cif", "international classification of functioning",
        "criteria", "criterios", "critérios", "work capacity", "functional capacity", "functioning",
        "medical evidence", "medical certificate", "medical certification",
    ],
    "auditoria_qualidade_integridade": [
        "audit", "auditoria", "quality", "quality assurance", "integrity", "integridade", "fraud",
        "fraude", "risk", "risco", "transparency", "transparencia", "transparência", "explainability",
        "explainable", "governance", "governanca", "governança", "consistency", "confiabilidade",
        "reliability", "review", "appeal", "appeals", "controle", "control",
    ],
    "fila_capacidade_equidade": [
        "waiting time", "wait time", "waiting", "fila", "queue", "capacity", "capacidade",
        "allocation", "alocacao", "alocação", "prioritization", "priorizacao", "priorização",
        "triage", "triagem", "access", "acesso", "equity", "equidade", "regional", "territorial",
        "application costs", "screened out",
    ],
    "avaliacao_medico_pericial": [
        "incapacity benefit", "disability benefit", "disability benefits", "sickness benefit",
        "social security disability", "beneficio por incapacidade", "benefício por incapacidade",
        "work disability", "medical assessment", "medical evaluation", "medical inspection",
        "pericia", "perícia", "inss", "bpc", "eligibility", "elegibilidade", "claim",
    ],
}


def log(kind: str, msg: str) -> None:
    print(f"[{kind}] {msg}")


def load_dotenv(path: Path = Path(".env")) -> None:
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


def norm_text(s: object) -> str:
    s = str(s or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", s).strip()


def norm_key(s: object) -> str:
    return norm_text(s).lower()


def clean(s: object, max_len: int | None = None) -> str:
    t = re.sub(r"\s+", " ", str(s or "")).strip()
    if max_len and len(t) > max_len:
        return t[: max_len - 1].rstrip() + "…"
    return t


def col_idx(ref: str) -> int:
    m = re.match(r"([A-Z]+)", ref or "")
    if not m:
        return 0
    n = 0
    for ch in m.group(1):
        n = n * 26 + ord(ch) - 64
    return n - 1


def read_xlsx(path: Path, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, str]], str]:
    with zipfile.ZipFile(path) as z:
        shared = []
        if "xl/sharedStrings.xml" in z.namelist():
            root = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for si in root.findall("a:si", NS):
                shared.append("".join(t.text or "" for t in si.findall(".//a:t", NS)))

        workbook = ET.fromstring(z.read("xl/workbook.xml"))
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels.findall("rel:Relationship", NS)}

        def resolve_target(target: str) -> str:
            target = (target or "").replace("\\", "/").lstrip("/")
            if target.startswith("xl/"):
                return target
            if target.startswith("../"):
                target = target[3:]
            return "xl/" + target

        target_xml = None
        chosen_sheet = None
        available = []
        for sheet in workbook.findall("a:sheets/a:sheet", NS):
            name = sheet.attrib.get("name", "")
            rid = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            available.append(name)
            if sheet_name is None or name == sheet_name:
                target_xml = resolve_target(rel_map.get(rid, ""))
                chosen_sheet = name
                if sheet_name is not None:
                    break

        if target_xml is None:
            raise SystemExit(f"Aba '{sheet_name}' não encontrada em {path}. Abas disponíveis: {available}")

        root = ET.fromstring(z.read(target_xml))
        rows = []
        for r in root.findall("a:sheetData/a:row", NS):
            vals = []
            for c in r.findall("a:c", NS):
                idx = col_idx(c.attrib.get("r", "A1"))
                while len(vals) <= idx:
                    vals.append("")
                t = c.attrib.get("t")
                if t == "inlineStr":
                    val = "".join(x.text or "" for x in c.findall(".//a:t", NS))
                else:
                    v = c.find("a:v", NS)
                    val = "" if v is None else (v.text or "")
                    if t == "s" and val:
                        val = shared[int(val)]
                vals[idx] = val
            rows.append(vals)

    if not rows:
        return [], [], chosen_sheet or ""

    headers = [str(h or "").strip() for h in rows[0]]
    data = []
    for row in rows[1:]:
        row = row + [""] * (len(headers) - len(row))
        if not any(str(x).strip() for x in row):
            continue
        data.append({headers[i]: str(row[i] if i < len(row) else "") for i in range(len(headers))})
    return headers, data, chosen_sheet or ""


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]], str]:
    for enc in ("utf-8-sig", "utf-8", "latin1"):
        try:
            with path.open("r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                return list(reader.fieldnames or []), [dict(r) for r in reader], "csv"
        except UnicodeDecodeError:
            continue
    raise SystemExit(f"Não consegui ler CSV: {path}")


def load_input(path: Path) -> tuple[list[str], list[dict[str, str]], str]:
    if path.suffix.lower() == ".xlsx":
        best_h, best_r, best_sheet = [], [], ""
        for sheet in ("Triagem completa", "Triagem e matriz", "Referências incluídas", None):
            try:
                h, r, s = read_xlsx(path, sheet)
            except Exception:
                continue
            if len(r) > len(best_r):
                best_h, best_r, best_sheet = h, r, s
            if sheet in ("Triagem completa", "Triagem e matriz") and len(r) >= 50:
                return h, r, f"{path} :: aba {s}"
        return best_h, best_r, f"{path} :: aba {best_sheet}"
    h, r, s = read_csv(path)
    return h, r, f"{path} :: {s}"


def find_default_input(out_dir: Path) -> Path:
    for p in [
        out_dir / f"{PREFIX}.triagem_titulo_resumo.xlsx",
        out_dir / f"{PREFIX}.triagem_titulo_resumo.csv",
        out_dir / f"{PREFIX}.curadoria_ia_referencias.xlsx",
        Path(f"{PREFIX}.triagem_titulo_resumo.xlsx"),
        Path(f"{PREFIX}.triagem_titulo_resumo.csv"),
    ]:
        if p.exists():
            return p
    raise SystemExit("Não encontrei triagem_titulo_resumo.xlsx/csv. Informe --input.")


def parse_number(text: str, key: str, default: int) -> int:
    m = re.search(rf"(?m)^\s*{re.escape(key)}\s*:\s*([0-9]+)\s*(?:#.*)?$", text)
    return int(m.group(1)) if m else default


def parse_section_numbers(text: str, section: str, defaults: dict[str, int]) -> dict[str, int]:
    out = dict(defaults)
    m = re.search(rf"(?ms)^\s*{re.escape(section)}\s*:\s*\n(?P<body>(?:\s{{2,}}[^\n]*\n?)+)", text)
    if not m:
        return out
    body = m.group("body")
    for raw in body.splitlines():
        line = raw.split("#", 1)[0].strip()
        if not line or ":" not in line:
            continue
        k, v = line.split(":", 1)
        k = k.strip()
        v = v.strip()
        if re.match(r"^-?\d+$", v):
            out[k] = int(v)
    return out


def parse_block_scalar(text: str, key: str, default: str = "") -> str:
    m = re.search(rf"(?ms)^\s*{re.escape(key)}\s*:\s*[>|]\s*\n(?P<body>(?:\s{{2,}}[^\n]*\n?)+)", text)
    if m:
        lines = [ln.strip() for ln in m.group("body").splitlines()]
        return " ".join(ln for ln in lines if ln)
    m2 = re.search(rf"(?m)^\s*{re.escape(key)}\s*:\s*(.+)$", text)
    if m2:
        return m2.group(1).strip().strip('"').strip("'")
    return default


def load_prompt(path: Path | None) -> dict:
    cfg = {
        "raw": "",
        "tema": DEFAULT_TEMA,
        "max_referencias": 27,
        "top_n_candidatos": 90,
        "limiar_minimo_inclusao": 45,
        "pesos": dict(DEFAULT_WEIGHTS),
        "minimos_por_categoria": dict(DEFAULT_MINIMUMS),
    }
    if not path:
        return cfg
    if not path.exists():
        log("WARN", f"Prompt de curadoria não encontrado: {path}. Usando defaults.")
        return cfg
    text = path.read_text(encoding="utf-8", errors="ignore")
    cfg["raw"] = text
    cfg["tema"] = parse_block_scalar(text, "tema", cfg["tema"])
    cfg["max_referencias"] = parse_number(text, "max_referencias", cfg["max_referencias"])
    cfg["top_n_candidatos"] = parse_number(text, "top_n_candidatos", cfg["top_n_candidatos"])
    cfg["limiar_minimo_inclusao"] = parse_number(text, "limiar_minimo_inclusao", cfg["limiar_minimo_inclusao"])
    cfg["pesos"] = parse_section_numbers(text, "pesos", cfg["pesos"])
    cfg["minimos_por_categoria"] = parse_section_numbers(text, "minimos_por_categoria", cfg["minimos_por_categoria"])
    return cfg


def row_text(row: dict[str, str]) -> str:
    fields = [
        "titulo", "title", "resumo", "abstract", "periodico", "journal", "palavras_chave",
        "bloco_tematico_ia", "recomendacao_ia", "autores",
    ]
    return norm_key(" ".join(str(row.get(f, "")) for f in fields))


def category_score(row: dict[str, str], category: str) -> int:
    text = row_text(row)
    hits = 0
    strong_hits = 0
    for term in TERMS.get(category, []):
        if norm_key(term) in text:
            hits += 1
            if len(term) >= 10:
                strong_hits += 1
    score = min(100, hits * 18 + strong_hits * 10)
    if category in ("ia_automacao", "telepericia_avaliacao_remota") and score > 0:
        if not any(t in text for t in ["disability", "incapacity", "work capacity", "benefit", "certificate", "certification", "pericia", "perícia", "incapacidade", "deficiencia", "deficiência", "functional"]):
            score = max(0, score - 35)
    return score


def compute_scores(row: dict[str, str], prompt_cfg: dict) -> dict[str, str]:
    weights = prompt_cfg["pesos"]
    scores = {k: category_score(row, k) for k in DEFAULT_WEIGHTS}
    weighted = 0.0
    weight_total = sum(max(0, int(v)) for v in weights.values()) or 100
    for k, score in scores.items():
        weighted += score * max(0, int(weights.get(k, 0))) / weight_total

    prior = norm_key(row.get("recomendacao_ia", ""))
    if "prioridade_alta" in prior:
        weighted += 8
    elif "revisar_humano" in prior:
        weighted += 3
    elif "incerto" in prior:
        weighted -= 3

    # Pequeno bônus para recorte previdenciário, sem deixar dominar a seleção.
    if scores["avaliacao_medico_pericial"] >= 35:
        weighted += 4

    weighted = max(0, min(100, int(round(weighted))))
    dominant = max(scores, key=lambda k: scores[k])
    classe = class_from_scores(scores, dominant)
    eixo = CATEGORY_LABELS.get(dominant, "Curadoria geral")

    return {
        "aderencia_ia_automacao": str(scores["ia_automacao"]),
        "aderencia_telepericia": str(scores["telepericia_avaliacao_remota"]),
        "aderencia_decisao_evidencias": str(scores["decisao_baseada_em_evidencias"]),
        "aderencia_auditoria_qualidade": str(scores["auditoria_qualidade_integridade"]),
        "aderencia_fila_capacidade": str(scores["fila_capacidade_equidade"]),
        "aderencia_avaliacao_pericial": str(scores["avaliacao_medico_pericial"]),
        "escore_curadoria": str(weighted),
        "categoria_dominante": dominant,
        "eixo_seminario": eixo,
        "classe_curadoria": classe,
    }


def class_from_scores(scores: dict[str, int], dominant: str) -> str:
    if scores["ia_automacao"] >= 45 or scores["telepericia_avaliacao_remota"] >= 50:
        return "NUCLEO_TECNOLOGICO"
    if scores["decisao_baseada_em_evidencias"] >= 50 or scores["auditoria_qualidade_integridade"] >= 50:
        return "NUCLEO_DECISORIO"
    if dominant == "avaliacao_medico_pericial" and scores[dominant] >= 45:
        return "APOIO_PERICIAL"
    if scores["fila_capacidade_equidade"] >= 45:
        return "APOIO_CONTEXTO"
    return "EXCLUIR"


def infer_design(row: dict[str, str]) -> str:
    text = row_text(row)
    if "scoping review" in text:
        return "Revisão de escopo"
    if "systematic review" in text:
        return "Revisão sistemática"
    if any(x in text for x in ["qualitative", "interviews", "ethnographic", "text analysis"]):
        return "Estudo qualitativo/análise documental"
    if any(x in text for x in ["machine learning", "artificial intelligence", "algorithm", "prediction", "classification"]):
        return "Estudo/modelo computacional ou suporte decisório"
    if any(x in text for x in ["pilot", "evaluation of a pilot"]):
        return "Avaliação de estudo-piloto"
    if any(x in text for x in ["instrument", "scale", "validation", "reliability"]):
        return "Desenvolvimento/validação de instrumento"
    if any(x in text for x in ["proposal", "protocol", "propose"]):
        return "Proposta de protocolo/modelo"
    if "review" in text:
        return "Revisão narrativa/documental"
    return "Estudo aplicado/documental"


def infer_country(row: dict[str, str]) -> str:
    text = row_text(row)
    checks = [
        ("Brasil/INSS", ["brazil", "brazilian", "brasil", "inss", "bpc"]),
        ("Itália/INPS", ["italy", "italian", "inps"]),
        ("Coreia do Sul", ["korea", "korean"]),
        ("Reino Unido", ["incapacity benefit", "britain", "uk", "england"]),
        ("Países Baixos", ["netherlands", "dutch"]),
        ("Estados Unidos", ["social security disability", "ssa", "ssi", "united states"]),
        ("Portugal", ["portugal", "aveiro"]),
        ("Internacional/comparativo", ["international", "oecd", "countries", "worldwide"]),
    ]
    for label, terms in checks:
        if any(t in text for t in terms):
            return label
    return "Internacional ou não especificado"


def abnt_ref(row: dict[str, str]) -> str:
    autores = clean(row.get("autores", ""))
    titulo = clean(row.get("titulo", "") or row.get("title", ""))
    periodico = clean(row.get("periodico", ""))
    ano = clean(row.get("ano", ""))
    doi = clean(row.get("doi", ""))
    url = clean(row.get("url", ""))
    parts = []
    if autores:
        parts.append(autores.upper())
    if titulo:
        parts.append(f"{titulo}.")
    if periodico:
        parts.append(f"{periodico},")
    if ano:
        parts.append(f"{ano}.")
    if doi:
        parts.append(f"DOI: {doi}.")
    elif url:
        parts.append(f"Disponível em: {url}.")
    return " ".join(parts)


def default_matrix_fields(row: dict[str, str]) -> dict[str, str]:
    return {
        "pais_contexto": row.get("pais_contexto") or infer_country(row),
        "objetivo_estudo": row.get("objetivo_estudo") or f"Analisar contribuição do estudo para {row.get('eixo_seminario', 'o recorte do seminário')}.",
        "desenho_metodo": row.get("desenho_metodo") or infer_design(row),
        "amostra_base": row.get("amostra_base") or "Não detalhado automaticamente; conferir texto completo.",
        "achados_principais": row.get("achados_principais") or clean(row.get("resumo", ""), 420),
        "limitacoes": row.get("limitacoes") or "Limitações não detalhadas automaticamente; confirmar no texto completo.",
        "contribuicao_pergunta": row.get("contribuicao_pergunta") or "Contribui para discutir IA, automação, teleperícia, decisão baseada em evidências, qualidade decisória ou gestão de fluxo.",
        "como_usar_no_seminario": row.get("como_usar_no_seminario") or f"Usar no eixo: {row.get('eixo_seminario', 'curadoria geral')}.",
        "referencia_abnt_rascunho": row.get("referencia_abnt_rascunho") or abnt_ref(row),
    }


def title_for_ai(row: dict[str, str]) -> str:
    return clean(row.get("titulo") or row.get("title") or "", 260)


def call_ai_batch(batch: list[dict[str, str]], prompt_cfg: dict, model: str, api_url: str, api_key: str) -> list[dict[str, str]]:
    items = []
    for r in batch:
        items.append({
            "row_id": r["_row_id"],
            "titulo": title_for_ai(r),
            "autores": clean(r.get("autores", ""), 180),
            "ano": clean(r.get("ano", ""), 20),
            "periodico": clean(r.get("periodico", ""), 140),
            "doi": clean(r.get("doi", ""), 100),
            "resumo": clean(r.get("resumo", ""), 1400),
            "escores_heuristicos": {
                "ia_automacao": r.get("aderencia_ia_automacao"),
                "telepericia": r.get("aderencia_telepericia"),
                "decisao_evidencias": r.get("aderencia_decisao_evidencias"),
                "auditoria_qualidade": r.get("aderencia_auditoria_qualidade"),
                "fila_capacidade": r.get("aderencia_fila_capacidade"),
                "avaliacao_pericial": r.get("aderencia_avaliacao_pericial"),
                "total": r.get("escore_curadoria"),
            }
        })

    system = (
        "Você é um pesquisador sênior em revisão PRISMA, evidências em políticas públicas, "
        "perícia médica, IA aplicada a decisões administrativas e avaliação remota. "
        "Responda somente JSON válido, sem markdown."
    )
    user = {
        "tarefa": "Curadoria calibrada de referências para seminário e matriz PRISMA.",
        "prompt_curadoria_estruturado": prompt_cfg.get("raw", ""),
        "tema_interpretado": prompt_cfg.get("tema", DEFAULT_TEMA),
        "instrucoes": [
            "Priorize IA, automação, teleperícia, decisão baseada em evidências, auditabilidade e suporte decisório.",
            "Não escolha artigos apenas por serem sobre benefício por incapacidade se não contribuírem ao recorte tecnológico/decisório.",
            "Preencha campos analíticos de forma sintética e útil para matriz PRISMA.",
            "Use NUCLEO_TECNOLOGICO, NUCLEO_DECISORIO, APOIO_PERICIAL, APOIO_CONTEXTO ou EXCLUIR.",
        ],
        "saida": {
            "tipo": "array",
            "campos": [
                "row_id", "decisao_titulo_resumo", "motivo_exclusao_titulo_resumo",
                "decisao_texto_completo", "motivo_exclusao_texto_completo", "incluir_final",
                "prioridade_seminario", "classe_curadoria", "eixo_seminario",
                "aderencia_ia_automacao", "aderencia_telepericia",
                "aderencia_decisao_evidencias", "aderencia_auditoria_qualidade",
                "aderencia_fila_capacidade", "justificativa_priorizacao",
                "pais_contexto", "objetivo_estudo", "desenho_metodo", "amostra_base",
                "achados_principais", "limitacoes", "contribuicao_pergunta", "como_usar_no_seminario"
            ],
            "valores_validos": {
                "decisao_titulo_resumo": ["INCLUIR", "EXCLUIR", "REVISAR"],
                "decisao_texto_completo": ["INCLUIR", "EXCLUIR", "PENDENTE", "NAO_APLICAVEL"],
                "incluir_final": ["SIM", "NAO"],
                "prioridade_seminario": ["NÚCLEO", "APOIO", "EXCLUÍDA"],
                "classe_curadoria": ["NUCLEO_TECNOLOGICO", "NUCLEO_DECISORIO", "APOIO_PERICIAL", "APOIO_CONTEXTO", "EXCLUIR"],
            }
        },
        "referencias": items,
    }

    payload = json.dumps({
        "model": model,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": json.dumps(user, ensure_ascii=False)},
        ],
        "temperature": 0.1,
    }, ensure_ascii=False).encode("utf-8")

    req = urllib.request.Request(
        api_url,
        data=payload,
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=180) as resp:
        raw = resp.read().decode("utf-8", errors="replace")
    obj = json.loads(raw)
    content = obj["choices"][0]["message"]["content"].strip()
    if content.startswith("```"):
        content = re.sub(r"^```(?:json)?", "", content).strip()
        content = re.sub(r"```$", "", content).strip()
    try:
        parsed = json.loads(content)
    except Exception:
        m = re.search(r"(\[.*\]|\{.*\})", content, re.S)
        if not m:
            raise ValueError("Resposta da IA não contém JSON.")
        parsed = json.loads(m.group(1))
    if isinstance(parsed, dict) and "items" in parsed:
        parsed = parsed["items"]
    if not isinstance(parsed, list):
        raise ValueError("Resposta JSON da IA não é uma lista.")
    return parsed


def chunked(items, size):
    for i in range(0, len(items), size):
        yield items[i:i+size]


def apply_ai(rows: list[dict[str, str]], args: argparse.Namespace, prompt_cfg: dict) -> tuple[dict[str, dict[str, str]], list[dict]]:
    if not args.usar_ia:
        return {}, []
    load_dotenv()
    api_key = os.environ.get("OPENAI_API_KEY", "")
    if not api_key:
        log("WARN", "OPENAI_API_KEY ausente. Seguirei apenas com heurística.")
        return {}, []
    model = args.model or os.environ.get("OPENAI_MODEL") or "gpt-4.1-mini"
    api_url = args.api_url or os.environ.get("OPENAI_API_BASE") or "https://api.openai.com/v1/chat/completions"

    candidates = sorted(rows, key=lambda r: int(r.get("escore_curadoria") or 0), reverse=True)[: args.top_n_candidatos]
    log("INFO", f"IA analisará {len(candidates)} candidato(s), em lotes de {args.batch_size}.")

    decisions = {}
    log_items = []
    for batch_no, batch in enumerate(chunked(candidates, args.batch_size), start=1):
        ok = False
        for attempt in range(1, args.max_retries + 1):
            try:
                result = call_ai_batch(batch, prompt_cfg, model, api_url, api_key)
                for item in result:
                    rid = str(item.get("row_id", "")).strip()
                    if rid:
                        decisions[rid] = {str(k): clean(v) for k, v in item.items()}
                log("OK", f"Lote IA {batch_no} concluído: {len(result)} item(ns).")
                log_items.append({"batch": batch_no, "attempt": attempt, "status": "ok", "items": len(result)})
                ok = True
                break
            except Exception as exc:
                wait = min(60, 3 * attempt)
                log("WARN", f"Falha no lote IA {batch_no}, tentativa {attempt}: {exc}. Aguardando {wait}s.")
                log_items.append({"batch": batch_no, "attempt": attempt, "status": "erro", "erro": str(exc)})
                time.sleep(wait)
        if not ok:
            log("WARN", f"Lote IA {batch_no} não foi concluído. Heurística preservada para esses itens.")
    return decisions, log_items


def fill_headers(headers: list[str]) -> list[str]:
    curated = [
        "ordem_curadoria", "prioridade_seminario", "classe_curadoria", "eixo_seminario",
        "aderencia_ia_automacao", "aderencia_telepericia", "aderencia_decisao_evidencias",
        "aderencia_auditoria_qualidade", "aderencia_fila_capacidade", "aderencia_avaliacao_pericial",
        "escore_curadoria", "categoria_dominante", "justificativa_priorizacao",
        "decisao_titulo_resumo", "motivo_exclusao_titulo_resumo",
        "texto_completo_local", "decisao_texto_completo", "motivo_exclusao_texto_completo",
        "incluir_final", "observacoes", "pais_contexto", "objetivo_estudo", "desenho_metodo",
        "amostra_base", "achados_principais", "limitacoes", "contribuicao_pergunta",
        "como_usar_no_seminario", "referencia_abnt_rascunho",
    ]
    out = list(headers)
    for h in curated:
        if h not in out:
            out.append(h)
    return out


def manual_reexport(headers: list[str], rows: list[dict[str, str]]) -> tuple[list[str], list[dict[str, str]]]:
    headers = fill_headers(headers)
    out = []
    order = 1
    for row in rows:
        r = dict(row)
        if str(r.get("incluir_final", "")).strip().upper() == "SIM":
            if not r.get("ordem_curadoria"):
                r["ordem_curadoria"] = str(order)
            order += 1
            r["decisao_titulo_resumo"] = r.get("decisao_titulo_resumo") or "INCLUIR"
            r["decisao_texto_completo"] = r.get("decisao_texto_completo") or "INCLUIR"
            r["prioridade_seminario"] = r.get("prioridade_seminario") or "APOIO"
            r["classe_curadoria"] = r.get("classe_curadoria") or "APOIO_PERICIAL"
        else:
            r["incluir_final"] = "NAO"
            r["decisao_titulo_resumo"] = r.get("decisao_titulo_resumo") or "EXCLUIR"
            r["decisao_texto_completo"] = r.get("decisao_texto_completo") or "NAO_APLICAVEL"
            r["prioridade_seminario"] = r.get("prioridade_seminario") or "EXCLUÍDA"
            r["classe_curadoria"] = r.get("classe_curadoria") or "EXCLUIR"
        for k, v in default_matrix_fields(r).items():
            r.setdefault(k, v)
            if not r.get(k):
                r[k] = v
        out.append(r)
    return headers, out


def select_final(rows: list[dict[str, str]], prompt_cfg: dict, max_incluir: int, threshold: int) -> set[str]:
    selected: set[str] = set()
    minima = prompt_cfg["minimos_por_categoria"]

    def eligible(r):
        return int(r.get("escore_curadoria") or 0) >= threshold

    # 1. Cumpre mínimos por categoria com os melhores candidatos.
    for category, minimum in minima.items():
        col = {
            "ia_automacao": "aderencia_ia_automacao",
            "telepericia_avaliacao_remota": "aderencia_telepericia",
            "decisao_baseada_em_evidencias": "aderencia_decisao_evidencias",
            "auditoria_qualidade_integridade": "aderencia_auditoria_qualidade",
            "fila_capacidade_equidade": "aderencia_fila_capacidade",
        }.get(category)
        if not col:
            continue
        candidates = sorted(
            [r for r in rows if eligible(r) and int(r.get(col) or 0) >= 35],
            key=lambda r: (int(r.get(col) or 0), int(r.get("escore_curadoria") or 0)),
            reverse=True,
        )
        count = 0
        for r in candidates:
            if len(selected) >= max_incluir:
                break
            if r["_row_id"] not in selected:
                selected.add(r["_row_id"])
                count += 1
            if count >= int(minimum):
                break

    # 2. Completa por escore total.
    for r in sorted(rows, key=lambda x: int(x.get("escore_curadoria") or 0), reverse=True):
        if len(selected) >= max_incluir:
            break
        if eligible(r):
            selected.add(r["_row_id"])

    return selected


def curate_rows(headers: list[str], rows: list[dict[str, str]], args: argparse.Namespace, prompt_cfg: dict) -> tuple[list[str], list[dict[str, str]], list[dict]]:
    headers = fill_headers(headers)

    for i, row in enumerate(rows, start=1):
        row["_row_id"] = str(i)
        scores = compute_scores(row, prompt_cfg)
        row.update(scores)
        row.update(default_matrix_fields(row))
        row["decisao_titulo_resumo"] = "EXCLUIR"
        row["decisao_texto_completo"] = "NAO_APLICAVEL"
        row["incluir_final"] = "NAO"
        row["prioridade_seminario"] = "EXCLUÍDA"
        row["motivo_exclusao_titulo_resumo"] = "Baixa aderência relativa ao prompt calibrado ou fora do recorte tecnológico/decisório."
        row["motivo_exclusao_texto_completo"] = "Excluído na triagem por título/resumo."
        row["observacoes"] = "Curadoria heurística local v2; revisar XLSX antes da submissão formal."
        row["justificativa_priorizacao"] = (
            f"Escore {row['escore_curadoria']}; classe {row['classe_curadoria']}; "
            f"categoria dominante {row['categoria_dominante']}."
        )

    ai_decisions, ai_log = apply_ai(rows, args, prompt_cfg)
    for row in rows:
        ai = ai_decisions.get(row["_row_id"])
        if not ai:
            continue
        for k, v in ai.items():
            if k in headers and v != "":
                row[k] = v
        if "aderencia_ia_automacao" in ai:
            row["aderencia_ia_automacao"] = ai.get("aderencia_ia_automacao", row["aderencia_ia_automacao"])
        row["observacoes"] = "Curadoria v2 gerada por IA conforme prompt estruturado; revisar XLSX antes da submissão formal."

    selected = select_final(rows, prompt_cfg, args.max_incluir, args.limiar_minimo_inclusao)

    order = 1
    for row in sorted(rows, key=lambda r: int(r["_row_id"])):
        if row["_row_id"] in selected:
            row["ordem_curadoria"] = str(order)
            row["incluir_final"] = "SIM"
            row["decisao_titulo_resumo"] = "INCLUIR"
            row["decisao_texto_completo"] = row.get("decisao_texto_completo") if row.get("decisao_texto_completo") not in ("", "NAO_APLICAVEL") else "INCLUIR"
            row["motivo_exclusao_titulo_resumo"] = ""
            row["motivo_exclusao_texto_completo"] = ""
            if row.get("classe_curadoria") in ("NUCLEO_TECNOLOGICO", "NUCLEO_DECISORIO"):
                row["prioridade_seminario"] = "NÚCLEO"
            else:
                row["prioridade_seminario"] = "APOIO"
            order += 1
        else:
            row["ordem_curadoria"] = ""
            row["incluir_final"] = "NAO"
            row["decisao_titulo_resumo"] = "EXCLUIR"
            row["decisao_texto_completo"] = "NAO_APLICAVEL"
            row["prioridade_seminario"] = "EXCLUÍDA"
            if not row.get("motivo_exclusao_titulo_resumo"):
                row["motivo_exclusao_titulo_resumo"] = "Excluído por curadoria final ou menor aderência relativa ao prompt calibrado."
            row["motivo_exclusao_texto_completo"] = "Excluído na triagem por título/resumo."

    return headers, rows, ai_log


def write_csv(path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    visible_headers = [h for h in headers if not h.startswith("_")]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=visible_headers, extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            writer.writerow({h: r.get(h, "") for h in visible_headers})


def write_xlsx(path: Path, headers: list[str], rows: list[dict[str, str]], prompt_cfg: dict) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.chart import BarChart, Reference
    except Exception as exc:
        log("WARN", f"openpyxl não disponível; XLSX não será gerado. Motivo: {exc}")
        return

    visible_headers = [h for h in headers if not h.startswith("_")]
    included = [r for r in rows if str(r.get("incluir_final", "")).upper() == "SIM"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Curadoria IA v2 - Referências PRISMA"])
    ws.append(["Tema", prompt_cfg.get("tema", DEFAULT_TEMA)])
    ws.append([])
    ws.append(["Indicador", "Valor"])
    ws.append(["Registros analisados", len(rows)])
    ws.append(["Referências incluídas", len(included)])
    ws.append(["Referências excluídas", len(rows) - len(included)])
    ws.append(["Max referências", prompt_cfg.get("max_referencias")])
    ws.append(["Prompt calibrado", "Sim" if prompt_cfg.get("raw") else "Não"])
    ws.append([])
    ws.append(["Classe/eixo", "Incluídas"])

    counts = {}
    for r in included:
        k = r.get("classe_curadoria") or r.get("eixo_seminario") or "Sem classe"
        counts[k] = counts.get(k, 0) + 1
    for k, v in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])):
        ws.append([k, v])

    def style_header(row):
        fill = PatternFill("solid", fgColor="1F4E79")
        for cell in row:
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    style_header(ws[1])
    style_header(ws[4])
    style_header(ws[11])

    if counts:
        chart = BarChart()
        chart.title = "Incluídas por classe"
        data = Reference(ws, min_col=2, min_row=11, max_row=11+len(counts))
        cats = Reference(ws, min_col=1, min_row=12, max_row=11+len(counts))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "D11")

    inc_headers = [
        "ordem_curadoria", "prioridade_seminario", "classe_curadoria", "eixo_seminario",
        "titulo", "autores", "ano", "periodico", "doi", "url", "url_pdf_aberto",
        "aderencia_ia_automacao", "aderencia_telepericia", "aderencia_decisao_evidencias",
        "aderencia_auditoria_qualidade", "aderencia_fila_capacidade", "aderencia_avaliacao_pericial",
        "escore_curadoria", "decisao_titulo_resumo", "decisao_texto_completo", "incluir_final",
        "justificativa_priorizacao", "pais_contexto", "objetivo_estudo", "desenho_metodo",
        "amostra_base", "achados_principais", "limitacoes", "contribuicao_pergunta",
        "como_usar_no_seminario", "referencia_abnt_rascunho"
    ]
    ws_inc = wb.create_sheet("Referências incluídas")
    ws_inc.append(inc_headers)
    for r in sorted(included, key=lambda x: int(x.get("ordem_curadoria") or 999999)):
        ws_inc.append([r.get(h, "") for h in inc_headers])

    ws_tri = wb.create_sheet("Triagem completa")
    ws_tri.append(visible_headers)
    for r in rows:
        ws_tri.append([r.get(h, "") for h in visible_headers])

    ws_prompt = wb.create_sheet("Prompt curadoria")
    ws_prompt.append(["Campo", "Valor"])
    for key in ["tema", "max_referencias", "top_n_candidatos", "limiar_minimo_inclusao"]:
        ws_prompt.append([key, str(prompt_cfg.get(key, ""))])
    ws_prompt.append([])
    ws_prompt.append(["pesos", json.dumps(prompt_cfg.get("pesos", {}), ensure_ascii=False)])
    ws_prompt.append(["minimos_por_categoria", json.dumps(prompt_cfg.get("minimos_por_categoria", {}), ensure_ascii=False)])
    ws_prompt.append([])
    ws_prompt.append(["prompt_raw", prompt_cfg.get("raw", "")])

    ws_abnt = wb.create_sheet("Referências ABNT")
    ws_abnt.append(["Ordem", "Classe", "Referência ABNT preliminar", "DOI", "URL"])
    for r in sorted(included, key=lambda x: int(x.get("ordem_curadoria") or 999999)):
        ws_abnt.append([r.get("ordem_curadoria", ""), r.get("classe_curadoria", ""), r.get("referencia_abnt_rascunho", ""), r.get("doi", ""), r.get("url", "")])

    for sheet in wb.worksheets:
        if sheet.max_row >= 1:
            style_header(sheet[1])
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(1, min(sheet.max_column, 40) + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in sheet[letter][: min(sheet.max_row, 80)]:
                max_len = max(max_len, len(str(cell.value or "")))
            sheet.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)

    for ws_table, name in [(ws_inc, "TabelaReferenciasIncluidas"), (ws_tri, "TabelaTriagemCompleta")]:
        if ws_table.max_row > 1 and ws_table.max_column > 1:
            tab = Table(displayName=name, ref=f"A1:{get_column_letter(ws_table.max_column)}{ws_table.max_row}")
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
            ws_table.add_table(tab)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_summary(path: Path, rows: list[dict[str, str]], source: str, prompt_cfg: dict, usar_ia: bool) -> None:
    included = [r for r in rows if str(r.get("incluir_final", "")).upper() == "SIM"]
    counts = {}
    for r in included:
        k = r.get("classe_curadoria") or "Sem classe"
        counts[k] = counts.get(k, 0) + 1

    lines = [
        "Resumo da curadoria IA v2 PRISMA",
        "=" * 48,
        f"Entrada: {source}",
        f"IA utilizada: {'sim' if usar_ia else 'não'}",
        f"Tema: {prompt_cfg.get('tema')}",
        f"Total de registros: {len(rows)}",
        f"Incluídos finais: {len(included)}",
        f"Excluídos finais: {len(rows)-len(included)}",
        "",
        "Incluídos por classe:",
    ]
    for k, v in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])):
        lines.append(f"- {k}: {v}")
    lines.append("")
    lines.append("Pesos:")
    for k, v in prompt_cfg.get("pesos", {}).items():
        lines.append(f"- {k}: {v}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default="app_bundle/projetos/prisma_fluxo_pmf/prisma_fluxo_pmf.toml")
    parser.add_argument("--input", default="")
    parser.add_argument("--out-dir", default=str(DEFAULT_OUT))
    parser.add_argument("--prompt-curadoria", default="")
    parser.add_argument("--usar-ia", action="store_true")
    parser.add_argument("--reexportar-xlsx", action="store_true", help="Preserva decisões já editadas no XLSX e apenas regenera CSVs.")
    parser.add_argument("--max-incluir", type=int, default=0)
    parser.add_argument("--top-n-candidatos", type=int, default=0)
    parser.add_argument("--limiar-minimo-inclusao", type=int, default=0)
    parser.add_argument("--batch-size", type=int, default=6)
    parser.add_argument("--max-retries", type=int, default=3)
    parser.add_argument("--model", default="")
    parser.add_argument("--api-url", default="")
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    input_path = Path(args.input) if args.input else find_default_input(out_dir)
    prompt_path = Path(args.prompt_curadoria) if args.prompt_curadoria else None
    prompt_cfg = load_prompt(prompt_path)

    if args.max_incluir > 0:
        prompt_cfg["max_referencias"] = args.max_incluir
    if args.top_n_candidatos > 0:
        prompt_cfg["top_n_candidatos"] = args.top_n_candidatos
    if args.limiar_minimo_inclusao > 0:
        prompt_cfg["limiar_minimo_inclusao"] = args.limiar_minimo_inclusao

    args.max_incluir = int(prompt_cfg.get("max_referencias") or 27)
    args.top_n_candidatos = int(prompt_cfg.get("top_n_candidatos") or 90)
    args.limiar_minimo_inclusao = int(prompt_cfg.get("limiar_minimo_inclusao") or 45)

    headers, rows, source = load_input(input_path)
    if not rows:
        log("ERRO", f"Nenhum registro lido de {input_path}")
        return 1

    if args.reexportar_xlsx:
        headers, final_rows = manual_reexport(headers, rows)
        ai_log = []
    else:
        headers, final_rows, ai_log = curate_rows(headers, rows, args, prompt_cfg)

    out_dir.mkdir(parents=True, exist_ok=True)
    triage_csv = out_dir / f"{PREFIX}.triagem_humana.csv"
    included_csv = out_dir / f"{PREFIX}.referencias_incluidas_seminario.csv"
    xlsx_out = out_dir / f"{PREFIX}.curadoria_ia_referencias.xlsx"
    summary = out_dir / f"{PREFIX}.curadoria_ia_resumo.txt"
    log_json = out_dir / f"{PREFIX}.curadoria_ia_log.json"

    write_csv(triage_csv, headers, final_rows)
    included = [r for r in final_rows if str(r.get("incluir_final", "")).upper() == "SIM"]
    write_csv(included_csv, headers, included)
    write_xlsx(xlsx_out, headers, final_rows, prompt_cfg)
    write_summary(summary, final_rows, source, prompt_cfg, args.usar_ia)
    log_json.write_text(json.dumps({
        "created_at": dt.datetime.now().isoformat(),
        "source": source,
        "prompt_curadoria": str(prompt_path or ""),
        "usar_ia": bool(args.usar_ia),
        "reexportar_xlsx": bool(args.reexportar_xlsx),
        "max_incluir": args.max_incluir,
        "top_n_candidatos": args.top_n_candidatos,
        "limiar_minimo_inclusao": args.limiar_minimo_inclusao,
        "registros": len(final_rows),
        "incluidos": len(included),
        "ai_log": ai_log,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    log("OK", f"XLSX de curadoria: {xlsx_out}")
    log("OK", f"CSV para pipeline: {triage_csv}")
    log("OK", f"CSV de incluídas: {included_csv}")
    log("OK", f"Resumo: {summary}")
    log("INFO", f"Registros: {len(final_rows)} | Incluídos: {len(included)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
